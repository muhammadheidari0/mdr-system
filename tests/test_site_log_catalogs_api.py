from __future__ import annotations

from io import BytesIO
from datetime import datetime
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect

from app.db.models import SiteLog
from app.db.session import SessionLocal
from app.main import app
from tests.site_logs_helpers import (
    admin_headers,
    create_scoped_user_and_login,
    ensure_project_discipline,
)


client = TestClient(app)


def _assert_catalog_schema_ready() -> None:
    with SessionLocal() as db:
        inspector = inspect(db.bind)
        missing = [
            table_name
            for table_name in (
                "site_log_role_catalog",
                "site_log_work_section_catalog",
                "site_log_equipment_catalog",
                "site_log_material_catalog",
                "site_log_equipment_status_catalog",
                "site_log_attachment_type_catalog",
                "site_log_issue_type_catalog",
                "site_log_shift_catalog",
                "site_log_weather_catalog",
                "site_log_activity_catalog",
                "site_log_pms_templates",
                "site_log_pms_template_steps",
                "site_log_activity_pms_mappings",
                "site_log_activity_pms_steps",
            )
            if not inspector.has_table(table_name)
        ]
    assert not missing, (
        "Site-log catalog tables are missing in the active database. "
        "Apply site-log catalog migrations through 20260508_0045 before running this test. "
        f"Missing: {', '.join(missing)}"
    )


def _upsert_catalog_item(
    headers: dict[str, str],
    *,
    catalog_type: str,
    code: str,
    label: str,
    sort_order: int = 10,
    is_active: bool = True,
    item_id: int | None = None,
) -> dict[str, object]:
    payload: dict[str, object] = {
        "catalog_type": catalog_type,
        "code": code,
        "label": label,
        "sort_order": sort_order,
        "is_active": is_active,
    }
    if item_id:
        payload["id"] = item_id
    response = client.post(
        "/api/v1/settings/site-log-catalogs/upsert",
        json=payload,
        headers=headers,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body.get("ok") is True
    return body


def _upsert_activity_catalog_item(
    headers: dict[str, str],
    *,
    project_code: str,
    activity_code: str,
    activity_title: str,
    organization_id: int | None = None,
    organization_contract_id: int | None = None,
    activity_type: str | None = None,
    activity_type_code: str | None = None,
    floor: str | None = None,
    wbs_code: str | None = None,
    default_quantity: float | None = None,
    default_location: str | None = None,
    default_unit: str | None = None,
    sort_order: int = 10,
    is_active: bool = True,
    item_id: int | None = None,
) -> dict[str, object]:
    payload: dict[str, object] = {
        "project_code": project_code,
        "organization_id": organization_id,
        "organization_contract_id": organization_contract_id,
        "activity_code": activity_code,
        "activity_title": activity_title,
        "activity_type": activity_type,
        "activity_type_code": activity_type_code,
        "floor": floor,
        "wbs_code": wbs_code,
        "default_quantity": default_quantity,
        "default_location": default_location,
        "default_unit": default_unit,
        "sort_order": sort_order,
        "is_active": is_active,
    }
    if item_id:
        payload["id"] = item_id
    response = client.post(
        "/api/v1/settings/site-log-activity-catalog/upsert",
        json=payload,
        headers=headers,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body.get("ok") is True
    return body


def _create_legacy_draft(
    headers: dict[str, str],
    *,
    project_code: str,
    discipline_code: str,
    organization_id: int,
    role_label: str,
    equipment_label: str,
    claimed_status: str,
    work_section_label: str | None = None,
) -> int:
    payload = {
        "log_type": "DAILY",
        "project_code": project_code,
        "discipline_code": discipline_code,
        "organization_id": int(organization_id),
        "log_date": datetime.utcnow().strftime("%Y-%m-%dT00:00:00"),
        "weather": "CLEAR",
        "summary": f"Legacy site log {uuid4().hex[:8]}",
        "manpower_rows": [
            {
                "role_code": None,
                "role_label": role_label,
                "work_section_label": work_section_label,
                "claimed_count": 2,
                "claimed_hours": 8.0,
                "sort_order": 0,
            }
        ],
        "equipment_rows": [
            {
                "equipment_code": None,
                "equipment_label": equipment_label,
                "claimed_status": claimed_status,
                "claimed_hours": 6.5,
                "sort_order": 0,
            }
        ],
        "activity_rows": [],
    }
    response = client.post("/api/v1/site-logs/create", json=payload, headers=headers)
    assert response.status_code == 200, response.text
    body = response.json()
    assert body.get("ok") is True
    return int(body.get("data", {}).get("id") or 0)


def _ensure_block(headers: dict[str, str], project_code: str) -> int:
    block_code = f"BLK{uuid4().hex[:4].upper()}"
    upsert_res = client.post(
        "/api/v1/settings/blocks/upsert",
        json={
            "project_code": project_code,
            "code": block_code,
            "name_e": f"Block {block_code}",
            "name_p": f"Block {block_code}",
            "sort_order": 10,
            "is_active": True,
        },
        headers=headers,
    )
    assert upsert_res.status_code == 200, upsert_res.text
    list_res = client.get("/api/v1/settings/blocks", headers=headers)
    assert list_res.status_code == 200, list_res.text
    row = next(
        (
            item
            for item in list_res.json().get("items", [])
            if str(item.get("project_code") or "").strip().upper() == project_code
            and str(item.get("code") or "").strip().upper() == block_code
        ),
        None,
    )
    assert row is not None
    return int(row.get("id") or 0)


def test_site_log_catalogs_settings_crud_and_runtime_visibility() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)
    project_code, discipline_code = ensure_project_discipline(client, admin)

    viewer = create_scoped_user_and_login(
        client,
        admin,
        org_type="contractor",
        project_code=project_code,
        discipline_code=discipline_code,
        email_prefix=f"site_log_catalog_viewer_{uuid4().hex[:6]}",
        organization_role="viewer",
    )
    viewer_headers = viewer["headers"]  # type: ignore[assignment]

    settings_response = client.get("/api/v1/settings/site-log-catalogs", headers=admin)
    assert settings_response.status_code == 200, settings_response.text
    settings_body = settings_response.json()
    assert settings_body.get("ok") is True
    expected_catalogs = {"role", "work_section", "equipment", "material", "equipment_status", "attachment_type", "issue_type", "shift", "weather"}
    assert set((settings_body.get("catalogs") or {}).keys()) == expected_catalogs
    assert set((settings_body.get("catalog_titles") or {}).keys()) == expected_catalogs

    blocked_settings_read = client.get("/api/v1/settings/site-log-catalogs", headers=viewer_headers)  # type: ignore[arg-type]
    assert blocked_settings_read.status_code == 403, blocked_settings_read.text

    sample_payload = {
        "catalog_type": "role",
        "code": f"VIEW{uuid4().hex[:4].upper()}",
        "label": "Blocked write",
        "sort_order": 10,
        "is_active": True,
    }
    blocked_upsert = client.post(
        "/api/v1/settings/site-log-catalogs/upsert",
        json=sample_payload,
        headers=viewer_headers,  # type: ignore[arg-type]
    )
    assert blocked_upsert.status_code == 403, blocked_upsert.text

    created_items: dict[str, dict[str, object]] = {}
    for catalog_type, label_prefix in (
        ("role", "Role"),
        ("work_section", "Work section"),
        ("equipment", "Equipment"),
        ("material", "Material"),
        ("equipment_status", "Status"),
        ("attachment_type", "Attachment type"),
        ("issue_type", "Issue type"),
        ("shift", "Shift"),
        ("weather", "Weather"),
    ):
        code = f"{catalog_type[:3].upper()}{uuid4().hex[:6].upper()}"
        label = f"{label_prefix} {uuid4().hex[:6]}"
        body = _upsert_catalog_item(
            admin,
            catalog_type=catalog_type,
            code=code,
            label=label,
            sort_order=25,
        )
        item = body.get("item") or {}
        assert str(item.get("code") or "") == code
        assert str(item.get("label") or "") == label
        assert bool(item.get("is_active")) is True
        created_items[catalog_type] = dict(item)

    runtime_catalog = client.get("/api/v1/site-logs/catalog", headers=viewer_headers)  # type: ignore[arg-type]
    assert runtime_catalog.status_code == 200, runtime_catalog.text
    runtime_body = runtime_catalog.json()
    assert runtime_body.get("ok") is True

    runtime_role_codes = {str(item.get("code") or "") for item in (runtime_body.get("role_catalog") or [])}
    runtime_work_section_codes = {str(item.get("code") or "") for item in (runtime_body.get("work_section_catalog") or [])}
    runtime_equipment_codes = {str(item.get("code") or "") for item in (runtime_body.get("equipment_catalog") or [])}
    runtime_material_codes = {str(item.get("code") or "") for item in (runtime_body.get("material_catalog") or [])}
    runtime_status_codes = {
        str(item.get("code") or "")
        for item in (runtime_body.get("equipment_status_catalog") or [])
    }
    runtime_attachment_type_codes = {
        str(item.get("code") or "")
        for item in (runtime_body.get("attachment_type_catalog") or [])
    }
    runtime_issue_type_codes = {
        str(item.get("code") or "")
        for item in (runtime_body.get("issue_type_catalog") or [])
    }
    runtime_shift_codes = {str(item.get("code") or "") for item in (runtime_body.get("shift_catalog") or [])}
    runtime_weather_codes = {str(item.get("code") or "") for item in (runtime_body.get("weather_catalog") or [])}
    assert str(created_items["role"]["code"]) in runtime_role_codes
    assert str(created_items["work_section"]["code"]) in runtime_work_section_codes
    assert str(created_items["equipment"]["code"]) in runtime_equipment_codes
    assert str(created_items["material"]["code"]) in runtime_material_codes
    assert str(created_items["equipment_status"]["code"]) in runtime_status_codes
    assert str(created_items["attachment_type"]["code"]) in runtime_attachment_type_codes
    assert str(created_items["issue_type"]["code"]) in runtime_issue_type_codes
    assert str(created_items["shift"]["code"]) in runtime_shift_codes
    assert str(created_items["weather"]["code"]) in runtime_weather_codes

    updated_role_label = f"Role Updated {uuid4().hex[:6]}"
    updated_role = _upsert_catalog_item(
        admin,
        catalog_type="role",
        item_id=int(created_items["role"]["id"]),
        code=str(created_items["role"]["code"]),
        label=updated_role_label,
        sort_order=35,
    )
    assert str((updated_role.get("item") or {}).get("label") or "") == updated_role_label

    delete_response = client.post(
        "/api/v1/settings/site-log-catalogs/delete",
        json={
            "catalog_type": "equipment_status",
            "id": int(created_items["equipment_status"]["id"]),
        },
        headers=admin,
    )
    assert delete_response.status_code == 200, delete_response.text
    delete_body = delete_response.json()
    assert delete_body.get("ok") is True
    assert bool((delete_body.get("item") or {}).get("is_active")) is False

    blocked_delete = client.post(
        "/api/v1/settings/site-log-catalogs/delete",
        json={
            "catalog_type": "equipment_status",
            "id": int(created_items["equipment_status"]["id"]),
        },
        headers=viewer_headers,  # type: ignore[arg-type]
    )
    assert blocked_delete.status_code == 403, blocked_delete.text

    settings_after_delete = client.get("/api/v1/settings/site-log-catalogs", headers=admin)
    assert settings_after_delete.status_code == 200, settings_after_delete.text
    settings_after_delete_body = settings_after_delete.json()
    settings_status_items = settings_after_delete_body.get("catalogs", {}).get("equipment_status") or []
    deleted_row = next(
        (
            item
            for item in settings_status_items
            if int(item.get("id") or 0) == int(created_items["equipment_status"]["id"])
        ),
        None,
    )
    assert deleted_row is not None
    assert bool(deleted_row.get("is_active")) is False

    runtime_after_delete = client.get("/api/v1/site-logs/catalog", headers=viewer_headers)  # type: ignore[arg-type]
    assert runtime_after_delete.status_code == 200, runtime_after_delete.text
    runtime_after_delete_body = runtime_after_delete.json()
    runtime_status_codes_after = {
        str(item.get("code") or "")
        for item in (runtime_after_delete_body.get("equipment_status_catalog") or [])
    }
    assert str(created_items["equipment_status"]["code"]) not in runtime_status_codes_after


def test_site_log_material_and_equipment_catalog_bulk_upsert() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)

    material_code = f"MATB{uuid4().hex[:5].upper()}"
    material_res = client.post(
        "/api/v1/settings/site-log-catalogs/bulk-upsert",
        json={
            "catalog_type": "material",
            "items": [
                {"code": material_code, "label": "Bulk Cement", "sort_order": 10},
                {"label": "Bulk Rebar Without Code", "sort_order": 20},
                {"code": material_code, "label": "Duplicate Material", "sort_order": 30},
            ],
            "overwrite_existing": False,
        },
        headers=admin,
    )
    assert material_res.status_code == 200, material_res.text
    material_body = material_res.json()
    assert material_body.get("ok") is True
    assert material_body.get("created") == 2
    assert material_body.get("updated") == 0
    assert material_body.get("skipped") == 1
    material_items = material_body.get("catalogs", {}).get("material") or []
    assert material_code in {str(row.get("code") or "") for row in material_items}
    generated_material = next(
        (
            row
            for row in material_items
            if str(row.get("label") or "") == "Bulk Rebar Without Code"
        ),
        None,
    )
    assert generated_material is not None
    assert str(generated_material.get("code") or "").startswith("MAT")

    equipment_code = f"EQPB{uuid4().hex[:5].upper()}"
    create_equipment = client.post(
        "/api/v1/settings/site-log-catalogs/bulk-upsert",
        json={
            "catalog_type": "equipment",
            "items": [{"code": equipment_code, "label": "Bulk Loader", "sort_order": 10}],
        },
        headers=admin,
    )
    assert create_equipment.status_code == 200, create_equipment.text
    update_equipment = client.post(
        "/api/v1/settings/site-log-catalogs/bulk-upsert",
        json={
            "catalog_type": "equipment",
            "items": [{"code": equipment_code, "label": "Bulk Loader Updated", "sort_order": 25}],
            "overwrite_existing": True,
        },
        headers=admin,
    )
    assert update_equipment.status_code == 200, update_equipment.text
    update_body = update_equipment.json()
    assert update_body.get("created") == 0
    assert update_body.get("updated") == 1
    equipment_items = update_body.get("catalogs", {}).get("equipment") or []
    updated = next((row for row in equipment_items if row.get("code") == equipment_code), None)
    assert updated is not None
    assert updated.get("label") == "Bulk Loader Updated"
    assert updated.get("sort_order") == 25

    unsupported = client.post(
        "/api/v1/settings/site-log-catalogs/bulk-upsert",
        json={"catalog_type": "role", "items": [{"label": "Role Bulk"}]},
        headers=admin,
    )
    assert unsupported.status_code == 400, unsupported.text


def test_site_log_catalogs_preserve_legacy_free_text_rows() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)
    project_code, discipline_code = ensure_project_discipline(client, admin)

    contractor = create_scoped_user_and_login(
        client,
        admin,
        org_type="contractor",
        project_code=project_code,
        discipline_code=discipline_code,
        email_prefix=f"site_log_catalog_editor_{uuid4().hex[:6]}",
        organization_role="manager",
    )
    contractor_headers = contractor["headers"]  # type: ignore[assignment]

    legacy_role_label = f"Legacy Role {uuid4().hex[:6]}"
    legacy_work_section_label = f"Legacy Section {uuid4().hex[:6]}"
    legacy_equipment_label = f"Legacy Equipment {uuid4().hex[:6]}"
    legacy_status = f"legacy_status_{uuid4().hex[:4]}"

    draft_id = _create_legacy_draft(
        contractor_headers,  # type: ignore[arg-type]
        project_code=project_code,
        discipline_code=discipline_code,
        organization_id=int(contractor.get("organization_id") or 0),
        role_label=legacy_role_label,
        equipment_label=legacy_equipment_label,
        claimed_status=legacy_status,
        work_section_label=legacy_work_section_label,
    )
    assert draft_id > 0

    get_response = client.get(f"/api/v1/site-logs/{draft_id}", headers=contractor_headers)  # type: ignore[arg-type]
    assert get_response.status_code == 200, get_response.text
    get_body = get_response.json()
    assert get_body.get("ok") is True
    data = get_body.get("data") or {}

    manpower_rows = data.get("manpower_rows") or []
    equipment_rows = data.get("equipment_rows") or []
    assert manpower_rows
    assert equipment_rows
    assert str(manpower_rows[0].get("role_label") or "") == legacy_role_label
    assert str(manpower_rows[0].get("work_section_label") or "") == legacy_work_section_label
    assert str(equipment_rows[0].get("equipment_label") or "") == legacy_equipment_label
    assert str(equipment_rows[0].get("claimed_status") or "") == legacy_status.upper()

    update_response = client.put(
        f"/api/v1/site-logs/{draft_id}",
        json={
            "summary": f"Legacy updated {uuid4().hex[:6]}",
            "manpower_rows": manpower_rows,
            "equipment_rows": equipment_rows,
            "activity_rows": data.get("activity_rows") or [],
        },
        headers=contractor_headers,  # type: ignore[arg-type]
    )
    assert update_response.status_code == 200, update_response.text
    update_body = update_response.json()
    assert update_body.get("ok") is True
    updated_data = update_body.get("data") or {}
    updated_manpower = updated_data.get("manpower_rows") or []
    updated_equipment = updated_data.get("equipment_rows") or []
    assert updated_manpower
    assert updated_equipment
    assert str(updated_manpower[0].get("role_label") or "") == legacy_role_label
    assert str(updated_manpower[0].get("work_section_label") or "") == legacy_work_section_label
    assert str(updated_equipment[0].get("equipment_label") or "") == legacy_equipment_label
    assert str(updated_equipment[0].get("claimed_status") or "") == legacy_status.upper()


def test_site_log_activity_catalog_crud_and_runtime_fallback() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)
    _existing_project_code, discipline_code = ensure_project_discipline(client, admin)
    project_code = f"ACTP{uuid4().hex[:5].upper()}"
    project_upsert = client.post(
        "/api/v1/settings/projects/upsert",
        json={"code": project_code, "name_e": f"Activity Project {project_code}", "is_active": True},
        headers=admin,
    )
    assert project_upsert.status_code == 200, project_upsert.text
    block_id = _ensure_block(admin, project_code)

    viewer = create_scoped_user_and_login(
        client,
        admin,
        org_type="contractor",
        project_code=project_code,
        discipline_code=discipline_code,
        email_prefix=f"site_log_activity_viewer_{uuid4().hex[:6]}",
        organization_role="viewer",
    )
    viewer_headers = viewer["headers"]  # type: ignore[assignment]

    org_code = f"ACTORG{uuid4().hex[:6].upper()}"
    create_org = client.post(
        "/api/v1/settings/organizations/upsert",
        json={
            "code": org_code,
            "name": f"Activity Org {org_code}",
            "org_type": "contractor",
            "is_active": True,
            "contracts": [
                {
                    "contract_number": f"CN-{uuid4().hex[:4].upper()}",
                    "subject": "Scoped Contract",
                    "block_id": block_id,
                }
            ],
        },
        headers=admin,
    )
    assert create_org.status_code == 200, create_org.text
    org_item = create_org.json().get("item") or {}
    organization_id = int(org_item.get("id") or 0)
    contracts = org_item.get("contracts") or []
    assert organization_id > 0
    assert contracts
    organization_contract_id = int(contracts[0].get("id") or 0)
    assert organization_contract_id > 0

    settings_list = client.get("/api/v1/settings/site-log-activity-catalog", headers=admin)
    assert settings_list.status_code == 200, settings_list.text
    assert settings_list.json().get("ok") is True

    blocked_settings_list = client.get("/api/v1/settings/site-log-activity-catalog", headers=viewer_headers)  # type: ignore[arg-type]
    assert blocked_settings_list.status_code == 403, blocked_settings_list.text

    code_project = f"A{uuid4().hex[:4].upper()}"
    code_duplicate = f"D{uuid4().hex[:4].upper()}"
    code_org_only = f"O{uuid4().hex[:4].upper()}"
    code_contract = f"C{uuid4().hex[:4].upper()}"

    project_only = _upsert_activity_catalog_item(
        admin,
        project_code=project_code,
        activity_code=code_project,
        activity_title="Project Default Activity",
        activity_type="Concrete",
        activity_type_code="CONC",
        floor="B1",
        wbs_code="1.2.3",
        default_quantity=125.5,
        default_location="Zone A",
        default_unit="Ton",
        sort_order=40,
    )
    duplicate_project = _upsert_activity_catalog_item(
        admin,
        project_code=project_code,
        activity_code=code_duplicate,
        activity_title="Project Duplicate Activity",
        default_location="Zone P",
        default_unit="m2",
        sort_order=50,
    )
    org_duplicate = _upsert_activity_catalog_item(
        admin,
        project_code=project_code,
        organization_id=organization_id,
        activity_code=code_duplicate,
        activity_title="Organization Duplicate Activity",
        default_location="Zone O",
        default_unit="m2",
        sort_order=20,
    )
    org_only = _upsert_activity_catalog_item(
        admin,
        project_code=project_code,
        organization_id=organization_id,
        activity_code=code_org_only,
        activity_title="Organization Only Activity",
        default_location="Zone O2",
        default_unit="Ton",
        sort_order=30,
    )
    contract_only = _upsert_activity_catalog_item(
        admin,
        project_code=project_code,
        organization_id=organization_id,
        organization_contract_id=organization_contract_id,
        activity_code=code_contract,
        activity_title="Contract Only Activity",
        default_location="Block B",
        default_unit="kg",
        sort_order=10,
    )
    assert (project_only.get("item") or {}).get("scope_code") == "project"
    assert (project_only.get("item") or {}).get("activity_type") == "Concrete"
    assert (project_only.get("item") or {}).get("activity_type_code") == "CONC"
    assert (project_only.get("item") or {}).get("floor") == "B1"
    assert (project_only.get("item") or {}).get("wbs_code") == "1.2.3"
    assert (project_only.get("item") or {}).get("default_quantity") == 125.5
    assert (org_only.get("item") or {}).get("scope_code") == "organization"
    assert (contract_only.get("item") or {}).get("scope_code") == "contract"

    runtime_res = client.get(
        f"/api/v1/site-logs/activity-options?project_code={project_code}&organization_id={organization_id}&organization_contract_id={organization_contract_id}",
        headers=admin,
    )
    assert runtime_res.status_code == 200, runtime_res.text
    runtime_body = runtime_res.json()
    runtime_rows = runtime_body.get("data") or []
    runtime_codes = [str(item.get("activity_code") or "") for item in runtime_rows]
    assert runtime_codes[:4] == [code_contract, code_duplicate, code_org_only, code_project]
    assert runtime_codes.count(code_duplicate) == 1
    runtime_project_row = next(row for row in runtime_rows if row.get("activity_code") == code_project)
    assert runtime_project_row.get("activity_type") == "Concrete"
    assert runtime_project_row.get("wbs_code") == "1.2.3"
    assert runtime_project_row.get("default_quantity") == 125.5

    delete_contract_item = client.post(
        "/api/v1/settings/site-log-activity-catalog/delete",
        json={"id": int((contract_only.get("item") or {}).get("id") or 0)},
        headers=admin,
    )
    assert delete_contract_item.status_code == 200, delete_contract_item.text
    assert bool((delete_contract_item.json().get("item") or {}).get("is_active")) is False

    runtime_after_delete = client.get(
        f"/api/v1/site-logs/activity-options?project_code={project_code}&organization_id={organization_id}&organization_contract_id={organization_contract_id}",
        headers=admin,
    )
    assert runtime_after_delete.status_code == 200, runtime_after_delete.text
    runtime_after_delete_codes = [
        str(item.get("activity_code") or "")
        for item in (runtime_after_delete.json().get("data") or [])
    ]
    assert code_contract not in runtime_after_delete_codes


def test_site_log_activity_catalog_excel_import_upserts_selected_scope() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)
    _existing_project_code, discipline_code = ensure_project_discipline(client, admin)
    project_code = f"ACTI{uuid4().hex[:5].upper()}"
    project_upsert = client.post(
        "/api/v1/settings/projects/upsert",
        json={"code": project_code, "name_e": f"Activity Import Project {project_code}", "is_active": True},
        headers=admin,
    )
    assert project_upsert.status_code == 200, project_upsert.text
    block_id = _ensure_block(admin, project_code)

    org_code = f"IMPACTORG{uuid4().hex[:5].upper()}"
    create_org = client.post(
        "/api/v1/settings/organizations/upsert",
        json={
            "code": org_code,
            "name": f"Import Activity Org {org_code}",
            "org_type": "contractor",
            "is_active": True,
            "contracts": [
                {
                    "contract_number": f"ICN-{uuid4().hex[:4].upper()}",
                    "subject": "Import Scoped Contract",
                    "block_id": block_id,
                }
            ],
        },
        headers=admin,
    )
    assert create_org.status_code == 200, create_org.text
    org_item = create_org.json().get("item") or {}
    organization_id = int(org_item.get("id") or 0)
    organization_contract_id = int((org_item.get("contracts") or [{}])[0].get("id") or 0)
    assert organization_id > 0
    assert organization_contract_id > 0

    template_res = client.get(
        "/api/v1/settings/site-log-activity-catalog/template",
        headers=admin,
    )
    assert template_res.status_code == 200, template_res.text
    assert "spreadsheetml.sheet" in (template_res.headers.get("content-type") or "")
    assert "site_log_activity_catalog_template.xlsx" in (template_res.headers.get("content-disposition") or "")
    template_workbook = load_workbook(BytesIO(template_res.content), read_only=True, data_only=True)
    try:
        template_sheet = template_workbook.active
        headers = [template_sheet.cell(row=1, column=index).value for index in range(1, 12)]
        assert headers == [
            "کد فعالیت",
            "عنوان فعالیت",
            "تیپ",
            "کد تیپ",
            "طبقه",
            "WBS",
            "مقدار",
            "محل پیش‌فرض",
            "واحد",
            "ترتیب",
            "وضعیت",
        ]
    finally:
        template_workbook.close()

    code_one = f"IMP{uuid4().hex[:4].upper()}"
    code_two = f"IMP{uuid4().hex[:4].upper()}"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["کد فعالیت", "عنوان فعالیت", "تیپ", "کد تیپ", "طبقه", "WBS", "مقدار", "محل پیش‌فرض", "واحد", "ترتیب", "وضعیت"])
    sheet.append([code_one, "Imported Activity One", "Concrete", "CONC", "B1", "1.2.3", 42.5, "Block A", "m3", 10, "فعال"])
    sheet.append([code_two, "Imported Activity Two", "Formwork", "FORM", "GF", "1.2.4", 10, "Block B", "kg", 20, "غیرفعال"])
    sheet.append(["", "Missing Code", "Steel", "STL", "B2", "1.2.5", 1, "Block C", "m2", 30, "فعال"])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)

    import_res = client.post(
        "/api/v1/settings/site-log-activity-catalog/import",
        data={
            "project_code": project_code,
            "organization_id": str(organization_id),
            "organization_contract_id": str(organization_contract_id),
        },
        files={
            "file": (
                "activity_import.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=admin,
    )
    assert import_res.status_code == 200, import_res.text
    import_body = import_res.json()
    assert import_body.get("ok") is True
    assert import_body.get("created") == 2
    assert import_body.get("updated") == 0
    assert import_body.get("skipped") == 1
    imported_rows = import_body.get("items") or []
    row_one = next((row for row in imported_rows if row.get("activity_code") == code_one), None)
    row_two = next((row for row in imported_rows if row.get("activity_code") == code_two), None)
    assert row_one is not None
    assert row_two is not None
    assert row_one.get("scope_code") == "contract"
    assert row_one.get("activity_type") == "Concrete"
    assert row_one.get("activity_type_code") == "CONC"
    assert row_one.get("floor") == "B1"
    assert row_one.get("wbs_code") == "1.2.3"
    assert row_one.get("default_quantity") == 42.5
    assert row_one.get("default_location") == "Block A"
    assert row_two.get("is_active") is False

    update_workbook = Workbook()
    update_sheet = update_workbook.active
    update_sheet.append(["activity_code", "activity_title", "activity_type", "activity_type_code", "floor", "wbs_code", "default_quantity", "default_location", "default_unit", "sort_order", "is_active"])
    update_sheet.append([code_one, "Imported Activity One Updated", "Concrete Updated", "CONC-U", "B2", "1.2.3.1", 55, "Block A1", "m3", 15, "yes"])
    update_buffer = BytesIO()
    update_workbook.save(update_buffer)
    update_workbook.close()
    update_buffer.seek(0)

    update_res = client.post(
        "/api/v1/settings/site-log-activity-catalog/import",
        data={
            "project_code": project_code,
            "organization_id": str(organization_id),
            "organization_contract_id": str(organization_contract_id),
        },
        files={
            "file": (
                "activity_update.xlsx",
                update_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=admin,
    )
    assert update_res.status_code == 200, update_res.text
    update_body = update_res.json()
    assert update_body.get("created") == 0
    assert update_body.get("updated") == 1
    updated_row = next((row for row in update_body.get("items", []) if row.get("activity_code") == code_one), None)
    assert updated_row is not None
    assert updated_row.get("activity_title") == "Imported Activity One Updated"
    assert updated_row.get("activity_type") == "Concrete Updated"
    assert updated_row.get("wbs_code") == "1.2.3.1"
    assert updated_row.get("default_quantity") == 55
    assert updated_row.get("sort_order") == 15

    runtime_res = client.get(
        f"/api/v1/site-logs/activity-options?project_code={project_code}&organization_id={organization_id}&organization_contract_id={organization_contract_id}",
        headers=admin,
    )
    assert runtime_res.status_code == 200, runtime_res.text
    runtime_codes = [str(row.get("activity_code") or "") for row in runtime_res.json().get("data", [])]
    assert code_one in runtime_codes
    assert code_two not in runtime_codes


def test_site_log_activity_pms_mapping_snapshot_reapply_and_report_step() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)
    project_code, discipline_code = ensure_project_discipline(client, admin)
    activity_code = f"PMS{uuid4().hex[:5].upper()}"
    activity = _upsert_activity_catalog_item(
        admin,
        project_code=project_code,
        activity_code=activity_code,
        activity_title="PMS mapped activity",
        default_location="Deck",
        default_unit="each",
    )
    activity_id = int((activity.get("item") or {}).get("id") or 0)
    assert activity_id > 0

    template_code = f"TPL{uuid4().hex[:5].upper()}"
    template_res = client.post(
        "/api/v1/settings/site-log-pms/templates/upsert",
        json={
            "code": template_code,
            "title": "Deck PMS",
            "sort_order": 10,
            "is_active": True,
            "steps": [
                {"step_code": "INSTALL", "step_title": "Install", "weight_pct": 80, "sort_order": 10, "is_active": True},
                {"step_code": "QC", "step_title": "QC Check", "weight_pct": 20, "sort_order": 20, "is_active": True},
            ],
        },
        headers=admin,
    )
    assert template_res.status_code == 200, template_res.text
    template = template_res.json().get("item") or {}
    template_id = int(template.get("id") or 0)
    assert template_id > 0
    assert template.get("version") == 1

    without_pms = client.get(
        f"/api/v1/settings/site-log-activity-catalog?project_code={project_code}&pms_status=none",
        headers=admin,
    )
    assert without_pms.status_code == 200, without_pms.text
    assert activity_code in {str(row.get("activity_code") or "") for row in without_pms.json().get("items", [])}

    apply_res = client.post(
        "/api/v1/settings/site-log-pms/mappings/apply",
        json={"activity_ids": [activity_id], "template_id": template_id, "overwrite": False},
        headers=admin,
    )
    assert apply_res.status_code == 200, apply_res.text
    mapped_item = (apply_res.json().get("items") or [])[0]
    mapping_id = int(mapped_item.get("pms_mapping_id") or 0)
    assert mapping_id > 0
    assert mapped_item.get("pms_status") == "mapped"
    assert mapped_item.get("pms_snapshot_version") == 1

    conflict_res = client.post(
        "/api/v1/settings/site-log-pms/mappings/apply",
        json={"activity_ids": [activity_id], "template_id": template_id, "overwrite": False},
        headers=admin,
    )
    assert conflict_res.status_code == 409, conflict_res.text

    runtime_res = client.get(f"/api/v1/site-logs/activity-options?project_code={project_code}", headers=admin)
    assert runtime_res.status_code == 200, runtime_res.text
    runtime_row = next(row for row in runtime_res.json().get("data", []) if row.get("activity_code") == activity_code)
    assert runtime_row.get("pms_mapping_id") == mapping_id
    assert [step.get("step_code") for step in runtime_row.get("pms_steps", [])] == ["INSTALL", "QC"]

    log_res = client.post(
        "/api/v1/site-logs/create",
        json={
            "log_type": "DAILY",
            "project_code": project_code,
            "discipline_code": discipline_code,
            "log_date": datetime.utcnow().strftime("%Y-%m-%dT00:00:00"),
            "current_work_summary": "PMS daily work",
            "activity_rows": [
                {
                    "activity_code": activity_code,
                    "activity_title": "PMS mapped activity",
                    "source_system": "CATALOG",
                    "external_ref": f"site_log_activity_catalog:{activity_id}",
                    "pms_mapping_id": mapping_id,
                    "pms_step_code": "INSTALL",
                    "today_quantity": 350,
                    "unit": "each",
                    "sort_order": 0,
                }
            ],
        },
        headers=admin,
    )
    assert log_res.status_code == 200, log_res.text
    saved_activity = (log_res.json().get("data") or {}).get("activity_rows", [])[0]
    assert saved_activity.get("pms_template_code") == template_code
    assert saved_activity.get("pms_step_code") == "INSTALL"
    assert saved_activity.get("pms_step_title") == "Install"
    assert saved_activity.get("pms_step_weight_pct") == 80

    update_template = client.post(
        "/api/v1/settings/site-log-pms/templates/upsert",
        json={
            "id": template_id,
            "code": template_code,
            "title": "Deck PMS updated",
            "sort_order": 10,
            "is_active": True,
            "steps": [
                {"step_code": "INSTALL", "step_title": "Install", "weight_pct": 70, "sort_order": 10, "is_active": True},
                {"step_code": "QC", "step_title": "QC Check", "weight_pct": 30, "sort_order": 20, "is_active": True},
            ],
        },
        headers=admin,
    )
    assert update_template.status_code == 200, update_template.text
    assert (update_template.json().get("item") or {}).get("version") == 2

    stale_list = client.get(
        f"/api/v1/settings/site-log-activity-catalog?project_code={project_code}&pms_status=stale",
        headers=admin,
    )
    assert stale_list.status_code == 200, stale_list.text
    stale_row = next(row for row in stale_list.json().get("items", []) if row.get("activity_code") == activity_code)
    assert stale_row.get("pms_status") == "stale"

    reapply_res = client.post(
        "/api/v1/settings/site-log-pms/mappings/reapply",
        json={"activity_ids": [activity_id]},
        headers=admin,
    )
    assert reapply_res.status_code == 200, reapply_res.text
    reapplied = (reapply_res.json().get("items") or [])[0]
    assert reapplied.get("pms_status") == "mapped"
    assert reapplied.get("pms_snapshot_version") == 2


def test_site_log_shift_weather_catalog_validation_and_labels() -> None:
    _assert_catalog_schema_ready()
    admin = admin_headers(client)
    project_code, discipline_code = ensure_project_discipline(client, admin)
    contractor = create_scoped_user_and_login(
        client,
        admin,
        org_type="contractor",
        project_code=project_code,
        discipline_code=discipline_code,
        email_prefix=f"site_log_shift_weather_{uuid4().hex[:6]}",
        organization_role="manager",
    )
    headers = contractor["headers"]  # type: ignore[assignment]
    organization_id = int(contractor.get("organization_id") or 0)

    inactive_shift = f"OFF{uuid4().hex[:5].upper()}"
    _upsert_catalog_item(
        admin,
        catalog_type="shift",
        code=inactive_shift,
        label="Inactive shift",
        is_active=False,
    )

    payload = {
        "log_type": "DAILY",
        "project_code": project_code,
        "discipline_code": discipline_code,
        "organization_id": organization_id,
        "log_date": datetime.utcnow().strftime("%Y-%m-%dT00:00:00"),
        "shift": inactive_shift,
        "weather": "CLEAR",
        "current_work_summary": "Shift validation current",
        "next_plan_summary": "Shift validation next",
        "manpower_rows": [],
        "equipment_rows": [],
        "activity_rows": [],
        "material_rows": [],
        "issue_rows": [],
        "attachment_rows": [],
    }
    invalid_create = client.post("/api/v1/site-logs/create", json=payload, headers=headers)  # type: ignore[arg-type]
    assert invalid_create.status_code == 400, invalid_create.text

    invalid_issue_payload = dict(payload)
    invalid_issue_payload["shift"] = "DAY"
    invalid_issue_payload["issue_rows"] = [{"issue_type": "UNLISTED_ISSUE", "description": "Invalid issue type"}]
    invalid_issue_create = client.post("/api/v1/site-logs/create", json=invalid_issue_payload, headers=headers)  # type: ignore[arg-type]
    assert invalid_issue_create.status_code == 400, invalid_issue_create.text

    payload["shift"] = "DAY"
    payload["issue_rows"] = [{"issue_type": "MATERIAL", "description": "Material delivery blocked", "sort_order": 0}]
    created = client.post("/api/v1/site-logs/create", json=payload, headers=headers)  # type: ignore[arg-type]
    assert created.status_code == 200, created.text
    created_data = created.json().get("data") or {}
    assert created_data.get("shift") == "DAY"
    assert created_data.get("shift_label") in {"روز", "DAY"}
    assert created_data.get("weather") == "CLEAR"
    assert created_data.get("weather_label") in {"صاف", "CLEAR"}
    created_issue = (created_data.get("issue_rows") or [])[0]
    assert created_issue.get("issue_type") == "MATERIAL"
    assert created_issue.get("issue_type_label")
    log_id = int(created_data.get("id") or 0)

    with SessionLocal() as db:
        row = db.query(SiteLog).filter(SiteLog.id == log_id).first()
        assert row is not None
        row.shift = "LEGACY_SHIFT"
        row.weather = "LEGACY_WEATHER"
        db.commit()

    unchanged_legacy = client.put(
        f"/api/v1/site-logs/{log_id}",
        json={"shift": "LEGACY_SHIFT", "weather": "LEGACY_WEATHER"},
        headers=headers,  # type: ignore[arg-type]
    )
    assert unchanged_legacy.status_code == 200, unchanged_legacy.text
    unchanged_data = unchanged_legacy.json().get("data") or {}
    assert unchanged_data.get("shift") == "LEGACY_SHIFT"
    assert unchanged_data.get("shift_label") == "LEGACY_SHIFT"
    unchanged_issue = (unchanged_data.get("issue_rows") or [])[0]
    assert unchanged_issue.get("issue_type") == "MATERIAL"

    invalid_update = client.put(
        f"/api/v1/site-logs/{log_id}",
        json={"shift": inactive_shift},
        headers=headers,  # type: ignore[arg-type]
    )
    assert invalid_update.status_code == 400, invalid_update.text
