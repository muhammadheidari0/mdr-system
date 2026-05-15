# app/api/v1/routers/settings.py
from __future__ import annotations

import csv
import io
import json
import re
from urllib.parse import urlparse
from datetime import datetime
from typing import Optional, List, Dict, Any
import traceback  # ✅ برای لاگ خطاهای دقیق

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy.exc import IntegrityError

from app.api.dependencies import _load_allowed_permissions, get_db, require_permission
from app.core.config import settings
from app.core.redaction import redact_secrets
from app.core.organizations import (
    ALL_ORG_TYPES,
    DEFAULT_PERMISSION_CATEGORY,
    PERMISSION_CATEGORIES,
    normalize_permission_category,
)
from app.core.access_matrix import (
    CANONICAL_MATRIX_ROLES,
    CANONICAL_PERMISSION_CATEGORIES,
    build_navigation_diagnostics,
    build_navigation_state,
    default_permission_matrix_for_category,
)
from app.core.roles import MATRIX_ROLES
from app.core.permission_catalog import feature_catalog, permission_keys, permission_meta_list
from app.db.models import (
    Project, Phase, Discipline, Package, Level,
    SettingsKV, Block, MdrCategory, DocStatus, User as DbUser, Organization, WorkboardItem,
    OrganizationContract,
    RoleCategoryDisciplineScope, RoleCategoryPermission, RoleCategoryProjectScope,
    RolePermission, RoleProjectScope, RoleDisciplineScope,
    UserProjectScope, UserDisciplineScope,
    SettingsAuditLog,
    PowerBiApiToken,
    ArchiveFile, Correspondence, CorrespondenceAction, CorrespondenceAttachment,
    CorrespondenceCategory, CorrespondenceDepartment, DocumentTag, IssuingEntity,
    WorkflowStatus, WorkflowTransition, TechSubtype, ReviewResult,
    SiteLogActivityCatalog, SiteLogRoleCatalog, SiteLogWorkSectionCatalog, SiteLogEquipmentCatalog, SiteLogMaterialCatalog, SiteLogEquipmentStatusCatalog,
    SiteLogAttachmentTypeCatalog, SiteLogIssueTypeCatalog, SiteLogShiftCatalog, SiteLogWeatherCatalog,
    SiteLogActivityPmsMapping, SiteLogActivityPmsStep, SiteLogPmsTemplate, SiteLogPmsTemplateStep,
)

# ✅ استفاده از سرویس Seed
from app.services import seed_service
from app.services.storage_policy import (
    DEFAULT_BIM_REVIT_INTEGRATION,
    DEFAULT_STORAGE_INTEGRATIONS,
    DEFAULT_STORAGE_POLICY,
    get_bim_revit_integration,
    get_storage_integrations,
    get_storage_policy,
    resolve_primary_storage_provider,
    set_bim_revit_integration,
    set_storage_integrations,
    set_storage_policy,
)
from app.services.bim_revit_security import encrypt_plugin_secret, generate_plugin_secret
from app.services.power_bi_tokens import (
    POWER_BI_SITE_LOG_SCOPE,
    create_power_bi_token,
    serialize_power_bi_token,
)
from app.services.storage_sync import resolve_nextcloud_runtime, resolve_openproject_runtime
from app.services.storage import StorageManager
from app.services.access_control import resolve_effective_access
from app.services import tag_service
from app.services.transmittal_options import (
    get_transmittal_parties,
    set_transmittal_parties,
)

router = APIRouter(
    prefix="/settings",
    tags=["settings"],
    dependencies=[Depends(require_permission("settings:read"))],
)

KV_RESERVED_KEYS = {
    "transmittal.state.v1",
    "rbac.matrix.v1",
    "rbac.scope.v1",
    "rbac.user_scope.v1",
}
KV_ALLOWED_WRITE_PREFIXES = ("ui.", "feature.", "custom.")
STORAGE_PATH_MDR_KEY = "mdr_storage_path"
STORAGE_PATH_CORRESPONDENCE_KEY = "correspondence_storage_path"
STORAGE_PATH_SITE_LOG_KEY = "site_log_storage_path"
DEFAULT_MDR_STORAGE_PATH = "./files/technical"
DEFAULT_CORRESPONDENCE_STORAGE_PATH = "./files/correspondence"


# -----------------------------
# Helpers
# -----------------------------
def _db_overview_payload() -> dict[str, str]:
    masked = settings.masked_database_url()
    return {
        "url": masked,
        "env": str(settings.APP_ENV or ""),
    }


# -----------------------------
# Pydantic Schemas
# -----------------------------
class ProjectIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=50)
    name_e: Optional[str] = Field(default=None, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    project_name: Optional[str] = Field(default=None, max_length=255)
    root_path: Optional[str] = Field(default=None, max_length=1024)
    is_active: Optional[bool] = None
    docnum_template: Optional[str] = Field(default=None, max_length=2000)

class PhaseIn(BaseModel):
    ph_code: str = Field(..., min_length=1, max_length=10)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: str = Field(..., min_length=1, max_length=255)

class DisciplineIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=20)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)

class PackageIn(BaseModel):
    discipline_code: str = Field(..., min_length=1, max_length=20)
    package_code: Optional[str] = Field(default=None, max_length=30)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)

class LevelIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=20)
    name_e: Optional[str] = Field(default=None, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    sort_order: int = 0

class LevelDeleteIn(BaseModel):
    code: str

class KvIn(BaseModel):
    key: str = Field(..., min_length=1, max_length=120)
    value: str = Field(..., min_length=0, max_length=5000)

class StoragePathsIn(BaseModel):
    mdr_storage_path: str = Field(..., min_length=1, max_length=1024)
    correspondence_storage_path: str = Field(..., min_length=1, max_length=1024)
    site_log_storage_path: Optional[str] = Field(default=None, max_length=1024)
    network_username: Optional[str] = Field(default=None, max_length=255)
    network_password: Optional[str] = Field(default=None, max_length=255)

class StoragePolicyIn(BaseModel):
    enforcement_mode: Optional[str] = Field(default=None, max_length=20)
    allowed_mimes: Optional[List[str]] = None
    allowed_mimes_by_kind: Optional[Dict[str, List[str]]] = None
    blocked_extensions: Optional[List[str]] = None
    dangerous_mimes: Optional[List[str]] = None
    max_size_mb: Optional[Dict[str, int]] = None


class StorageIntegrationsIn(BaseModel):
    primary: Optional[Dict[str, Any]] = None
    mirror: Optional[Dict[str, Any]] = None
    google_drive: Optional[Dict[str, Any]] = None
    openproject: Optional[Dict[str, Any]] = None
    nextcloud: Optional[Dict[str, Any]] = None
    local_cache: Optional[Dict[str, Any]] = None


class BimRevitSettingsIn(BaseModel):
    enabled: Optional[bool] = None
    api_endpoint_url: Optional[str] = Field(default=None, max_length=1024)
    require_plugin_signature: Optional[bool] = None
    plugin_key_id: Optional[str] = Field(default=None, max_length=128)
    plugin_secret: Optional[str] = Field(default=None, max_length=512)
    default_category_id: Optional[int] = Field(default=None, ge=1)
    default_folder_id: Optional[int] = Field(default=None, ge=1)
    allowed_mime: Optional[List[str]] = None
    max_batch_size: Optional[int] = Field(default=None, ge=1, le=5000)


class PowerBiTokenMintIn(BaseModel):
    name: str = Field(default="Power BI Dashboard", min_length=1, max_length=255)
    expires_at: Optional[datetime] = None
    allowed_project_codes: Optional[List[str]] = None
    allowed_report_sections: Optional[List[str]] = None
    allowed_ip_ranges: Optional[List[str]] = None


class BlockIn(BaseModel):
    project_code: str = Field(..., min_length=1, max_length=50)
    code: str = Field(..., min_length=1, max_length=10)
    name_e: Optional[str] = Field(default=None, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    sort_order: int = 0
    is_active: bool = True

class BlockDeleteIn(BaseModel):
    project_code: str
    code: str
    hard_delete: bool = False

class MdrCategoryIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=10)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    folder_name: Optional[str] = Field(default=None, max_length=255)
    sort_order: int = 0
    is_active: bool = True

class MdrCategoryDeleteIn(BaseModel):
    code: str
    hard_delete: bool = False

class ProjectDeleteIn(BaseModel):
    code: str
    hard_delete: bool = False

class PhaseDeleteIn(BaseModel):
    ph_code: str

class DisciplineDeleteIn(BaseModel):
    code: str

class PackageDeleteIn(BaseModel):
    discipline_code: str
    package_code: str

class DocStatusIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=20)
    name: str = Field(..., min_length=1, max_length=100)
    description: Optional[str] = Field(default=None, max_length=255)
    sort_order: int = 0

class DocStatusDeleteIn(BaseModel):
    code: str

class CorrespondenceIssuingIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=20)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    project_code: Optional[str] = Field(default=None, max_length=50)
    is_active: bool = True
    sort_order: int = 0

class CorrespondenceIssuingDeleteIn(BaseModel):
    code: str
    hard_delete: bool = False

class CorrespondenceCategoryIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=20)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    is_active: bool = True
    sort_order: int = 0

class CorrespondenceCategoryDeleteIn(BaseModel):
    code: str
    hard_delete: bool = False


class CorrespondenceDepartmentIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=32)
    name_e: str = Field(..., min_length=1, max_length=255)
    name_p: Optional[str] = Field(default=None, max_length=255)
    is_active: bool = True
    sort_order: int = 0


class CorrespondenceDepartmentDeleteIn(BaseModel):
    code: str
    hard_delete: bool = False


class CorrespondenceTagIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    name: str = Field(..., min_length=1, max_length=64)
    color: Optional[str] = Field(default=None, max_length=16)


class CorrespondenceTagDeleteIn(BaseModel):
    id: int = Field(..., ge=1)


class TransmittalPartyOptionIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=20)
    label: str = Field(..., min_length=1, max_length=100)
    is_active: bool = True
    sort_order: int = 0


class TransmittalPartiesIn(BaseModel):
    direction_options: List[TransmittalPartyOptionIn] = Field(default_factory=list)
    recipient_options: List[TransmittalPartyOptionIn] = Field(default_factory=list)


class WorkflowStatusIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    item_type: str = Field(..., min_length=1, max_length=16)
    code: str = Field(..., min_length=1, max_length=64)
    label: str = Field(..., min_length=1, max_length=128)
    is_terminal: bool = False
    sort_order: int = 0
    is_active: bool = True


class WorkflowTransitionIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    item_type: str = Field(..., min_length=1, max_length=16)
    from_status_code: str = Field(..., min_length=1, max_length=64)
    to_status_code: str = Field(..., min_length=1, max_length=64)
    requires_note: bool = False
    is_active: bool = True


class TechSubtypeIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=32)
    label: str = Field(..., min_length=1, max_length=128)
    sort_order: int = 0
    is_active: bool = True


class ReviewResultIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=32)
    label: str = Field(..., min_length=1, max_length=128)
    sort_order: int = 0
    is_active: bool = True


class OrganizationContractIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    contract_number: Optional[str] = Field(default=None, max_length=128)
    subject: Optional[str] = Field(default=None, max_length=500)
    block_id: Optional[int] = Field(default=None, ge=1)


class OrganizationIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    code: str = Field(..., min_length=1, max_length=64)
    name: str = Field(..., min_length=1, max_length=255)
    org_type: str = Field(default="contractor", min_length=1, max_length=32)
    parent_id: Optional[int] = Field(default=None, ge=1)
    is_active: bool = True
    contracts: Optional[List[OrganizationContractIn]] = None


class OrganizationDeleteIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    code: Optional[str] = Field(default=None, max_length=64)
    hard_delete: bool = False


class PermissionMatrixIn(BaseModel):
    matrix: Dict[str, Dict[str, bool]]


class PermissionScopeIn(BaseModel):
    scope: Dict[str, Dict[str, List[str]]]


class UserPermissionScopeIn(BaseModel):
    user_id: int
    projects: List[str] = Field(default_factory=list)
    disciplines: List[str] = Field(default_factory=list)


class SiteLogCatalogUpsertIn(BaseModel):
    catalog_type: str = Field(..., min_length=1, max_length=32)
    id: Optional[int] = Field(default=None, ge=1)
    code: str = Field(..., min_length=1, max_length=32)
    label: str = Field(..., min_length=1, max_length=255)
    sort_order: int = 0
    is_active: bool = True


class SiteLogCatalogBulkItemIn(BaseModel):
    code: Optional[str] = Field(default=None, max_length=128)
    label: str = Field(..., min_length=1, max_length=255)
    sort_order: Optional[int] = None
    is_active: bool = True


class SiteLogCatalogBulkUpsertIn(BaseModel):
    catalog_type: str = Field(..., min_length=1, max_length=32)
    items: List[SiteLogCatalogBulkItemIn] = Field(..., min_length=1, max_length=500)
    overwrite_existing: bool = False


class SiteLogCatalogDeleteIn(BaseModel):
    catalog_type: str = Field(..., min_length=1, max_length=32)
    id: int = Field(..., ge=1)


class SiteLogActivityCatalogUpsertIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    project_code: str = Field(..., min_length=1, max_length=50)
    organization_id: Optional[int] = Field(default=None, ge=1)
    organization_contract_id: Optional[int] = Field(default=None, ge=1)
    activity_code: str = Field(..., min_length=1, max_length=64)
    activity_title: str = Field(..., min_length=1, max_length=255)
    default_location: Optional[str] = Field(default=None, max_length=255)
    default_unit: Optional[str] = Field(default=None, max_length=64)
    sort_order: int = 0
    is_active: bool = True


class SiteLogActivityCatalogDeleteIn(BaseModel):
    id: int = Field(..., ge=1)


class SiteLogPmsStepIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    step_code: str = Field(..., min_length=1, max_length=64)
    step_title: str = Field(..., min_length=1, max_length=255)
    weight_pct: float = Field(default=0, ge=0, le=100)
    sort_order: int = 0
    is_active: bool = True


class SiteLogPmsTemplateUpsertIn(BaseModel):
    id: Optional[int] = Field(default=None, ge=1)
    code: str = Field(..., min_length=1, max_length=64)
    title: str = Field(..., min_length=1, max_length=255)
    description: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True
    steps: List[SiteLogPmsStepIn] = Field(default_factory=list)


class SiteLogPmsTemplateDeleteIn(BaseModel):
    id: int = Field(..., ge=1)


class SiteLogPmsMappingApplyIn(BaseModel):
    activity_ids: List[int] = Field(..., min_length=1)
    template_id: Optional[int] = Field(default=None, ge=1)
    template_code: Optional[str] = Field(default=None, max_length=64)
    overwrite: bool = False


class SiteLogPmsMappingDeleteIn(BaseModel):
    activity_id: int = Field(..., ge=1)


class SiteLogPmsMappingReapplyIn(BaseModel):
    activity_ids: List[int] = Field(..., min_length=1)


# -----------------------------
# Helpers
# -----------------------------
def _norm(s: Any) -> str:
    if s is None: return ""
    return str(s).strip()

def _upper(s: Any) -> str:
    return _norm(s).upper()


SITE_LOG_CATALOG_MODELS: Dict[str, Any] = {
    "role": SiteLogRoleCatalog,
    "work_section": SiteLogWorkSectionCatalog,
    "equipment": SiteLogEquipmentCatalog,
    "material": SiteLogMaterialCatalog,
    "equipment_status": SiteLogEquipmentStatusCatalog,
    "attachment_type": SiteLogAttachmentTypeCatalog,
    "issue_type": SiteLogIssueTypeCatalog,
    "shift": SiteLogShiftCatalog,
    "weather": SiteLogWeatherCatalog,
}

SITE_LOG_CATALOG_TITLES: Dict[str, str] = {
    "role": "فهرست نقش‌ها",
    "equipment": "فهرست تجهیزات",
    "material": "فهرست مصالح",
    "equipment_status": "فهرست وضعیت تجهیزات",
    "attachment_type": "فهرست نوع پیوست گزارش",
    "issue_type": "فهرست نوع موانع",
}
SITE_LOG_CATALOG_TITLES.update(
    {
        "shift": "فهرست شیفت‌های گزارش",
        "weather": "فهرست وضعیت‌های جوی",
    }
)


SITE_LOG_CATALOG_TITLES["work_section"] = "فهرست واحد / بخش کاری نفرات"


SITE_LOG_BULK_CATALOG_PREFIXES: Dict[str, str] = {
    "equipment": "EQP",
    "material": "MAT",
}


def _normalize_site_log_catalog_type_or_400(value: Optional[str]) -> str:
    key = _norm(value).lower()
    if key not in SITE_LOG_CATALOG_MODELS:
        raise HTTPException(status_code=400, detail=f"Invalid catalog_type: {value}")
    return key


def _site_log_catalog_model(catalog_type: str):
    return SITE_LOG_CATALOG_MODELS[_normalize_site_log_catalog_type_or_400(catalog_type)]


def _next_site_log_bulk_code(db: Session, model: Any, prefix: str, used_codes: set[str]) -> str:
    prefix_value = _upper(prefix) or "CAT"
    max_number = 0
    for (raw_code,) in db.query(model.code).filter(func.upper(model.code).like(f"{prefix_value}%")).all():
        code = _upper(raw_code)
        suffix = code[len(prefix_value):]
        if suffix.isdigit():
            max_number = max(max_number, int(suffix))

    next_number = max_number + 1
    while True:
        code = f"{prefix_value}{next_number:04d}"
        if code not in used_codes:
            used_codes.add(code)
            return code
        next_number += 1


def _serialize_site_log_catalog_row(row: Any) -> Dict[str, Any]:
    return {
        "id": int(getattr(row, "id", 0) or 0),
        "code": _upper(getattr(row, "code", None)),
        "label": _norm(getattr(row, "label", None)),
        "sort_order": int(getattr(row, "sort_order", 0) or 0),
        "is_active": bool(getattr(row, "is_active", False)),
    }


def _load_site_log_catalog_items(db: Session, catalog_type: str) -> List[Dict[str, Any]]:
    model = _site_log_catalog_model(catalog_type)
    rows = (
        db.query(model)
        .order_by(model.sort_order.asc(), model.code.asc(), model.id.asc())
        .all()
    )
    return [_serialize_site_log_catalog_row(row) for row in rows]


def _load_site_log_catalogs_payload(db: Session) -> Dict[str, List[Dict[str, Any]]]:
    return {
        catalog_type: _load_site_log_catalog_items(db, catalog_type)
        for catalog_type in SITE_LOG_CATALOG_MODELS.keys()
    }


def _serialize_site_log_activity_catalog_row(row: SiteLogActivityCatalog) -> Dict[str, Any]:
    contract = row.organization_contract
    scope_code = "project"
    if row.organization_contract_id:
        scope_code = "contract"
    elif row.organization_id:
        scope_code = "organization"
    pms_payload = _serialize_activity_pms_summary(getattr(row, "pms_mapping", None))
    return {
        "id": int(row.id or 0),
        "project_code": _upper(row.project_code),
        "project_name": _norm(getattr(getattr(row, "project", None), "name_e", None))
        or _norm(getattr(getattr(row, "project", None), "name_p", None))
        or _upper(row.project_code),
        "organization_id": int(row.organization_id or 0) or None,
        "organization_name": _norm(getattr(getattr(row, "organization", None), "name", None)) or None,
        "organization_contract_id": int(row.organization_contract_id or 0) or None,
        "contract_number": _norm(getattr(contract, "contract_number", None)) or None,
        "contract_subject": _norm(getattr(contract, "subject", None)) or None,
        "activity_code": _upper(row.activity_code),
        "activity_title": _norm(row.activity_title),
        "default_location": _norm(row.default_location) or None,
        "default_unit": _norm(row.default_unit) or None,
        "sort_order": int(row.sort_order or 0),
        "is_active": bool(row.is_active),
        "scope_code": scope_code,
        "scope_label": (
            "سطح قرارداد"
            if scope_code == "contract"
            else "سطح سازمان" if scope_code == "organization" else "سطح پروژه"
        ),
        **pms_payload,
    }


def _serialize_pms_step(row: SiteLogPmsTemplateStep | SiteLogActivityPmsStep) -> Dict[str, Any]:
    return {
        "id": int(getattr(row, "id", 0) or 0),
        "step_code": _upper(getattr(row, "step_code", None)),
        "step_title": _norm(getattr(row, "step_title", None)),
        "weight_pct": float(getattr(row, "weight_pct", 0) or 0),
        "sort_order": int(getattr(row, "sort_order", 0) or 0),
        "is_active": bool(getattr(row, "is_active", False)),
    }


def _active_template_steps(template: SiteLogPmsTemplate) -> List[SiteLogPmsTemplateStep]:
    return [
        step
        for step in sorted(template.steps or [], key=lambda item: (int(item.sort_order or 0), int(item.id or 0)))
        if bool(step.is_active)
    ]


def _template_weight_total(template: SiteLogPmsTemplate) -> float:
    return round(sum(float(step.weight_pct or 0) for step in _active_template_steps(template)), 6)


def _serialize_pms_template(row: SiteLogPmsTemplate, *, include_steps: bool = True) -> Dict[str, Any]:
    steps = sorted(row.steps or [], key=lambda item: (int(item.sort_order or 0), int(item.id or 0)))
    active_steps = [step for step in steps if bool(step.is_active)]
    payload: Dict[str, Any] = {
        "id": int(row.id or 0),
        "code": _upper(row.code),
        "title": _norm(row.title),
        "description": _norm(row.description) or None,
        "version": int(row.version or 1),
        "sort_order": int(row.sort_order or 0),
        "is_active": bool(row.is_active),
        "active_step_count": len(active_steps),
        "weight_total": _template_weight_total(row),
    }
    if include_steps:
        payload["steps"] = [_serialize_pms_step(step) for step in steps]
    return payload


def _serialize_activity_pms_summary(mapping: SiteLogActivityPmsMapping | None) -> Dict[str, Any]:
    if not mapping:
        return {
            "pms_mapping_id": None,
            "pms_template_id": None,
            "pms_template_code": None,
            "pms_template_title": None,
            "pms_snapshot_version": None,
            "pms_template_version": None,
            "pms_status": "none",
            "pms_status_label": "بدون PMS",
            "pms_steps": [],
        }
    template = mapping.template
    template_version = int(getattr(template, "version", None) or mapping.snapshot_version or 1)
    snapshot_version = int(mapping.snapshot_version or 1)
    status = "stale" if template_version != snapshot_version else "mapped"
    steps = [
        _serialize_pms_step(step)
        for step in sorted(mapping.steps or [], key=lambda item: (int(item.sort_order or 0), int(item.id or 0)))
        if bool(step.is_active)
    ]
    return {
        "pms_mapping_id": int(mapping.id or 0),
        "pms_template_id": int(mapping.template_id or 0),
        "pms_template_code": _upper(mapping.template_code),
        "pms_template_title": _norm(mapping.template_title),
        "pms_snapshot_version": snapshot_version,
        "pms_template_version": template_version,
        "pms_status": status,
        "pms_status_label": "قدیمی شده" if status == "stale" else "دارای PMS",
        "pms_steps": steps,
    }


def _load_pms_templates(db: Session, *, active_only: bool = False) -> List[SiteLogPmsTemplate]:
    query = db.query(SiteLogPmsTemplate).options(selectinload(SiteLogPmsTemplate.steps))
    if active_only:
        query = query.filter(SiteLogPmsTemplate.is_active == True)
    return query.order_by(
        SiteLogPmsTemplate.sort_order.asc(),
        SiteLogPmsTemplate.code.asc(),
        SiteLogPmsTemplate.id.asc(),
    ).all()


def _load_pms_template_or_404(
    db: Session,
    template_id: int | None = None,
    template_code: str | None = None,
) -> SiteLogPmsTemplate:
    query = db.query(SiteLogPmsTemplate).options(selectinload(SiteLogPmsTemplate.steps))
    if template_id:
        query = query.filter(SiteLogPmsTemplate.id == int(template_id))
    else:
        code = _upper(template_code)
        if not code:
            raise HTTPException(status_code=400, detail="template_id or template_code is required")
        query = query.filter(func.upper(SiteLogPmsTemplate.code) == code)
    row = query.first()
    if not row:
        raise HTTPException(status_code=404, detail="PMS Template not found")
    return row


def _ensure_pms_template_ready(template: SiteLogPmsTemplate) -> None:
    if not bool(template.is_active):
        raise HTTPException(status_code=400, detail="PMS Template is inactive")
    active_steps = _active_template_steps(template)
    if not active_steps:
        raise HTTPException(status_code=400, detail="PMS Template must have at least one active step")
    if abs(_template_weight_total(template) - 100.0) > 0.001:
        raise HTTPException(status_code=400, detail="Active PMS Step weights must sum to 100")


def _copy_template_to_activity_mapping(
    db: Session,
    *,
    activity: SiteLogActivityCatalog,
    template: SiteLogPmsTemplate,
    overwrite: bool,
) -> SiteLogActivityPmsMapping:
    _ensure_pms_template_ready(template)
    mapping = activity.pms_mapping
    if mapping and not overwrite:
        raise HTTPException(status_code=409, detail=f"Activity already has PMS: {activity.activity_code}")
    if not mapping:
        mapping = SiteLogActivityPmsMapping(activity_catalog=activity)
        db.add(mapping)
    mapping.template = template
    mapping.template_id = int(template.id or 0)
    mapping.template_code = _upper(template.code)
    mapping.template_title = _norm(template.title)
    mapping.snapshot_version = int(template.version or 1)
    mapping.updated_at = datetime.utcnow()
    if int(mapping.id or 0) > 0 and mapping.steps:
        mapping.steps[:] = []
        db.flush()
    for step in _active_template_steps(template):
        mapping.steps.append(
            SiteLogActivityPmsStep(
                source_template_step_id=int(step.id or 0) or None,
                step_code=_upper(step.step_code),
                step_title=_norm(step.step_title),
                weight_pct=float(step.weight_pct or 0),
                sort_order=int(step.sort_order or 0),
                is_active=True,
            )
        )
    return mapping


def _check_optional_project_or_404(db: Session, project_code: str | None) -> str | None:
    value = _upper(project_code)
    if not value:
        return None
    if not db.query(Project.code).filter(Project.code == value).first():
        raise HTTPException(status_code=404, detail=f"Project not found: {project_code}")
    return value


def _check_optional_organization_or_404(db: Session, organization_id: int | None) -> int | None:
    if not organization_id:
        return None
    if not db.query(Organization.id).filter(Organization.id == int(organization_id)).first():
        raise HTTPException(status_code=404, detail=f"Organization not found: {organization_id}")
    return int(organization_id)


def _load_site_log_activity_catalog_row_or_404(db: Session, item_id: int) -> SiteLogActivityCatalog:
    row = (
        db.query(SiteLogActivityCatalog)
        .options(
            joinedload(SiteLogActivityCatalog.project),
            joinedload(SiteLogActivityCatalog.organization),
            joinedload(SiteLogActivityCatalog.organization_contract).joinedload(OrganizationContract.block),
            joinedload(SiteLogActivityCatalog.pms_mapping).joinedload(SiteLogActivityPmsMapping.template),
            joinedload(SiteLogActivityCatalog.pms_mapping).selectinload(SiteLogActivityPmsMapping.steps),
        )
        .filter(SiteLogActivityCatalog.id == int(item_id))
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Activity catalog item not found")
    return row


def _activity_catalog_scope_query(
    db: Session,
    *,
    project_code: str | None,
    organization_id: int | None,
    organization_contract_id: int | None,
):
    query = db.query(SiteLogActivityCatalog).options(
        joinedload(SiteLogActivityCatalog.project),
        joinedload(SiteLogActivityCatalog.organization),
        joinedload(SiteLogActivityCatalog.organization_contract).joinedload(OrganizationContract.block),
        joinedload(SiteLogActivityCatalog.pms_mapping).joinedload(SiteLogActivityPmsMapping.template),
        joinedload(SiteLogActivityCatalog.pms_mapping).selectinload(SiteLogActivityPmsMapping.steps),
    )
    if project_code:
        query = query.filter(SiteLogActivityCatalog.project_code == project_code)
    if organization_id is not None:
        query = query.filter(SiteLogActivityCatalog.organization_id == organization_id)
    if organization_contract_id is not None:
        query = query.filter(SiteLogActivityCatalog.organization_contract_id == organization_contract_id)
    return query


def _load_site_log_activity_catalog_payload(
    db: Session,
    *,
    project_code: str | None = None,
    organization_id: int | None = None,
    organization_contract_id: int | None = None,
    pms_status: str | None = None,
    pms_template_id: int | None = None,
    default_unit: str | None = None,
    default_location: str | None = None,
    reference_search: str | None = None,
) -> Dict[str, Any]:
    rows = (
        _activity_catalog_scope_query(
            db,
            project_code=project_code,
            organization_id=organization_id,
            organization_contract_id=organization_contract_id,
        )
        .order_by(
            SiteLogActivityCatalog.project_code.asc(),
            SiteLogActivityCatalog.sort_order.asc(),
            SiteLogActivityCatalog.activity_code.asc(),
            SiteLogActivityCatalog.id.asc(),
        )
        .all()
    )
    status_filter = _norm(pms_status).lower()
    template_filter = int(pms_template_id or 0) or None
    unit_filter = _norm(default_unit).lower()
    location_filter = _norm(default_location).lower()
    reference_filter = _norm(reference_search).lower()
    if any([status_filter, template_filter, unit_filter, location_filter, reference_filter]):
        filtered_rows: List[SiteLogActivityCatalog] = []
        for item in rows:
            pms = _serialize_activity_pms_summary(getattr(item, "pms_mapping", None))
            item_status = str(pms.get("pms_status") or "none").lower()
            if status_filter in {"none", "without", "without_pms"} and item_status != "none":
                continue
            if status_filter in {"mapped", "with", "with_pms"} and item_status == "none":
                continue
            if status_filter == "stale" and item_status != "stale":
                continue
            if template_filter and int(pms.get("pms_template_id") or 0) != template_filter:
                continue
            if unit_filter and unit_filter not in _norm(item.default_unit).lower():
                continue
            if location_filter and location_filter not in _norm(item.default_location).lower():
                continue
            reference_text = " ".join(
                [
                    _norm(getattr(getattr(item, "organization", None), "name", None)),
                    _norm(getattr(getattr(item, "organization_contract", None), "contract_number", None)),
                    _norm(getattr(getattr(item, "organization_contract", None), "subject", None)),
                    _upper(item.project_code),
                ]
            ).lower()
            if reference_filter and reference_filter not in reference_text:
                continue
            filtered_rows.append(item)
        rows = filtered_rows
    pms_summary = {"total": len(rows), "mapped": 0, "none": 0, "stale": 0}
    for item in rows:
        status = str(_serialize_activity_pms_summary(getattr(item, "pms_mapping", None)).get("pms_status") or "none")
        if status == "stale":
            pms_summary["stale"] += 1
        elif status == "mapped":
            pms_summary["mapped"] += 1
        else:
            pms_summary["none"] += 1
    projects = db.query(Project).filter(Project.is_active == True).order_by(Project.code.asc()).all()
    organizations = (
        db.query(Organization)
        .options(joinedload(Organization.contracts).joinedload(OrganizationContract.block))
        .filter(Organization.is_active == True)
        .order_by(Organization.name.asc())
        .all()
    )
    return {
        "items": [_serialize_site_log_activity_catalog_row(row) for row in rows],
        "pms_summary": pms_summary,
        "pms_templates": [_serialize_pms_template(row) for row in _load_pms_templates(db)],
        "projects": [{"code": row.code, "name": row.name_e or row.name_p or row.code} for row in projects],
        "organizations": [
            {
                "id": row.id,
                "name": row.name,
                "org_type": row.org_type,
                "contracts": [
                    _serialize_organization_contract(contract)
                    for contract in sorted(
                        list(row.contracts or []),
                        key=lambda item: (int(item.sort_order or 0), int(item.id or 0)),
                    )
                ],
            }
            for row in organizations
        ],
    }


_ACTIVITY_IMPORT_HEADER_ALIASES: Dict[str, str] = {
    "activitycode": "activity_code",
    "code": "activity_code",
    "کد": "activity_code",
    "کدفعالیت": "activity_code",
    "activitytitle": "activity_title",
    "title": "activity_title",
    "name": "activity_title",
    "شرح": "activity_title",
    "عنوان": "activity_title",
    "عنوانفعالیت": "activity_title",
    "defaultlocation": "default_location",
    "location": "default_location",
    "محل": "default_location",
    "محلپیشفرض": "default_location",
    "جبههکاری": "default_location",
    "defaultunit": "default_unit",
    "unit": "default_unit",
    "واحد": "default_unit",
    "واحدپیشفرض": "default_unit",
    "sortorder": "sort_order",
    "order": "sort_order",
    "sort": "sort_order",
    "ترتیب": "sort_order",
    "ردیف": "sort_order",
    "isactive": "is_active",
    "active": "is_active",
    "status": "is_active",
    "وضعیت": "is_active",
    "فعال": "is_active",
}


def _activity_import_header_key(value: Any) -> str:
    raw = _norm(value).lower().replace("ي", "ی").replace("ك", "ک")
    raw = raw.replace("\u200c", "").replace("\u200f", "").replace("\u200e", "")
    return re.sub(r"[\s_\-‐‑‒–—./()]+", "", raw)


def _activity_import_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _norm(value)


def _activity_import_bool(value: Any, default: bool = True) -> bool:
    text = _activity_import_cell_text(value).lower().replace("ي", "ی").replace("ك", "ک")
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return default
    if compact in {"0", "false", "no", "n", "inactive", "disabled", "غیرفعال", "غيرفعال", "نه", "خیر"}:
        return False
    if compact in {"1", "true", "yes", "y", "active", "enabled", "فعال", "بله"}:
        return True
    return default


def _activity_import_int(value: Any, default: int) -> int:
    text = _activity_import_cell_text(value)
    if not text:
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return default


def _resolve_activity_catalog_scope(
    db: Session,
    *,
    project_code: str | None,
    organization_id: int | None,
    organization_contract_id: int | None,
) -> tuple[str, int | None, OrganizationContract | None]:
    project_value = _check_optional_project_or_404(db, project_code)
    if not project_value:
        raise HTTPException(status_code=400, detail="project_code is required")
    organization_value = _check_optional_organization_or_404(db, organization_id)
    contract = None
    if organization_contract_id:
        contract = (
            db.query(OrganizationContract)
            .options(joinedload(OrganizationContract.block))
            .filter(OrganizationContract.id == int(organization_contract_id))
            .first()
        )
        if not contract:
            raise HTTPException(status_code=404, detail="Organization contract not found")
        if organization_value and int(contract.organization_id or 0) != int(organization_value):
            raise HTTPException(status_code=400, detail="Selected contract does not belong to the selected organization.")
        if organization_value is None:
            organization_value = int(contract.organization_id or 0) or None
    return project_value, organization_value, contract


def _find_activity_catalog_by_scope_code(
    db: Session,
    *,
    project_code: str,
    organization_id: int | None,
    organization_contract_id: int | None,
    activity_code: str,
) -> SiteLogActivityCatalog | None:
    query = db.query(SiteLogActivityCatalog).filter(
        SiteLogActivityCatalog.project_code == project_code,
        func.upper(SiteLogActivityCatalog.activity_code) == _upper(activity_code),
    )
    if organization_id is None:
        query = query.filter(SiteLogActivityCatalog.organization_id.is_(None))
    else:
        query = query.filter(SiteLogActivityCatalog.organization_id == int(organization_id))
    if organization_contract_id is None:
        query = query.filter(SiteLogActivityCatalog.organization_contract_id.is_(None))
    else:
        query = query.filter(SiteLogActivityCatalog.organization_contract_id == int(organization_contract_id))
    return query.first()


def _parse_site_log_activity_catalog_import(content: bytes, filename: str) -> List[Dict[str, Any]]:
    if not content:
        raise HTTPException(status_code=400, detail="Import file is empty")
    name = _norm(filename).lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported for activity catalog import")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is expected in the app image
        raise HTTPException(status_code=500, detail="Excel parser dependency is not installed") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}") from exc
    try:
        sheet = workbook.active
        header_map: Dict[int, str] = {}
        header_row_no = 0
        buffered_rows = list(sheet.iter_rows(values_only=True))
        for row_no, row in enumerate(buffered_rows, start=1):
            candidate: Dict[int, str] = {}
            for index, cell in enumerate(row):
                alias = _ACTIVITY_IMPORT_HEADER_ALIASES.get(_activity_import_header_key(cell))
                if alias:
                    candidate[index] = alias
            if "activity_code" in candidate.values() and "activity_title" in candidate.values():
                header_map = candidate
                header_row_no = row_no
                break
        if not header_map:
            raise HTTPException(
                status_code=400,
                detail="Excel header must include activity code and activity title columns.",
            )

        parsed: List[Dict[str, Any]] = []
        for row_no, row in enumerate(buffered_rows[header_row_no:], start=header_row_no + 1):
            if not any(_activity_import_cell_text(cell) for cell in row):
                continue
            item: Dict[str, Any] = {"row_no": row_no}
            for index, key in header_map.items():
                item[key] = _activity_import_cell_text(row[index] if index < len(row) else None)
            parsed.append(item)
        return parsed
    finally:
        workbook.close()


def _site_log_activity_catalog_template_bytes() -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover - dependency is expected in the app image
        raise HTTPException(status_code=500, detail="Excel writer dependency is not installed") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Activity Catalog"
    headers = ["کد فعالیت", "عنوان فعالیت", "محل پیش‌فرض", "واحد", "ترتیب", "وضعیت"]
    sheet.append(headers)
    sheet.append(["CV-101", "آرماتوربندی فونداسیون", "بلوک B", "تن", 10, "فعال"])
    sheet.append(["CV-118", "قالب بندی دیوار حائل", "ضلع شمالی", "مترمربع", 20, "فعال"])
    sheet.freeze_panes = "A2"
    widths = [18, 34, 24, 14, 12, 14]
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="0B3A7A")
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width
        cell = sheet.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="center")

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


_PMS_MAPPING_IMPORT_HEADER_ALIASES: Dict[str, str] = {
    "activitycode": "activity_code",
    "code": "activity_code",
    "pmsactivitycode": "activity_code",
    "activitytitle": "activity_title",
    "title": "activity_title",
    "pmstemplatecode": "pms_template_code",
    "templatecode": "pms_template_code",
    "pmscode": "pms_template_code",
}


def _parse_site_log_pms_mapping_import(content: bytes, filename: str) -> List[Dict[str, Any]]:
    if not content:
        raise HTTPException(status_code=400, detail="Import file is empty")
    name = _norm(filename).lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported for PMS mapping import")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Excel parser dependency is not installed") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}") from exc
    try:
        sheet = workbook.active
        buffered_rows = list(sheet.iter_rows(values_only=True))
        header_map: Dict[int, str] = {}
        header_row_no = 0
        for row_no, row in enumerate(buffered_rows, start=1):
            candidate: Dict[int, str] = {}
            for index, cell in enumerate(row):
                alias = _PMS_MAPPING_IMPORT_HEADER_ALIASES.get(_activity_import_header_key(cell))
                if alias:
                    candidate[index] = alias
            if "activity_code" in candidate.values() and "pms_template_code" in candidate.values():
                header_map = candidate
                header_row_no = row_no
                break
        if not header_map:
            raise HTTPException(
                status_code=400,
                detail="Excel header must include activity_code and pms_template_code columns.",
            )
        parsed: List[Dict[str, Any]] = []
        for row_no, row in enumerate(buffered_rows[header_row_no:], start=header_row_no + 1):
            if not any(_activity_import_cell_text(cell) for cell in row):
                continue
            item: Dict[str, Any] = {"row_no": row_no}
            for index, key in header_map.items():
                item[key] = _activity_import_cell_text(row[index] if index < len(row) else None)
            parsed.append(item)
        return parsed
    finally:
        workbook.close()


def _site_log_pms_mapping_workbook_bytes(rows: List[Dict[str, Any]] | None = None) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Excel writer dependency is not installed") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Activity PMS Mapping"
    headers = [
        "activity_code",
        "activity_title",
        "default_location",
        "default_unit",
        "reference",
        "pms_template_code",
        "pms_template_title",
        "pms_status",
        "pms_version",
    ]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append(
                [
                    row.get("activity_code") or "",
                    row.get("activity_title") or "",
                    row.get("default_location") or "",
                    row.get("default_unit") or "",
                    row.get("organization_name") or row.get("contract_subject") or row.get("project_code") or "",
                    row.get("pms_template_code") or "",
                    row.get("pms_template_title") or "",
                    row.get("pms_status") or "none",
                    row.get("pms_snapshot_version") or "",
                ]
            )
    else:
        sheet.append(["CV-101", "Concrete work", "Block A", "m3", "Contract A", "CONC", "Concrete PMS", "", ""])
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="0B3A7A")
    widths = [18, 34, 24, 14, 24, 22, 32, 16, 14]
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width
        cell = sheet.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="center")
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


VALID_WORKFLOW_ITEM_TYPES = {"RFI", "NCR", "WORK_INSTRUCTION"}


def _normalize_workflow_item_type_or_400(value: Optional[str]) -> str:
    item_type = _upper(value)
    if item_type not in VALID_WORKFLOW_ITEM_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid workflow item_type: {value}")
    return item_type


_PKG_SEQ_RE = re.compile(r"(\d{1,3})$")


def _extract_package_sequence(code: Any, discipline_code: Any = "") -> Optional[int]:
    raw = _upper(code)
    if not raw:
        return None
    dcode = _upper(discipline_code)
    candidate = raw
    if dcode and candidate.startswith(dcode):
        candidate = candidate[len(dcode):]

    if candidate.isdigit():
        seq = int(candidate)
    else:
        match = _PKG_SEQ_RE.search(candidate)
        if not match:
            return None
        seq = int(match.group(1))

    if seq < 1 or seq > 99:
        return None
    return seq


def _normalize_package_code_by_discipline(code: Any, discipline_code: Any = "") -> str:
    seq = _extract_package_sequence(code, discipline_code)
    if seq is None:
        return _upper(code)
    return f"{seq:02d}"


def _next_package_code_for_discipline(db: Session, discipline_code: str) -> str:
    dcode = _upper(discipline_code)
    used: set[int] = set()
    rows = db.query(Package.package_code).filter(Package.discipline_code == dcode).all()
    for (pkg_code,) in rows:
        seq = _extract_package_sequence(pkg_code, dcode)
        if seq is not None:
            used.add(seq)

    for seq in range(1, 100):
        if seq not in used:
            return f"{seq:02d}"

    raise HTTPException(
        status_code=400,
        detail=f"No available 2-digit package code left for discipline {dcode}.",
    )


def _normalize_permission_category_or_400(value: Optional[str]) -> str:
    raw = _norm(value).lower()
    if not raw:
        return DEFAULT_PERMISSION_CATEGORY
    if raw in PERMISSION_CATEGORIES:
        return raw
    if raw in ALL_ORG_TYPES:
        mapped = normalize_permission_category(raw)
        if mapped in PERMISSION_CATEGORIES:
            return mapped
    raise HTTPException(status_code=400, detail=f"Invalid category: {value}")


def _normalize_org_type_or_400(value: Optional[str]) -> str:
    key = _norm(value).lower()
    if not key:
        raise HTTPException(status_code=400, detail="org_type is required")
    if key not in ALL_ORG_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid org_type: {value}")
    return key


def _count(db: Session, model) -> int:
    return db.query(model).count()


def _safe_json_dump(payload: Any) -> str:
    return json.dumps(redact_secrets(payload), ensure_ascii=False, default=str)


def _storage_integrations_openproject_token_source(integrations: Dict[str, Any]) -> str:
    openproject = dict((integrations or {}).get("openproject") or {})
    env_token = str(settings.OPENPROJECT_API_TOKEN or "").strip()
    settings_token = str(openproject.get("api_token") or "").strip()
    if env_token:
        return "env"
    if settings_token:
        return "settings"
    return "none"


def _storage_integrations_nextcloud_credential_source(integrations: Dict[str, Any]) -> str:
    nextcloud = dict((integrations or {}).get("nextcloud") or {})
    env_username = str(settings.NEXTCLOUD_USERNAME or "").strip()
    env_password = str(settings.NEXTCLOUD_APP_PASSWORD or "").strip()
    settings_username = str(nextcloud.get("username") or "").strip()
    settings_password = str(nextcloud.get("app_password") or "").strip()
    if env_username and env_password:
        return "env"
    if settings_username and settings_password:
        return "settings"
    return "none"


def _path_under_storage_root(path_value: str, root_value: str) -> bool:
    raw_path = _norm(path_value)
    raw_root = _norm(root_value)
    if not raw_path or not raw_root:
        return False
    return StorageManager._path_is_under_root_value(raw_path, raw_root)


def _path_under_nextcloud_root(path_value: str, root_value: str) -> bool:
    raw_path = _norm(path_value)
    raw_root = _norm(root_value)
    if not raw_path or not raw_root:
        return False
    return StorageManager._is_under_remote_root(raw_path, raw_root)


def _storage_primary_runtime_payload(
    db: Session | None,
    integrations: Dict[str, Any],
) -> Dict[str, Any]:
    selected_provider = resolve_primary_storage_provider(integrations)
    nextcloud_runtime = resolve_nextcloud_runtime(integrations)
    mode = _norm(nextcloud_runtime.get("mode")).lower() or "mount"
    mount_root = _norm(nextcloud_runtime.get("local_mount_root_effective"))
    root_path = _norm(nextcloud_runtime.get("root_path")) or "/"
    nextcloud_ready = bool(
        nextcloud_runtime.get("enabled")
        and _norm(nextcloud_runtime.get("base_url"))
        and _norm(nextcloud_runtime.get("username"))
        and _norm(nextcloud_runtime.get("app_password"))
        and (root_path if mode == "webdav" else mount_root)
    )
    mdr_path = _kv_get_value(db, STORAGE_PATH_MDR_KEY, DEFAULT_MDR_STORAGE_PATH) if db else ""
    corr_path = _kv_get_value(
        db,
        STORAGE_PATH_CORRESPONDENCE_KEY,
        DEFAULT_CORRESPONDENCE_STORAGE_PATH,
    ) if db else ""
    if mode == "webdav":
        mdr_on_target = bool(db and root_path and _path_under_nextcloud_root(mdr_path, root_path))
        corr_on_target = bool(db and root_path and _path_under_nextcloud_root(corr_path, root_path))
    else:
        mdr_on_target = bool(db and mount_root and _path_under_storage_root(mdr_path, mount_root))
        corr_on_target = bool(db and mount_root and _path_under_storage_root(corr_path, mount_root))

    effective_provider = "local"
    status = "local_active"
    status_message = "Primary storage is local filesystem / UNC path."
    if selected_provider == "nextcloud":
        if nextcloud_ready and mdr_on_target and corr_on_target:
            effective_provider = "nextcloud"
            status = "ready"
            status_message = (
                "Primary storage is Nextcloud via direct WebDAV API."
                if mode == "webdav"
                else "Primary storage is Nextcloud via mounted local/UNC path."
            )
        elif not nextcloud_ready:
            status = "misconfigured"
            status_message = (
                "Nextcloud primary storage needs enabled integration, credentials, and Root Path in WebDAV mode."
                if mode == "webdav"
                else "Nextcloud primary storage needs enabled integration, credentials, and Local Mount Root."
            )
        else:
            status = "paths_pending"
            status_message = (
                "Move MDR and Correspondence paths under the Nextcloud Root Path to activate WebDAV primary storage."
                if mode == "webdav"
                else "Move MDR and Correspondence paths under the Nextcloud Local Mount Root to activate primary storage."
            )

    return {
        "provider": selected_provider,
        "effective_provider": effective_provider,
        "status": status,
        "status_message": status_message,
        "nextcloud_mode": mode,
        "nextcloud_ready": nextcloud_ready,
        "nextcloud_root_path": root_path,
        "local_mount_root_effective": mount_root,
        "local_mount_root_source": str(nextcloud_runtime.get("local_mount_root_source") or "none"),
        "mdr_path_under_mount": mdr_on_target if mode != "webdav" else False,
        "correspondence_path_under_mount": corr_on_target if mode != "webdav" else False,
        "mdr_path_under_root": mdr_on_target if mode == "webdav" else False,
        "correspondence_path_under_root": corr_on_target if mode == "webdav" else False,
    }


def _masked_storage_integrations_payload(
    integrations: Dict[str, Any],
    db: Session | None = None,
) -> Dict[str, Any]:
    masked = dict(integrations or {})
    masked["primary"] = _storage_primary_runtime_payload(db, integrations)

    openproject = dict(masked.get("openproject", {}))
    if not str(openproject.get("default_work_package_id") or "").strip():
        openproject["default_work_package_id"] = str(openproject.get("default_project_id") or "").strip()
    openproject.pop("default_project_id", None)
    runtime = resolve_openproject_runtime(integrations)
    token_source = _storage_integrations_openproject_token_source(integrations)
    openproject["token_source"] = token_source
    openproject["api_token_configured"] = token_source in {"env", "settings"}
    openproject["skip_ssl_verify"] = bool(runtime.get("skip_ssl_verify_effective"))
    openproject["ssl_source"] = str(runtime.get("ssl_source") or "env_default")
    openproject["ssl_force_active"] = bool(runtime.get("ssl_force_active"))
    openproject.pop("api_token", None)
    masked["openproject"] = openproject

    gdrive = dict(masked.get("google_drive", {}))
    oauth_client_id = str(gdrive.get("oauth_client_id") or "").strip()
    oauth_client_secret = str(gdrive.get("oauth_client_secret") or "").strip()
    oauth_refresh_token = str(gdrive.get("oauth_refresh_token") or "").strip()
    gdrive["oauth_configured"] = bool(oauth_client_id and oauth_client_secret and oauth_refresh_token)
    if str(settings.GDRIVE_SERVICE_ACCOUNT_JSON or "").strip():
        gdrive["service_account_configured"] = True
    gdrive.pop("oauth_client_secret", None)
    gdrive.pop("oauth_refresh_token", None)
    masked["google_drive"] = gdrive

    mirror = dict(masked.get("mirror") or {})
    provider = str(mirror.get("provider") or "").strip().lower()
    if provider not in {"none", "google_drive", "nextcloud"}:
        provider = "none"
    mirror["provider"] = provider
    masked["mirror"] = mirror

    nextcloud = dict(masked.get("nextcloud") or {})
    nextcloud_runtime = resolve_nextcloud_runtime(integrations)
    credential_source = _storage_integrations_nextcloud_credential_source(integrations)
    nextcloud["credential_source"] = credential_source
    nextcloud["credentials_configured"] = credential_source in {"env", "settings"}
    nextcloud["skip_ssl_verify"] = bool(nextcloud_runtime.get("skip_ssl_verify_effective"))
    nextcloud["ssl_source"] = str(nextcloud_runtime.get("ssl_source") or "env_default")
    nextcloud["ssl_force_active"] = bool(nextcloud_runtime.get("ssl_force_active"))
    nextcloud["local_mount_root"] = str(nextcloud_runtime.get("local_mount_root_effective") or "")
    nextcloud["local_mount_root_source"] = str(
        nextcloud_runtime.get("local_mount_root_source") or "none"
    )
    nextcloud["local_mount_root_configured"] = bool(
        nextcloud_runtime.get("local_mount_root_configured")
    )
    nextcloud["public_share_password_configured"] = bool(
        str((integrations.get("nextcloud") or {}).get("public_share_password") or "").strip()
    )
    nextcloud["public_share_password_required"] = bool(
        (integrations.get("nextcloud") or {}).get("public_share_password_required", True)
    )
    nextcloud.pop("app_password", None)
    nextcloud.pop("public_share_password", None)
    masked["nextcloud"] = nextcloud

    return redact_secrets(masked)


def _is_valid_http_url(value: str) -> bool:
    raw = _norm(value)
    if not raw:
        return False
    if raw.startswith("/"):
        return True
    try:
        parsed = urlparse(raw)
    except Exception:
        return False
    if parsed.scheme not in {"http", "https"}:
        return False
    return bool(parsed.netloc)


def _runtime_secret_key_or_500() -> str:
    value = _norm(settings.SECRET_KEY)
    if not value:
        raise HTTPException(status_code=500, detail="SECRET_KEY is not configured.")
    return value


def _masked_bim_revit_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    raw = dict(payload or {})
    secret = _norm(raw.get("plugin_secret_encrypted"))
    raw.pop("plugin_secret_encrypted", None)
    masked = redact_secrets(raw)
    masked["has_secret"] = bool(secret)
    return masked


def _as_dict(obj: Any, fields: List[str]) -> Dict[str, Any] | None:
    if obj is None:
        return None
    return {field: getattr(obj, field, None) for field in fields}


def _audit_log(
    db: Session,
    *,
    actor: DbUser,
    action: str,
    target_type: str,
    target_key: str | None,
    before: Any,
    after: Any,
) -> None:
    db.add(
        SettingsAuditLog(
            action=action,
            target_type=target_type,
            target_key=_norm(target_key) or None,
            actor_user_id=actor.id,
            actor_email=actor.email,
            actor_name=actor.full_name,
            before_json=_safe_json_dump(before),
            after_json=_safe_json_dump(after),
            created_at=datetime.utcnow(),
        )
    )


def _effective_scope_values(role_values: List[str], user_values: List[str]) -> tuple[List[str], bool]:
    role_set = set(role_values)
    user_set = set(user_values)
    if role_set and user_set:
        return sorted(role_set & user_set), True
    if role_set:
        return sorted(role_set), True
    if user_set:
        return sorted(user_set), True
    return [], False


def _build_access_report_items(
    db: Session,
    *,
    project_code: Optional[str],
    discipline_code: Optional[str],
    include_inactive: bool,
    include_denied: bool,
) -> List[Dict[str, Any]]:
    project = _upper(project_code)
    discipline = _upper(discipline_code)
    user_scope = _load_user_scope_rules(db)
    role_scope_cache: Dict[str, Dict[str, Dict[str, List[str]]]] = {}

    def _scope_for_category(category_key: str) -> Dict[str, Dict[str, List[str]]]:
        cache_key = _normalize_permission_category_or_400(category_key)
        if cache_key not in role_scope_cache:
            role_scope_cache[cache_key] = _load_scope_rules(db, category=cache_key)
        return role_scope_cache[cache_key]

    q = db.query(DbUser).options(joinedload(DbUser.organization)).order_by(DbUser.id)
    if not include_inactive:
        q = q.filter(DbUser.is_active == True)
    users = q.all()

    items: List[Dict[str, Any]] = []
    for user in users:
        access = resolve_effective_access(user)
        role = access.effective_role
        category = access.permission_category
        project_allowed = True
        discipline_allowed = True
        has_access = True
        effective_projects: List[str] = []
        effective_disciplines: List[str] = []
        project_restricted = False
        discipline_restricted = False

        if not access.full_access:
            role_scope = _scope_for_category(category)
            role_projects = role_scope.get(role, {}).get("projects", [])
            role_disciplines = role_scope.get(role, {}).get("disciplines", [])
            user_projects = user_scope.get(str(user.id), {}).get("projects", [])
            user_disciplines = user_scope.get(str(user.id), {}).get("disciplines", [])

            effective_projects, project_restricted = _effective_scope_values(role_projects, user_projects)
            effective_disciplines, discipline_restricted = _effective_scope_values(role_disciplines, user_disciplines)

            if project and project_restricted:
                project_allowed = project in effective_projects
            if discipline and discipline_restricted:
                discipline_allowed = discipline in effective_disciplines
            has_access = bool(project_allowed and discipline_allowed)

        if has_access or include_denied:
            items.append(
                {
                    "user_id": user.id,
                    "email": user.email,
                    "full_name": user.full_name,
                    "role": role,
                    "organization_role": getattr(user, "organization_role", None),
                    "category": category,
                    "is_system_admin": access.is_system_admin,
                    "is_active": user.is_active,
                    "has_access": has_access,
                    "project_allowed": project_allowed,
                    "discipline_allowed": discipline_allowed,
                    "projects_restricted": project_restricted,
                    "disciplines_restricted": discipline_restricted,
                    "effective_projects": effective_projects,
                    "effective_disciplines": effective_disciplines,
                }
            )
    return items


# -----------------------------
# KV Helpers
# -----------------------------
def _kv_set(db: Session, key: str, value: str) -> None:
    kv = db.query(SettingsKV).filter(SettingsKV.key == key).first()
    if kv:
        kv.value = value
        kv.updated_at = datetime.utcnow()
    else:
        kv = SettingsKV(key=key, value=value, updated_at=datetime.utcnow())
        db.add(kv)

def _kv_get_all(db: Session, *, include_legacy: bool = False) -> List[Dict[str, Any]]:
    rows = db.query(SettingsKV).order_by(SettingsKV.key).all()
    if not include_legacy:
        rows = [
            r for r in rows
            if r.key not in KV_RESERVED_KEYS and not str(r.key or "").startswith("legacy.")
        ]
    return [{"key": r.key, "value": r.value, "updated_at": r.updated_at.isoformat()} for r in rows]


def _kv_get_value(db: Session, key: str, default: str = "") -> str:
    row = db.query(SettingsKV).filter(SettingsKV.key == key).first()
    if not row:
        return default
    value = _norm(row.value)
    return value if value else default


def _is_allowed_kv_write_key(key: str) -> bool:
    if key in KV_RESERVED_KEYS:
        return False
    if key.startswith("legacy."):
        return False
    return any(key.startswith(prefix) for prefix in KV_ALLOWED_WRITE_PREFIXES)


def _permission_keys() -> List[str]:
    return permission_keys()


def _permission_meta() -> List[Dict[str, Any]]:
    return permission_meta_list()


def _feature_catalog() -> List[Dict[str, Any]]:
    return feature_catalog()


def _default_permission_matrix(category: Optional[str] = None) -> Dict[str, Dict[str, bool]]:
    return default_permission_matrix_for_category(category)


def _matrix_role_labels() -> Dict[str, str]:
    return {
        "manager": "سرپرست",
        "dcc": "کنترل مدارک (DCC)",
        "project_control": "کنترل پروژه",
        "user": "کاربر عادی",
        "viewer": "مشاهده‌گر",
    }


def _permission_category_labels() -> Dict[str, str]:
    return {
        "consultant": "مشاور",
        "contractor": "پیمانکار",
        "employer": "کارفرما",
        "dcc": "DCC",
    }


def _normalize_permission_matrix(raw: Any, category: Optional[str] = None) -> Dict[str, Dict[str, bool]]:
    normalized = _default_permission_matrix(category)
    perms = _permission_keys()

    if not isinstance(raw, dict):
        return normalized

    for role in MATRIX_ROLES:
        role_data = raw.get(role)
        if not isinstance(role_data, dict):
            continue

        for perm in perms:
            if perm in role_data:
                normalized[role][perm] = bool(role_data.get(perm))

    return normalized


def _load_permission_matrix(
    db: Session,
    *,
    category: Optional[str] = None,
) -> Dict[str, Dict[str, bool]]:
    rows = []

    if category is not None:
        category_key = _normalize_permission_category_or_400(category)
        matrix = _default_permission_matrix(category_key)
        perms = _permission_keys()
        rows = (
            db.query(RoleCategoryPermission)
            .filter(RoleCategoryPermission.category == category_key)
            .all()
        )
    else:
        matrix = _default_permission_matrix()
        perms = _permission_keys()
        rows = db.query(RolePermission).all()

    for row in rows:
        role = _norm(getattr(row, "role", None)).lower()
        perm = _norm(getattr(row, "permission", None))
        if role not in MATRIX_ROLES or perm not in perms:
            continue
        matrix[role][perm] = bool(getattr(row, "allowed", False))
    return matrix


def _default_scope_rules() -> Dict[str, Dict[str, List[str]]]:
    # لیست خالی یعنی محدودیت اعمال نشود (دسترسی به همه)
    return {
        role: {
            "projects": [],
            "disciplines": [],
        }
        for role in MATRIX_ROLES
    }


def _normalize_scope_values(values: Any, *, upper: bool = True) -> List[str]:
    if not isinstance(values, list):
        return []
    normalized: List[str] = []
    for raw in values:
        value = _norm(raw)
        if not value:
            continue
        normalized.append(value.upper() if upper else value)
    return sorted(set(normalized))


def _normalize_scope_rules(raw: Any) -> Dict[str, Dict[str, List[str]]]:
    normalized = _default_scope_rules()
    if not isinstance(raw, dict):
        return normalized

    for role in MATRIX_ROLES:
        role_data = raw.get(role)
        if not isinstance(role_data, dict):
            continue

        normalized[role]["projects"] = _normalize_scope_values(role_data.get("projects"))
        normalized[role]["disciplines"] = _normalize_scope_values(role_data.get("disciplines"))

    return normalized


def _load_scope_rules(
    db: Session,
    *,
    category: Optional[str] = None,
) -> Dict[str, Dict[str, List[str]]]:
    normalized = _default_scope_rules()

    role_projects: List[tuple[str, str]] = []
    role_disciplines: List[tuple[str, str]] = []

    if category is not None:
        category_key = _normalize_permission_category_or_400(category)
        has_category_context = (
            db.query(RoleCategoryPermission.id)
            .filter(RoleCategoryPermission.category == category_key)
            .first()
            is not None
        )
        if has_category_context:
            role_projects = (
                db.query(RoleCategoryProjectScope.role, RoleCategoryProjectScope.project_code)
                .filter(RoleCategoryProjectScope.category == category_key)
                .all()
            )
            role_disciplines = (
                db.query(RoleCategoryDisciplineScope.role, RoleCategoryDisciplineScope.discipline_code)
                .filter(RoleCategoryDisciplineScope.category == category_key)
                .all()
            )

    if not role_projects and not role_disciplines:
        role_projects = db.query(RoleProjectScope.role, RoleProjectScope.project_code).all()
        role_disciplines = db.query(RoleDisciplineScope.role, RoleDisciplineScope.discipline_code).all()

    for role, project_code in role_projects:
        role_key = _norm(role).lower()
        if role_key in normalized and project_code:
            normalized[role_key]["projects"].append(_upper(project_code))

    for role, discipline_code in role_disciplines:
        role_key = _norm(role).lower()
        if role_key in normalized and discipline_code:
            normalized[role_key]["disciplines"].append(_upper(discipline_code))

    for role in MATRIX_ROLES:
        normalized[role]["projects"] = sorted(set(normalized[role]["projects"]))
        normalized[role]["disciplines"] = sorted(set(normalized[role]["disciplines"]))
    return normalized


def _default_user_scope_rules() -> Dict[str, Dict[str, List[str]]]:
    return {}


def _normalize_user_scope_rules(raw: Any) -> Dict[str, Dict[str, List[str]]]:
    if not isinstance(raw, dict):
        return _default_user_scope_rules()

    normalized: Dict[str, Dict[str, List[str]]] = {}
    for user_id, values in raw.items():
        key = _norm(user_id)
        if not key:
            continue
        if not isinstance(values, dict):
            continue
        normalized[key] = {
            "projects": _normalize_scope_values(values.get("projects")),
            "disciplines": _normalize_scope_values(values.get("disciplines")),
        }
    return normalized


def _load_user_scope_rules(db: Session) -> Dict[str, Dict[str, List[str]]]:
    normalized: Dict[str, Dict[str, List[str]]] = {}

    project_rows = db.query(UserProjectScope.user_id, UserProjectScope.project_code).all()
    for user_id, project_code in project_rows:
        key = str(user_id)
        normalized.setdefault(key, {"projects": [], "disciplines": []})
        if project_code:
            normalized[key]["projects"].append(_upper(project_code))

    discipline_rows = db.query(UserDisciplineScope.user_id, UserDisciplineScope.discipline_code).all()
    for user_id, discipline_code in discipline_rows:
        key = str(user_id)
        normalized.setdefault(key, {"projects": [], "disciplines": []})
        if discipline_code:
            normalized[key]["disciplines"].append(_upper(discipline_code))

    for key, values in normalized.items():
        values["projects"] = sorted(set(values["projects"]))
        values["disciplines"] = sorted(set(values["disciplines"]))
    return normalized


def _organization_sort_key(row: Organization) -> tuple[str, str, int]:
    return (_norm(row.name).lower(), _norm(row.code).lower(), int(row.id))


def _flatten_organizations(rows: List[Organization]) -> List[tuple[Organization, int]]:
    rows_by_id: Dict[int, Organization] = {int(row.id): row for row in rows}
    children_map: Dict[Optional[int], List[Organization]] = {}
    for row in rows:
        parent_id = int(row.parent_id) if row.parent_id is not None and int(row.parent_id) in rows_by_id else None
        children_map.setdefault(parent_id, []).append(row)

    for items in children_map.values():
        items.sort(key=_organization_sort_key)

    ordered: List[tuple[Organization, int]] = []

    def _walk(parent_id: Optional[int], depth: int, trail: set[int]) -> None:
        for item in children_map.get(parent_id, []):
            item_id = int(item.id)
            if item_id in trail:
                continue
            ordered.append((item, depth))
            _walk(item_id, depth + 1, trail | {item_id})

    _walk(None, 0, set())

    if len(ordered) < len(rows):
        visited = {int(row.id) for row, _ in ordered}
        leftovers = [row for row in rows if int(row.id) not in visited]
        leftovers.sort(key=_organization_sort_key)
        for row in leftovers:
            ordered.append((row, 0))

    return ordered


def _build_organization_tree(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    nodes = {
        int(item["id"]): {
            **item,
            "children": [],
        }
        for item in items
    }
    roots: List[Dict[str, Any]] = []
    for item in items:
        item_id = int(item["id"])
        parent_id = item.get("parent_id")
        node = nodes[item_id]
        if parent_id is not None and int(parent_id) in nodes:
            nodes[int(parent_id)]["children"].append(node)
        else:
            roots.append(node)
    return roots


def _organization_contract_block_label(block: Block | None) -> Optional[str]:
    if block is None:
        return None
    primary = "/".join(part for part in [_norm(block.project_code), _norm(block.code)] if part)
    name = _norm(getattr(block, "name_p", None)) or _norm(getattr(block, "name_e", None))
    if not primary:
        return name or None
    return f"{primary} - {name}" if name else primary


def _serialize_organization_contract(row: OrganizationContract) -> Dict[str, Any]:
    block = row.block
    return {
        "id": row.id,
        "contract_number": row.contract_number,
        "subject": row.subject,
        "block_id": row.block_id,
        "block_code": block.code if block else None,
        "block_project_code": block.project_code if block else None,
        "block_label": _organization_contract_block_label(block),
        "sort_order": row.sort_order,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


def _serialize_organization_snapshot(row: Organization | None) -> Dict[str, Any] | None:
    if row is None:
        return None
    contracts = [
        _serialize_organization_contract(contract)
        for contract in sorted(
            list(row.contracts or []),
            key=lambda item: (int(item.sort_order or 0), int(item.id or 0)),
        )
    ]
    return {
        "id": row.id,
        "code": row.code,
        "name": row.name,
        "org_type": row.org_type,
        "parent_id": row.parent_id,
        "is_active": bool(row.is_active),
        "contracts": contracts,
        "contracts_count": len(contracts),
    }


def _serialize_organization_list_item(
    row: Organization,
    *,
    parent: Organization | None,
    depth: int,
    users_count: Dict[int, int],
    children_count: Dict[int, int],
) -> Dict[str, Any]:
    snapshot = _serialize_organization_snapshot(row) or {}
    snapshot.update(
        {
            "parent_code": parent.code if parent else None,
            "parent_name": parent.name if parent else None,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "users_count": users_count.get(int(row.id), 0),
            "children_count": children_count.get(int(row.id), 0),
            "depth": int(depth),
        }
    )
    return snapshot


def _normalize_organization_contracts_or_400(
    payload_contracts: Optional[List[OrganizationContractIn]],
    db: Session,
) -> tuple[List[Dict[str, Any]], Dict[int, Block]]:
    normalized: List[Dict[str, Any]] = []
    block_ids: set[int] = set()

    for index, item in enumerate(payload_contracts or []):
        contract_number = _norm(item.contract_number)
        subject = _norm(item.subject)
        block_id = int(item.block_id) if item.block_id is not None else None

        if not contract_number and not subject and block_id is None:
            continue
        if not contract_number:
            raise HTTPException(
                status_code=400,
                detail=f"Contract #{index + 1}: contract_number is required",
            )
        if not subject:
            raise HTTPException(
                status_code=400,
                detail=f"Contract #{index + 1}: subject is required",
            )

        normalized.append(
            {
                "id": int(item.id) if item.id is not None else None,
                "contract_number": contract_number,
                "subject": subject,
                "block_id": block_id,
                "sort_order": index,
            }
        )
        if block_id is not None:
            block_ids.add(block_id)

    block_lookup: Dict[int, Block] = {}
    if block_ids:
        block_lookup = {
            int(block.id): block
            for block in db.query(Block).filter(Block.id.in_(sorted(block_ids))).all()
        }
        missing_ids = sorted(block_ids - set(block_lookup))
        if missing_ids:
            raise HTTPException(
                status_code=400,
                detail=f"Block not found for contract(s): {', '.join(str(item) for item in missing_ids)}",
            )

    return normalized, block_lookup


def _sync_organization_contracts(
    row: Organization,
    *,
    contracts_payload: List[Dict[str, Any]],
    block_lookup: Dict[int, Block],
) -> None:
    existing_by_id = {
        int(contract.id): contract
        for contract in list(row.contracts or [])
        if contract.id is not None
    }
    next_contracts: List[OrganizationContract] = []

    for item in contracts_payload:
        contract_id = item.get("id")
        contract = existing_by_id.get(int(contract_id)) if contract_id is not None else None
        if contract_id is not None and contract is None:
            raise HTTPException(
                status_code=400,
                detail=f"Organization contract not found: {contract_id}",
            )
        if contract is None:
            contract = OrganizationContract()

        contract.contract_number = item["contract_number"]
        contract.subject = item["subject"]
        contract.block_id = item["block_id"]
        contract.block = block_lookup.get(item["block_id"]) if item["block_id"] is not None else None
        contract.sort_order = int(item["sort_order"])
        next_contracts.append(contract)

    row.contracts[:] = next_contracts


# -----------------------------
# Endpoints
# -----------------------------

@router.get("/overview")
def overview(db: Session = Depends(get_db)):
    """
    نمایش آمار کلی سیستم
    """
    return {
        "ok": True,
        "app": settings.APP_NAME,
        "db": _db_overview_payload(),
        "counts": {
            "projects": _count(db, Project),
            "organizations": _count(db, Organization),
            "phases": _count(db, Phase),
            "disciplines": _count(db, Discipline),
            "packages": _count(db, Package),
            "levels": _count(db, Level),
            "blocks": _count(db, Block),
            "mdr_categories": _count(db, MdrCategory),
            "settings_kv": _count(db, SettingsKV),
            "statuses": _count(db, DocStatus),
            "role_permissions": _count(db, RolePermission),
            "role_category_permissions": _count(db, RoleCategoryPermission),
            "role_project_scopes": _count(db, RoleProjectScope),
            "role_category_project_scopes": _count(db, RoleCategoryProjectScope),
            "role_discipline_scopes": _count(db, RoleDisciplineScope),
            "role_category_discipline_scopes": _count(db, RoleCategoryDisciplineScope),
            "user_project_scopes": _count(db, UserProjectScope),
            "user_discipline_scopes": _count(db, UserDisciplineScope),
            "settings_audit_logs": _count(db, SettingsAuditLog),
            "correspondences": _count(db, Correspondence),
            "correspondence_actions": _count(db, CorrespondenceAction),
            "correspondence_attachments": _count(db, CorrespondenceAttachment),
            "issuing_entities": _count(db, IssuingEntity),
            "correspondence_categories": _count(db, CorrespondenceCategory),
            "correspondence_departments": _count(db, CorrespondenceDepartment),
        },
    }

@router.post("/seed")
def seed_all(
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("settings:update")),
):
    """
    فراخوانی سرویس Seed برای خواندن اکسل Master Data و آپدیت دیتابیس.
    همراه با لاگ‌گیری دقیق خطاها.
    """
    print(f"[INFO] Seeding triggered via API")
    
    try:
        # بررسی وجود فایل قبل از اجرا
        excel_path = settings.BASE_DIR / "data_sources" / "master_data.xlsx"
        if not excel_path.exists():
            print(f"[ERROR] Excel file not found at: {excel_path}")
            return {"ok": False, "message": f"File not found: {excel_path}"}

        stats = seed_service.seed_from_excel(db)
        
        if "error" in stats:
             print(f"[ERROR] Seed service returned error: {stats['error']}")
             return {"ok": False, "message": stats["error"]}
             
        db.commit()
        print("[SUCCESS] Database seeded successfully.")
        return {"ok": True, "stats": stats, "source": "Excel"}
        
    except Exception as e:
        db.rollback()
        print("!!!!!!!!!!!!!! SEEDING ERROR !!!!!!!!!!!!!!")
        traceback.print_exc()  # چاپ کامل متن خطا در ترمینال
        raise HTTPException(status_code=500, detail=f"Seeding failed: {str(e)}")


# --- Organizations ---
@router.get("/organizations")
def list_organizations_settings(
    include_inactive: bool = Query(default=False),
    tree: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("organizations:read")),
):
    all_rows = (
        db.query(Organization)
        .options(joinedload(Organization.contracts).joinedload(OrganizationContract.block))
        .order_by(Organization.id.asc())
        .all()
    )
    parent_lookup = {int(row.id): row for row in all_rows}
    rows = [row for row in all_rows if include_inactive or bool(row.is_active)]

    users_count = {
        int(org_id): int(count)
        for org_id, count in (
            db.query(DbUser.organization_id, func.count(DbUser.id))
            .group_by(DbUser.organization_id)
            .all()
        )
        if org_id is not None
    }
    children_count = {
        int(parent_id): int(count)
        for parent_id, count in (
            db.query(Organization.parent_id, func.count(Organization.id))
            .group_by(Organization.parent_id)
            .all()
        )
        if parent_id is not None
    }

    flat_rows = _flatten_organizations(rows)
    items: List[Dict[str, Any]] = []
    for row, depth in flat_rows:
        parent = parent_lookup.get(int(row.parent_id)) if row.parent_id is not None else None
        items.append(
            _serialize_organization_list_item(
                row,
                parent=parent,
                depth=depth,
                users_count=users_count,
                children_count=children_count,
            )
        )

    payload: Dict[str, Any] = {
        "ok": True,
        "count": len(items),
        "items": items,
    }
    if tree:
        payload["tree"] = _build_organization_tree(items)
    return payload


@router.post("/organizations/upsert")
def upsert_organization_settings(
    payload: OrganizationIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("organizations:manage")),
):
    code = _upper(payload.code)
    name = _norm(payload.name)
    org_type = _normalize_org_type_or_400(payload.org_type)
    parent_id = int(payload.parent_id) if payload.parent_id is not None else None
    payload_fields = payload.model_fields_set if hasattr(payload, "model_fields_set") else getattr(payload, "__fields_set__", set())
    contracts_provided = "contracts" in payload_fields
    contracts_payload: Optional[List[Dict[str, Any]]] = None
    block_lookup: Dict[int, Block] = {}

    if not code:
        raise HTTPException(status_code=400, detail="code is required")
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    if contracts_provided:
        contracts_payload, block_lookup = _normalize_organization_contracts_or_400(payload.contracts, db)

    row = None
    if payload.id is not None:
        row = (
            db.query(Organization)
            .options(joinedload(Organization.contracts).joinedload(OrganizationContract.block))
            .filter(Organization.id == int(payload.id))
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail=f"Organization not found: {payload.id}")
    if row is None:
        row = (
            db.query(Organization)
            .options(joinedload(Organization.contracts).joinedload(OrganizationContract.block))
            .filter(Organization.code == code)
            .first()
        )

    parent = None
    if parent_id is not None:
        parent = db.query(Organization).filter(Organization.id == parent_id).first()
        if not parent:
            raise HTTPException(status_code=400, detail=f"Parent organization not found: {parent_id}")

    if row is not None and parent is not None:
        if int(row.id) == int(parent.id):
            raise HTTPException(status_code=400, detail="Organization cannot be its own parent")
        probe_id: Optional[int] = int(parent.id)
        visited: set[int] = set()
        while probe_id is not None and probe_id not in visited:
            if probe_id == int(row.id):
                raise HTTPException(
                    status_code=400,
                    detail="Invalid parent_id: cycle detected in organization hierarchy",
                )
            visited.add(probe_id)
            probe_id = (
                db.query(Organization.parent_id)
                .filter(Organization.id == probe_id)
                .scalar()
            )

    duplicate = db.query(Organization).filter(Organization.code == code)
    if row is not None:
        duplicate = duplicate.filter(Organization.id != int(row.id))
    if duplicate.first():
        raise HTTPException(status_code=409, detail=f"Organization code already exists: {code}")

    before = _serialize_organization_snapshot(row)
    if row:
        row.code = code
        row.name = name
        row.org_type = org_type
        row.parent_id = parent_id
        row.is_active = bool(payload.is_active)
    else:
        row = Organization(
            code=code,
            name=name,
            org_type=org_type,
            parent_id=parent_id,
            is_active=bool(payload.is_active),
        )
        db.add(row)

    if contracts_payload is not None:
        _sync_organization_contracts(
            row,
            contracts_payload=contracts_payload,
            block_lookup=block_lookup,
        )

    db.flush()
    after = _serialize_organization_snapshot(row)
    _audit_log(
        db,
        actor=current_user,
        action="organization.upsert",
        target_type="organization",
        target_key=code,
        before=before,
        after=after,
    )
    db.commit()
    return {
        "ok": True,
        "message": "Organization upserted",
        "id": after["id"] if after else None,
        "code": after["code"] if after else code,
        "item": after,
    }


@router.post("/organizations/delete")
def delete_organization_settings(
    payload: OrganizationDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("organizations:manage")),
):
    if payload.id is None and not _norm(payload.code):
        raise HTTPException(status_code=400, detail="id or code is required")

    row = None
    if payload.id is not None:
        row = (
            db.query(Organization)
            .options(joinedload(Organization.contracts).joinedload(OrganizationContract.block))
            .filter(Organization.id == int(payload.id))
            .first()
        )
    if row is None and _norm(payload.code):
        row = (
            db.query(Organization)
            .options(joinedload(Organization.contracts).joinedload(OrganizationContract.block))
            .filter(Organization.code == _upper(payload.code))
            .first()
        )
    if row is None:
        return {"ok": True, "message": "Organization not found (noop)"}

    code = _upper(row.code)
    if code == "SYSTEM_ROOT":
        raise HTTPException(status_code=409, detail="SYSTEM_ROOT is protected and cannot be deleted")

    children_count = db.query(Organization).filter(Organization.parent_id == int(row.id)).count()
    users_count = db.query(DbUser).filter(DbUser.organization_id == int(row.id)).count()
    workboard_count = db.query(WorkboardItem).filter(WorkboardItem.organization_id == int(row.id)).count()
    dependencies = {
        "children": int(children_count),
        "users": int(users_count),
        "workboard_items": int(workboard_count),
    }

    before = _serialize_organization_snapshot(row)
    if payload.hard_delete:
        if children_count > 0 or users_count > 0 or workboard_count > 0:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Cannot hard delete organization with dependencies. "
                    f"children={children_count}, users={users_count}, workboard_items={workboard_count}"
                ),
            )
        _audit_log(
            db,
            actor=current_user,
            action="organization.delete.hard",
            target_type="organization",
            target_key=code,
            before=before,
            after=None,
        )
        db.delete(row)
        db.commit()
        return {"ok": True, "message": "Organization deleted", "id": row.id, "code": code}

    row.is_active = False
    after = _serialize_organization_snapshot(row)
    _audit_log(
        db,
        actor=current_user,
        action="organization.delete.soft",
        target_type="organization",
        target_key=code,
        before=before,
        after=after,
    )
    db.commit()
    return {
        "ok": True,
        "message": "Organization disabled",
        "id": row.id,
        "code": code,
        "dependencies": dependencies,
    }


# --- Projects ---
@router.get("/projects")
def list_projects(db: Session = Depends(get_db)):
    projects = db.query(Project).order_by(Project.code).all()
    return {"ok": True, "items": [
        {
            "id": p.id, "code": p.code, "project_code": p.code,
            "project_name": p.name_e or p.name_p,
            "root_path": p.root_path, "is_active": p.is_active,
            "docnum_template": p.docnum_template
        } for p in projects
    ]}

@router.post("/projects/upsert")
def upsert_project(
    payload: ProjectIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    if not code: raise HTTPException(status_code=400, detail="code is required")
    proj = db.query(Project).filter(Project.code == code).first()
    before = _as_dict(proj, ["code", "name_e", "name_p", "root_path", "is_active", "docnum_template"])
    pname = _norm(payload.project_name) or _norm(payload.name_e) or _norm(payload.name_p)
    if proj:
        if pname: proj.name_e = pname
        proj.root_path = _norm(payload.root_path) or proj.root_path
        if payload.is_active is not None: proj.is_active = payload.is_active
        proj.docnum_template = _norm(payload.docnum_template) or proj.docnum_template
    else:
        proj = Project(code=code, name_e=pname, root_path=_norm(payload.root_path),
                       is_active=payload.is_active if payload.is_active is not None else True,
                       docnum_template=_norm(payload.docnum_template) or "{PROJECT}-{MDR}{PKG}-{BLK}{LVL}")
        db.add(proj)
    _audit_log(
        db,
        actor=current_user,
        action="project.upsert",
        target_type="project",
        target_key=code,
        before=before,
        after=_as_dict(proj, ["code", "name_e", "name_p", "root_path", "is_active", "docnum_template"]),
    )
    db.commit()
    return {"ok": True, "message": "Project upserted", "code": code}

@router.post("/projects/delete")
def delete_project(
    payload: ProjectDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(Project).filter(Project.code == code).first()
    if not row:
        return {"ok": True, "message": "Project not found (noop)"}
    before = _as_dict(row, ["code", "name_e", "name_p", "root_path", "is_active", "docnum_template"])

    if payload.hard_delete:
        try:
            _audit_log(
                db,
                actor=current_user,
                action="project.delete.hard",
                target_type="project",
                target_key=code,
                before=before,
                after=None,
            )
            db.delete(row)
            db.commit()
            return {"ok": True, "message": "Project deleted", "code": code}
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=f"Project {code} is in use and cannot be hard deleted.",
            )

    row.is_active = False
    _audit_log(
        db,
        actor=current_user,
        action="project.delete.soft",
        target_type="project",
        target_key=code,
        before=before,
        after=_as_dict(row, ["code", "name_e", "name_p", "root_path", "is_active", "docnum_template"]),
    )
    db.commit()
    return {"ok": True, "message": "Project disabled", "code": code}


# --- Blocks ---
@router.get("/blocks")
def list_blocks(project_code: Optional[str] = Query(default=None), db: Session = Depends(get_db)):
    q = db.query(Block)
    if project_code: q = q.filter(Block.project_code == _upper(project_code))
    rows = q.order_by(Block.project_code, Block.sort_order, Block.code).all()
    return {"ok": True, "items": [
        {"id": b.id, "project_code": b.project_code, "code": b.code, "name_e": b.name_e, "name_p": b.name_p,
         "sort_order": b.sort_order, "is_active": b.is_active} for b in rows
    ]}

@router.post("/blocks/upsert")
def upsert_block(
    payload: BlockIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    pcode = _upper(payload.project_code)
    bcode = _upper(payload.code)
    proj = db.query(Project).filter(Project.code == pcode).first()
    if not proj: raise HTTPException(status_code=400, detail=f"Project {pcode} not found")
    blk = db.query(Block).filter(Block.project_code == pcode, Block.code == bcode).first()
    before = _as_dict(blk, ["project_code", "code", "name_e", "name_p", "sort_order", "is_active"])
    if blk:
        blk.name_e = _norm(payload.name_e) or blk.name_e
        blk.name_p = _norm(payload.name_p) or blk.name_p
        blk.sort_order = payload.sort_order
        blk.is_active = payload.is_active
    else:
        blk = Block(project_code=pcode, code=bcode, name_e=_norm(payload.name_e) or bcode,
                    name_p=_norm(payload.name_p), sort_order=payload.sort_order, is_active=payload.is_active)
        db.add(blk)
    _audit_log(
        db,
        actor=current_user,
        action="block.upsert",
        target_type="block",
        target_key=f"{pcode}:{bcode}",
        before=before,
        after=_as_dict(blk, ["project_code", "code", "name_e", "name_p", "sort_order", "is_active"]),
    )
    db.commit()
    return {"ok": True, "message": "Block upserted", "project_code": pcode, "code": bcode}

@router.post("/blocks/delete")
def delete_block(
    payload: BlockDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    pcode = _upper(payload.project_code)
    bcode = _upper(payload.code)
    blk = db.query(Block).filter(Block.project_code == pcode, Block.code == bcode).first()
    if not blk: return {"ok": True, "message": "Block not found (noop)"}
    before = _as_dict(blk, ["project_code", "code", "name_e", "name_p", "sort_order", "is_active"])
    if payload.hard_delete: db.delete(blk)
    else: blk.is_active = False
    _audit_log(
        db,
        actor=current_user,
        action="block.delete.hard" if payload.hard_delete else "block.delete.soft",
        target_type="block",
        target_key=f"{pcode}:{bcode}",
        before=before,
        after=None if payload.hard_delete else _as_dict(blk, ["project_code", "code", "name_e", "name_p", "sort_order", "is_active"]),
    )
    db.commit()
    return {"ok": True, "message": "Block deleted" if payload.hard_delete else "Block disabled"}


# --- MDR Categories ---
@router.get("/mdr-categories")
def list_mdr_categories(db: Session = Depends(get_db)):
    rows = db.query(MdrCategory).order_by(MdrCategory.sort_order, MdrCategory.code).all()
    return {"ok": True, "items": [
        {"code": r.code, "name_e": r.name_e, "name_p": r.name_p, "folder_name": r.folder_name,
         "sort_order": r.sort_order, "is_active": r.is_active} for r in rows
    ]}

@router.post("/mdr-categories/upsert")
def upsert_mdr_category(
    payload: MdrCategoryIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    if not code: raise HTTPException(status_code=400, detail="code is required")
    row = db.query(MdrCategory).filter(MdrCategory.code == code).first()
    before = _as_dict(row, ["code", "name_e", "name_p", "folder_name", "sort_order", "is_active"])
    if row:
        row.name_e = _norm(payload.name_e) or row.name_e
        row.name_p = _norm(payload.name_p) or row.name_p
        row.folder_name = _norm(payload.folder_name) or row.folder_name
        row.sort_order = payload.sort_order
        row.is_active = payload.is_active
    else:
        row = MdrCategory(code=code, name_e=_norm(payload.name_e), name_p=_norm(payload.name_p),
                          folder_name=_norm(payload.folder_name) or _norm(payload.name_e),
                          sort_order=payload.sort_order, is_active=payload.is_active)
        db.add(row)
    _audit_log(
        db,
        actor=current_user,
        action="mdr_category.upsert",
        target_type="mdr_category",
        target_key=code,
        before=before,
        after=_as_dict(row, ["code", "name_e", "name_p", "folder_name", "sort_order", "is_active"]),
    )
    db.commit()
    from app.api.v1.routers.lookup import invalidate_dictionary_cache

    invalidate_dictionary_cache()
    return {"ok": True, "message": "MDR Category upserted", "code": code}

@router.post("/mdr-categories/delete")
def delete_mdr_category(
    payload: MdrCategoryDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(MdrCategory).filter(MdrCategory.code == code).first()
    if not row: return {"ok": True, "message": "MDR Category not found (noop)"}
    before = _as_dict(row, ["code", "name_e", "name_p", "folder_name", "sort_order", "is_active"])
    if payload.hard_delete: db.delete(row)
    else: row.is_active = False
    _audit_log(
        db,
        actor=current_user,
        action="mdr_category.delete.hard" if payload.hard_delete else "mdr_category.delete.soft",
        target_type="mdr_category",
        target_key=code,
        before=before,
        after=None if payload.hard_delete else _as_dict(row, ["code", "name_e", "name_p", "folder_name", "sort_order", "is_active"]),
    )
    db.commit()
    from app.api.v1.routers.lookup import invalidate_dictionary_cache

    invalidate_dictionary_cache()
    return {"ok": True, "message": "MDR Category deleted" if payload.hard_delete else "MDR Category disabled"}


# --- Phases ---
@router.get("/phases")
def list_phases_settings(db: Session = Depends(get_db)):
    phases = db.query(Phase).order_by(Phase.ph_code).all()
    return {"ok": True, "items": [{"ph_code": p.ph_code, "name_e": p.name_e, "name_p": p.name_p} for p in phases]}

@router.post("/phases/upsert")
def upsert_phase_settings(
    payload: PhaseIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.ph_code)
    ph = db.query(Phase).filter(Phase.ph_code == code).first()
    before = _as_dict(ph, ["ph_code", "name_e", "name_p"])
    if ph: ph.name_e = payload.name_e; ph.name_p = payload.name_p
    else:
        ph = Phase(ph_code=code, name_e=payload.name_e, name_p=payload.name_p)
        db.add(ph)
    _audit_log(
        db,
        actor=current_user,
        action="phase.upsert",
        target_type="phase",
        target_key=code,
        before=before,
        after=_as_dict(ph, ["ph_code", "name_e", "name_p"]),
    )
    db.commit()
    return {"ok": True, "message": "Phase upserted"}

@router.post("/phases/delete")
def delete_phase_settings(
    payload: PhaseDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.ph_code)
    row = db.query(Phase).filter(Phase.ph_code == code).first()
    if not row:
        return {"ok": True, "message": "Phase not found (noop)"}
    before = _as_dict(row, ["ph_code", "name_e", "name_p"])
    try:
        _audit_log(
            db,
            actor=current_user,
            action="phase.delete.hard",
            target_type="phase",
            target_key=code,
            before=before,
            after=None,
        )
        db.delete(row)
        db.commit()
        return {"ok": True, "message": "Phase deleted", "ph_code": code}
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"Phase {code} is in use by documents and cannot be deleted.",
        )


# --- Disciplines ---
@router.get("/disciplines")
def list_disciplines_settings(db: Session = Depends(get_db)):
    discs = db.query(Discipline).order_by(Discipline.code).all()
    return {"ok": True, "items": [{"code": d.code, "discipline_code": d.code, "name_e": d.name_e, "name_p": d.name_p} for d in discs]}

@router.post("/disciplines/upsert")
def upsert_discipline_settings(
    payload: DisciplineIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    disc = db.query(Discipline).filter(Discipline.code == code).first()
    before = _as_dict(disc, ["code", "name_e", "name_p"])
    if disc:
        disc.name_e = payload.name_e
        if payload.name_p: disc.name_p = payload.name_p
    else:
        disc = Discipline(code=code, name_e=payload.name_e, name_p=payload.name_p)
        db.add(disc)
    _audit_log(
        db,
        actor=current_user,
        action="discipline.upsert",
        target_type="discipline",
        target_key=code,
        before=before,
        after=_as_dict(disc, ["code", "name_e", "name_p"]),
    )
    db.commit()
    return {"ok": True, "message": "Discipline upserted"}

@router.post("/disciplines/delete")
def delete_discipline_settings(
    payload: DisciplineDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(Discipline).filter(Discipline.code == code).first()
    if not row:
        return {"ok": True, "message": "Discipline not found (noop)"}
    before = _as_dict(row, ["code", "name_e", "name_p"])
    try:
        _audit_log(
            db,
            actor=current_user,
            action="discipline.delete.hard",
            target_type="discipline",
            target_key=code,
            before=before,
            after=None,
        )
        db.delete(row)
        db.commit()
        return {"ok": True, "message": "Discipline deleted", "code": code}
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"Discipline {code} is in use and cannot be deleted.",
        )


# --- Packages ---
@router.get("/packages")
def list_packages_settings(discipline_code: Optional[str] = Query(default=None), db: Session = Depends(get_db)):
    q = db.query(Package)
    if discipline_code:
        q = q.filter(Package.discipline_code == _upper(discipline_code))
    rows = q.order_by(Package.discipline_code, Package.package_code).all()
    return {"ok": True, "items": [
        {
            "discipline_code": r.discipline_code,
            "package_code": r.package_code,
            "name_e": r.name_e,
            "name_p": r.name_p,
        }
        for r in rows
    ]}

@router.post("/packages/upsert")
def upsert_package_settings(
    payload: PackageIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    dcode = _upper(payload.discipline_code)
    raw_pcode = _upper(payload.package_code)

    disc = db.query(Discipline).filter(Discipline.code == dcode).first()
    if not disc:
        raise HTTPException(status_code=400, detail=f"Discipline {dcode} not found")

    normalized_input_code = _normalize_package_code_by_discipline(raw_pcode, dcode)
    candidate_codes: List[str] = []
    if raw_pcode:
        candidate_codes.append(raw_pcode)
    if normalized_input_code and normalized_input_code not in candidate_codes:
        candidate_codes.append(normalized_input_code)

    row = None
    for code in candidate_codes:
        row = db.query(Package).filter(Package.discipline_code == dcode, Package.package_code == code).first()
        if row:
            break

    before = _as_dict(row, ["discipline_code", "package_code", "name_e", "name_p"])
    if row:
        pcode = row.package_code
        row.name_e = _norm(payload.name_e)
        row.name_p = _norm(payload.name_p) or row.name_p
    else:
        pcode = _next_package_code_for_discipline(db, dcode)
        row = Package(
            discipline_code=dcode,
            package_code=pcode,
            name_e=_norm(payload.name_e),
            name_p=_norm(payload.name_p),
        )
        db.add(row)
    _audit_log(
        db,
        actor=current_user,
        action="package.upsert",
        target_type="package",
        target_key=f"{dcode}:{pcode}",
        before=before,
        after=_as_dict(row, ["discipline_code", "package_code", "name_e", "name_p"]),
    )
    db.commit()
    return {"ok": True, "message": "Package upserted", "discipline_code": dcode, "package_code": pcode}

@router.post("/packages/delete")
def delete_package_settings(
    payload: PackageDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    dcode = _upper(payload.discipline_code)
    pcode = _upper(payload.package_code)
    row = db.query(Package).filter(Package.discipline_code == dcode, Package.package_code == pcode).first()
    if not row:
        return {"ok": True, "message": "Package not found (noop)"}
    before = _as_dict(row, ["discipline_code", "package_code", "name_e", "name_p"])
    try:
        _audit_log(
            db,
            actor=current_user,
            action="package.delete.hard",
            target_type="package",
            target_key=f"{dcode}:{pcode}",
            before=before,
            after=None,
        )
        db.delete(row)
        db.commit()
        return {"ok": True, "message": "Package deleted", "discipline_code": dcode, "package_code": pcode}
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"Package {pcode} in discipline {dcode} is in use and cannot be deleted.",
        )


# --- Levels ---
@router.get("/levels")
def list_levels_settings(db: Session = Depends(get_db)):
    levels = db.query(Level).order_by(Level.sort_order, Level.code).all()
    return {"ok": True, "items": [{"code": l.code, "name_e": l.name_e, "name_p": l.name_p, "sort_order": l.sort_order} for l in levels]}

@router.post("/levels/upsert")
def upsert_level_settings(
    payload: LevelIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _norm(payload.code)
    lvl = db.query(Level).filter(Level.code == code).first()
    before = _as_dict(lvl, ["code", "name_e", "name_p", "sort_order"])
    if lvl:
        if payload.name_e: lvl.name_e = payload.name_e
        if payload.name_p: lvl.name_p = payload.name_p
        lvl.sort_order = payload.sort_order
    else:
        lvl = Level(code=code, name_e=payload.name_e or code, name_p=payload.name_p, sort_order=payload.sort_order)
        db.add(lvl)
    _audit_log(
        db,
        actor=current_user,
        action="level.upsert",
        target_type="level",
        target_key=code,
        before=before,
        after=_as_dict(lvl, ["code", "name_e", "name_p", "sort_order"]),
    )
    db.commit()
    return {"ok": True, "message": "Level upserted"}

@router.post("/levels/delete")
def delete_level_settings(
    payload: LevelDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _norm(payload.code)
    row = db.query(Level).filter(Level.code == code).first()
    if not row:
        return {"ok": True, "message": "Level not found (noop)"}
    before = _as_dict(row, ["code", "name_e", "name_p", "sort_order"])
    try:
        _audit_log(
            db,
            actor=current_user,
            action="level.delete.hard",
            target_type="level",
            target_key=code,
            before=before,
            after=None,
        )
        db.delete(row)
        db.commit()
        return {"ok": True, "message": "Level deleted", "code": code}
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"Level {code} is in use by documents and cannot be deleted.",
        )


# --- Statuses (✅ ADDED) ---
@router.get("/statuses")
def list_statuses_settings(db: Session = Depends(get_db)):
    statuses = db.query(DocStatus).order_by(DocStatus.sort_order).all()
    return {"ok": True, "items": [{"code": s.code, "name": s.name, "description": s.description, "sort_order": s.sort_order} for s in statuses]}

@router.post("/statuses/upsert")
def upsert_status_settings(
    payload: DocStatusIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    st = db.query(DocStatus).filter(DocStatus.code == code).first()
    before = _as_dict(st, ["code", "name", "description", "sort_order"])
    if st:
        st.name = payload.name
        st.description = payload.description
        st.sort_order = payload.sort_order
    else:
        st = DocStatus(code=code, name=payload.name, description=payload.description, sort_order=payload.sort_order)
        db.add(st)
    _audit_log(
        db,
        actor=current_user,
        action="status.upsert",
        target_type="status",
        target_key=code,
        before=before,
        after=_as_dict(st, ["code", "name", "description", "sort_order"]),
    )
    db.commit()
    return {"ok": True, "message": "Status upserted"}

@router.post("/statuses/delete")
def delete_status_settings(
    payload: DocStatusDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(DocStatus).filter(DocStatus.code == code).first()
    if not row:
        return {"ok": True, "message": "Status not found (noop)"}
    before = _as_dict(row, ["code", "name", "description", "sort_order"])
    _audit_log(
        db,
        actor=current_user,
        action="status.delete.hard",
        target_type="status",
        target_key=code,
        before=before,
        after=None,
    )
    db.delete(row)
    db.commit()
    return {"ok": True, "message": "Status deleted", "code": code}


# --- Correspondence Parameters ---
@router.get("/correspondence-issuing")
def list_correspondence_issuing_settings(db: Session = Depends(get_db)):
    rows = (
        db.query(IssuingEntity)
        .order_by(IssuingEntity.sort_order.asc(), IssuingEntity.code.asc())
        .all()
    )
    return {
        "ok": True,
        "items": [
            {
                "code": row.code,
                "name_e": row.name_e,
                "name_p": row.name_p,
                "project_code": row.project_code,
                "is_active": bool(row.is_active),
                "sort_order": int(row.sort_order or 0),
            }
            for row in rows
        ],
    }


@router.post("/correspondence-issuing/upsert")
def upsert_correspondence_issuing_settings(
    payload: CorrespondenceIssuingIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")

    project_code = _upper(payload.project_code) or None
    if project_code:
        project = db.query(Project).filter(Project.code == project_code).first()
        if not project:
            raise HTTPException(status_code=404, detail=f"Project not found: {project_code}")

    name_e = _norm(payload.name_e) or _norm(payload.name_p)
    if not name_e:
        raise HTTPException(status_code=400, detail="name_e is required")

    row = db.query(IssuingEntity).filter(IssuingEntity.code == code).first()
    before = _as_dict(row, ["code", "name_e", "name_p", "project_code", "is_active", "sort_order"])
    if row:
        row.name_e = name_e
        row.name_p = _norm(payload.name_p) or None
        row.project_code = project_code
        row.is_active = bool(payload.is_active)
        row.sort_order = int(payload.sort_order or 0)
    else:
        row = IssuingEntity(
            code=code,
            name_e=name_e,
            name_p=_norm(payload.name_p) or None,
            project_code=project_code,
            is_active=bool(payload.is_active),
            sort_order=int(payload.sort_order or 0),
        )
        db.add(row)

    _audit_log(
        db,
        actor=current_user,
        action="correspondence_issuing.upsert",
        target_type="correspondence_issuing",
        target_key=code,
        before=before,
        after=_as_dict(row, ["code", "name_e", "name_p", "project_code", "is_active", "sort_order"]),
    )
    db.commit()
    return {"ok": True, "message": "Correspondence issuing upserted", "code": code}


@router.post("/correspondence-issuing/delete")
def delete_correspondence_issuing_settings(
    payload: CorrespondenceIssuingDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(IssuingEntity).filter(IssuingEntity.code == code).first()
    if not row:
        return {"ok": True, "message": "Issuing entity not found (noop)"}

    before = _as_dict(row, ["code", "name_e", "name_p", "project_code", "is_active", "sort_order"])
    if payload.hard_delete:
        try:
            _audit_log(
                db,
                actor=current_user,
                action="correspondence_issuing.delete.hard",
                target_type="correspondence_issuing",
                target_key=code,
                before=before,
                after=None,
            )
            db.delete(row)
            db.commit()
            return {"ok": True, "message": "Issuing entity deleted", "code": code}
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=f"Issuing entity {code} is in use by correspondences and cannot be hard deleted.",
            )

    row.is_active = False
    _audit_log(
        db,
        actor=current_user,
        action="correspondence_issuing.delete.soft",
        target_type="correspondence_issuing",
        target_key=code,
        before=before,
        after=_as_dict(row, ["code", "name_e", "name_p", "project_code", "is_active", "sort_order"]),
    )
    db.commit()
    return {"ok": True, "message": "Issuing entity disabled", "code": code}


@router.get("/correspondence-categories")
def list_correspondence_categories_settings(db: Session = Depends(get_db)):
    rows = (
        db.query(CorrespondenceCategory)
        .order_by(CorrespondenceCategory.sort_order.asc(), CorrespondenceCategory.code.asc())
        .all()
    )
    return {
        "ok": True,
        "items": [
            {
                "code": row.code,
                "name_e": row.name_e,
                "name_p": row.name_p,
                "is_active": bool(row.is_active),
                "sort_order": int(row.sort_order or 0),
            }
            for row in rows
        ],
    }


@router.post("/correspondence-categories/upsert")
def upsert_correspondence_categories_settings(
    payload: CorrespondenceCategoryIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")

    name_e = _norm(payload.name_e) or _norm(payload.name_p)
    if not name_e:
        raise HTTPException(status_code=400, detail="name_e is required")

    row = db.query(CorrespondenceCategory).filter(CorrespondenceCategory.code == code).first()
    before = _as_dict(row, ["code", "name_e", "name_p", "is_active", "sort_order"])
    if row:
        row.name_e = name_e
        row.name_p = _norm(payload.name_p) or None
        row.is_active = bool(payload.is_active)
        row.sort_order = int(payload.sort_order or 0)
    else:
        row = CorrespondenceCategory(
            code=code,
            name_e=name_e,
            name_p=_norm(payload.name_p) or None,
            is_active=bool(payload.is_active),
            sort_order=int(payload.sort_order or 0),
        )
        db.add(row)

    _audit_log(
        db,
        actor=current_user,
        action="correspondence_category.upsert",
        target_type="correspondence_category",
        target_key=code,
        before=before,
        after=_as_dict(row, ["code", "name_e", "name_p", "is_active", "sort_order"]),
    )
    db.commit()
    return {"ok": True, "message": "Correspondence category upserted", "code": code}


@router.post("/correspondence-categories/delete")
def delete_correspondence_categories_settings(
    payload: CorrespondenceCategoryDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(CorrespondenceCategory).filter(CorrespondenceCategory.code == code).first()
    if not row:
        return {"ok": True, "message": "Correspondence category not found (noop)"}

    before = _as_dict(row, ["code", "name_e", "name_p", "is_active", "sort_order"])
    if payload.hard_delete:
        try:
            _audit_log(
                db,
                actor=current_user,
                action="correspondence_category.delete.hard",
                target_type="correspondence_category",
                target_key=code,
                before=before,
                after=None,
            )
            db.delete(row)
            db.commit()
            return {"ok": True, "message": "Correspondence category deleted", "code": code}
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=f"Correspondence category {code} is in use and cannot be hard deleted.",
            )

    row.is_active = False
    _audit_log(
        db,
        actor=current_user,
        action="correspondence_category.delete.soft",
        target_type="correspondence_category",
        target_key=code,
        before=before,
        after=_as_dict(row, ["code", "name_e", "name_p", "is_active", "sort_order"]),
    )
    db.commit()
    return {"ok": True, "message": "Correspondence category disabled", "code": code}


def _serialize_correspondence_department(row: CorrespondenceDepartment) -> Dict[str, Any]:
    return {
        "code": row.code,
        "name_e": row.name_e,
        "name_p": row.name_p,
        "is_active": bool(row.is_active),
        "sort_order": int(row.sort_order or 0),
    }


@router.get("/correspondence-departments")
def list_correspondence_departments_settings(db: Session = Depends(get_db)):
    rows = (
        db.query(CorrespondenceDepartment)
        .order_by(CorrespondenceDepartment.sort_order.asc(), CorrespondenceDepartment.code.asc())
        .all()
    )
    return {"ok": True, "items": [_serialize_correspondence_department(row) for row in rows]}


@router.post("/correspondence-departments/upsert")
def upsert_correspondence_departments_settings(
    payload: CorrespondenceDepartmentIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")

    name_e = _norm(payload.name_e) or _norm(payload.name_p)
    if not name_e:
        raise HTTPException(status_code=400, detail="name_e is required")

    row = db.query(CorrespondenceDepartment).filter(CorrespondenceDepartment.code == code).first()
    before = _as_dict(row, ["code", "name_e", "name_p", "is_active", "sort_order"])
    if row:
        row.name_e = name_e
        row.name_p = _norm(payload.name_p) or None
        row.is_active = bool(payload.is_active)
        row.sort_order = int(payload.sort_order or 0)
    else:
        row = CorrespondenceDepartment(
            code=code,
            name_e=name_e,
            name_p=_norm(payload.name_p) or None,
            is_active=bool(payload.is_active),
            sort_order=int(payload.sort_order or 0),
        )
        db.add(row)

    _audit_log(
        db,
        actor=current_user,
        action="correspondence_department.upsert",
        target_type="correspondence_department",
        target_key=code,
        before=before,
        after=_serialize_correspondence_department(row),
    )
    db.commit()
    return {"ok": True, "message": "Correspondence department upserted", "code": code}


@router.post("/correspondence-departments/delete")
def delete_correspondence_departments_settings(
    payload: CorrespondenceDepartmentDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    row = db.query(CorrespondenceDepartment).filter(CorrespondenceDepartment.code == code).first()
    if not row:
        return {"ok": True, "message": "Correspondence department not found (noop)"}

    before = _serialize_correspondence_department(row)
    if payload.hard_delete:
        try:
            _audit_log(
                db,
                actor=current_user,
                action="correspondence_department.delete.hard",
                target_type="correspondence_department",
                target_key=code,
                before=before,
                after=None,
            )
            db.delete(row)
            db.commit()
            return {"ok": True, "message": "Correspondence department deleted", "code": code}
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=f"Correspondence department {code} is in use and cannot be hard deleted.",
            )

    row.is_active = False
    _audit_log(
        db,
        actor=current_user,
        action="correspondence_department.delete.soft",
        target_type="correspondence_department",
        target_key=code,
        before=before,
        after=_serialize_correspondence_department(row),
    )
    db.commit()
    return {"ok": True, "message": "Correspondence department disabled", "code": code}


def _serialize_correspondence_tag(row: DocumentTag) -> Dict[str, Any]:
    return {
        "id": int(row.id or 0),
        "name": row.name,
        "color": row.color,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


@router.get("/correspondence-tags")
def list_correspondence_tags_settings(db: Session = Depends(get_db)):
    rows = tag_service.list_tags(db)
    return {"ok": True, "items": [_serialize_correspondence_tag(row) for row in rows]}


@router.post("/correspondence-tags/upsert")
def upsert_correspondence_tag_settings(
    payload: CorrespondenceTagIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    if payload.id:
        row = tag_service.update_tag(
            db,
            int(payload.id),
            name=payload.name,
            color=payload.color,
        )
        action = "correspondence_tag.update"
    else:
        row = tag_service.create_tag(
            db,
            name=payload.name,
            color=payload.color,
            user=current_user,
        )
        action = "correspondence_tag.create"
    _audit_log(
        db,
        actor=current_user,
        action=action,
        target_type="correspondence_tag",
        target_key=str(int(row.id or 0)),
        before=None,
        after=_serialize_correspondence_tag(row),
    )
    db.commit()
    return {"ok": True, "message": "Correspondence tag upserted", "id": int(row.id or 0)}


@router.post("/correspondence-tags/delete")
def delete_correspondence_tag_settings(
    payload: CorrespondenceTagDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    row = tag_service.get_tag_or_404(db, int(payload.id))
    before = _serialize_correspondence_tag(row)
    tag_service.delete_tag(db, int(payload.id))
    _audit_log(
        db,
        actor=current_user,
        action="correspondence_tag.delete",
        target_type="correspondence_tag",
        target_key=str(int(payload.id)),
        before=before,
        after=None,
    )
    db.commit()
    return {"ok": True, "message": "Correspondence tag deleted", "id": int(payload.id)}


@router.get("/transmittal-parties")
def get_transmittal_parties_settings(db: Session = Depends(get_db)):
    return {"ok": True, **get_transmittal_parties(db)}


@router.post("/transmittal-parties")
def save_transmittal_parties_settings(
    payload: TransmittalPartiesIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    before = get_transmittal_parties(db)
    after = set_transmittal_parties(db, payload.model_dump())
    _audit_log(
        db,
        actor=current_user,
        action="transmittal_parties.update",
        target_type="transmittal_parties",
        target_key="custom.transmittal.parties.v1",
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "message": "Transmittal parties saved", **after}


# --- KV ---
@router.get("/kv")
def kv_list(
    include_legacy: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    return {"ok": True, "items": _kv_get_all(db, include_legacy=include_legacy)}

@router.post("/kv/set")
def kv_set(
    payload: KvIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    key = _norm(payload.key)
    if not _is_allowed_kv_write_key(key):
        raise HTTPException(
            status_code=403,
            detail=(
                "KV key is not allowed. "
                f"Allowed prefixes: {', '.join(KV_ALLOWED_WRITE_PREFIXES)}"
            ),
        )

    before_row = db.query(SettingsKV).filter(SettingsKV.key == key).first()
    before_val = before_row.value if before_row else None

    _kv_set(db, key, payload.value)
    _audit_log(
        db,
        actor=current_user,
        action="kv.set",
        target_type="kv",
        target_key=key,
        before={"value": before_val},
        after={"value": payload.value},
    )
    db.commit()
    return {"ok": True, "message": "Setting saved"}


@router.get("/storage-paths")
def get_storage_paths(db: Session = Depends(get_db)):
    correspondence_path = _kv_get_value(
        db,
        STORAGE_PATH_CORRESPONDENCE_KEY,
        DEFAULT_CORRESPONDENCE_STORAGE_PATH,
    )
    site_log_path = _kv_get_value(db, STORAGE_PATH_SITE_LOG_KEY, "")
    return {
        "ok": True,
        "mdr_storage_path": _kv_get_value(db, STORAGE_PATH_MDR_KEY, DEFAULT_MDR_STORAGE_PATH),
        "correspondence_storage_path": correspondence_path,
        "site_log_storage_path": site_log_path,
        "site_log_storage_path_effective": site_log_path or correspondence_path,
    }


@router.post("/storage-paths")
def save_storage_paths(
    payload: StoragePathsIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    mdr_storage_path = _norm(payload.mdr_storage_path)
    correspondence_storage_path = _norm(payload.correspondence_storage_path)
    site_log_storage_path = _norm(payload.site_log_storage_path)
    before_site_log = _kv_get_value(db, STORAGE_PATH_SITE_LOG_KEY, "")
    integrations = get_storage_integrations(db)
    primary_provider = resolve_primary_storage_provider(integrations)
    nextcloud_runtime = resolve_nextcloud_runtime(integrations)
    use_nextcloud_webdav = (
        primary_provider == "nextcloud"
        and bool(nextcloud_runtime.get("mode_is_webdav"))
    )
    remote_root = _norm(nextcloud_runtime.get("root_path")) or "/"
    errors: list[dict[str, str]] = []
    normalized_mdr, mdr_errors = StorageManager.validate_storage_path(
        mdr_storage_path,
        field="mdr_storage_path",
        network_username=payload.network_username,
        network_password=payload.network_password,
        allow_remote_webdav_path=use_nextcloud_webdav,
        remote_root_value=remote_root if use_nextcloud_webdav else None,
    )
    normalized_corr, corr_errors = StorageManager.validate_storage_path(
        correspondence_storage_path,
        field="correspondence_storage_path",
        network_username=payload.network_username,
        network_password=payload.network_password,
        allow_remote_webdav_path=use_nextcloud_webdav,
        remote_root_value=remote_root if use_nextcloud_webdav else None,
    )
    errors.extend(mdr_errors)
    errors.extend(corr_errors)
    normalized_site_log = before_site_log if payload.site_log_storage_path is None else ""
    if payload.site_log_storage_path is not None and site_log_storage_path:
        normalized_site_log, site_log_errors = StorageManager.validate_storage_path(
            site_log_storage_path,
            field="site_log_storage_path",
            network_username=payload.network_username,
            network_password=payload.network_password,
            allow_remote_webdav_path=use_nextcloud_webdav,
            remote_root_value=remote_root if use_nextcloud_webdav else None,
        )
        errors.extend(site_log_errors)
    if errors:
        raise HTTPException(status_code=422, detail=errors)

    before = {
        "mdr_storage_path": _kv_get_value(db, STORAGE_PATH_MDR_KEY, DEFAULT_MDR_STORAGE_PATH),
        "correspondence_storage_path": _kv_get_value(
            db,
            STORAGE_PATH_CORRESPONDENCE_KEY,
            DEFAULT_CORRESPONDENCE_STORAGE_PATH,
        ),
        "site_log_storage_path": before_site_log,
    }
    after = {
        "mdr_storage_path": normalized_mdr,
        "correspondence_storage_path": normalized_corr,
        "site_log_storage_path": normalized_site_log,
        "site_log_storage_path_effective": normalized_site_log or normalized_corr,
    }

    _kv_set(db, STORAGE_PATH_MDR_KEY, normalized_mdr)
    _kv_set(db, STORAGE_PATH_CORRESPONDENCE_KEY, normalized_corr)
    if payload.site_log_storage_path is not None:
        _kv_set(db, STORAGE_PATH_SITE_LOG_KEY, normalized_site_log)
    _audit_log(
        db,
        actor=current_user,
        action="storage_paths.update",
        target_type="storage_paths",
        target_key=None,
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "message": "Storage paths updated", **after}


@router.get("/storage-policy")
def get_storage_policy_settings(db: Session = Depends(get_db)):
    policy = get_storage_policy(db)
    return {
        "ok": True,
        "policy": policy,
        "defaults": DEFAULT_STORAGE_POLICY,
    }


@router.post("/storage-policy")
def save_storage_policy_settings(
    payload: StoragePolicyIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    before = get_storage_policy(db)
    incoming = payload.model_dump(exclude_unset=True)
    merged = dict(before)
    merged.update(incoming)
    policy = set_storage_policy(db, merged)
    _audit_log(
        db,
        actor=current_user,
        action="storage_policy.update",
        target_type="storage_policy",
        target_key=None,
        before=before,
        after=policy,
    )
    db.commit()
    return {"ok": True, "policy": policy}


@router.get("/storage-integrations")
def get_storage_integrations_settings(db: Session = Depends(get_db)):
    integrations = get_storage_integrations(db)
    masked = _masked_storage_integrations_payload(integrations, db)
    return {
        "ok": True,
        "integrations": masked,
        "defaults": DEFAULT_STORAGE_INTEGRATIONS,
    }


@router.post("/storage-integrations")
def save_storage_integrations_settings(
    payload: StorageIntegrationsIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    before = get_storage_integrations(db)
    incoming = payload.model_dump(exclude_unset=True)
    gdrive_incoming = incoming.get("google_drive")
    if isinstance(gdrive_incoming, dict):
        before_google = dict(before.get("google_drive") or {})
        if "oauth_client_secret" in gdrive_incoming and not str(
            gdrive_incoming.get("oauth_client_secret") or ""
        ).strip():
            gdrive_incoming["oauth_client_secret"] = str(before_google.get("oauth_client_secret") or "")
        if "oauth_refresh_token" in gdrive_incoming and not str(
            gdrive_incoming.get("oauth_refresh_token") or ""
        ).strip():
            gdrive_incoming["oauth_refresh_token"] = str(before_google.get("oauth_refresh_token") or "")

    openproject_incoming = incoming.get("openproject")
    if isinstance(openproject_incoming, dict):
        raw_default_wp = str(
            openproject_incoming.get("default_work_package_id") or openproject_incoming.get("default_project_id") or ""
        ).strip()
        openproject_incoming["default_work_package_id"] = raw_default_wp
        openproject_incoming.pop("default_project_id", None)
        if "api_token" in openproject_incoming and not str(openproject_incoming.get("api_token") or "").strip():
            openproject_incoming["api_token"] = str((before.get("openproject") or {}).get("api_token") or "")

    nextcloud_incoming = incoming.get("nextcloud")
    if isinstance(nextcloud_incoming, dict):
        if "app_password" in nextcloud_incoming and not str(nextcloud_incoming.get("app_password") or "").strip():
            nextcloud_incoming["app_password"] = str((before.get("nextcloud") or {}).get("app_password") or "")
        if "public_share_password" in nextcloud_incoming and not str(
            nextcloud_incoming.get("public_share_password") or ""
        ).strip():
            nextcloud_incoming["public_share_password"] = str(
                (before.get("nextcloud") or {}).get("public_share_password") or ""
            )

        # CRITICAL: Warn if root_path is changing and WebDAV files exist
        if "root_path" in nextcloud_incoming:
            old_root = str((before.get("nextcloud") or {}).get("root_path") or "").strip()
            new_root = str(nextcloud_incoming.get("root_path") or "").strip()

            if old_root and new_root and old_root != new_root:
                # Check if any WebDAV files exist
                webdav_count = (
                    db.query(ArchiveFile)
                    .filter(ArchiveFile.stored_path.like("webdav://%"))
                    .count()
                )

                if webdav_count > 0:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Cannot change Nextcloud Root Path from '{old_root}' to '{new_root}' "
                            f"because {webdav_count} WebDAV file(s) exist in the database. "
                            "Changing the root path will break access to existing files. "
                            "If you must change it, you need to: "
                            "1) Backup your database, "
                            "2) Run a migration script to update stored_path for all WebDAV files, "
                            "3) Move files on Nextcloud to match the new structure."
                        ),
                    )

    merged = dict(before)
    for key in ("primary", "mirror", "google_drive", "openproject", "nextcloud", "local_cache"):
        if isinstance(incoming.get(key), dict):
            current = dict(merged.get(key) or {})
            current.update(incoming.get(key) or {})
            merged[key] = current

    primary_status = _storage_primary_runtime_payload(db, merged)
    if (
        str(primary_status.get("effective_provider") or "") == "nextcloud"
        and str(((merged.get("mirror") or {}).get("provider")) or "").strip().lower() == "nextcloud"
    ):
        raise HTTPException(
            status_code=400,
            detail="Nextcloud mirror cannot be enabled when Nextcloud is already the effective primary storage.",
        )

    integrations = set_storage_integrations(db, merged)
    masked = _masked_storage_integrations_payload(integrations, db)
    _audit_log(
        db,
        actor=current_user,
        action="storage_integrations.update",
        target_type="storage_integrations",
        target_key=None,
        before=before,
        after=masked,
    )
    db.commit()
    return {"ok": True, "integrations": masked}


@router.post("/storage-integrations/openproject/clear-token")
def clear_storage_openproject_token(
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    before = get_storage_integrations(db)
    merged = dict(before)
    openproject = dict(merged.get("openproject") or {})
    openproject["api_token"] = ""
    merged["openproject"] = openproject
    integrations = set_storage_integrations(db, merged)
    masked = _masked_storage_integrations_payload(integrations, db)
    _audit_log(
        db,
        actor=current_user,
        action="storage_integrations.openproject.clear_token",
        target_type="storage_integrations",
        target_key="openproject",
        before=before,
        after=masked,
    )
    db.commit()
    return {"ok": True, "integrations": masked}


@router.get("/power-bi/tokens")
def list_power_bi_tokens(
    include_inactive: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("settings:update")),
):
    query = db.query(PowerBiApiToken)
    if not include_inactive:
        query = query.filter(PowerBiApiToken.is_active.is_(True), PowerBiApiToken.revoked_at.is_(None))
    rows = query.order_by(PowerBiApiToken.id.desc()).all()
    return {
        "ok": True,
        "items": [serialize_power_bi_token(row) for row in rows],
        "defaults": {"scopes": [POWER_BI_SITE_LOG_SCOPE]},
    }


@router.post("/power-bi/tokens/mint")
def mint_power_bi_token(
    payload: PowerBiTokenMintIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    row, raw_token = create_power_bi_token(
        db,
        name=payload.name,
        created_by_id=int(current_user.id or 0) if getattr(current_user, "id", None) else None,
        expires_at=payload.expires_at,
        allowed_project_codes=payload.allowed_project_codes or [],
        allowed_report_sections=payload.allowed_report_sections or [],
        allowed_ip_ranges=payload.allowed_ip_ranges or [],
    )
    item = serialize_power_bi_token(row)
    _audit_log(
        db,
        actor=current_user,
        action="power_bi_token.mint",
        target_type="power_bi_api_token",
        target_key=str(row.id),
        before=None,
        after=item,
    )
    db.commit()
    db.refresh(row)
    item = serialize_power_bi_token(row)
    return {
        "ok": True,
        "item": item,
        "token": raw_token,
        "warning": "Power BI token is shown once. Save it securely now.",
    }


@router.post("/power-bi/tokens/{token_id}/revoke")
def revoke_power_bi_token(
    token_id: int,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    row = db.query(PowerBiApiToken).filter(PowerBiApiToken.id == int(token_id)).first()
    if not row:
        raise HTTPException(status_code=404, detail="Power BI token not found")
    before = serialize_power_bi_token(row)
    row.is_active = False
    row.revoked_at = datetime.utcnow()
    after = serialize_power_bi_token(row)
    _audit_log(
        db,
        actor=current_user,
        action="power_bi_token.revoke",
        target_type="power_bi_api_token",
        target_key=str(row.id),
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "item": after}


@router.get("/bim-revit")
def get_bim_revit_settings(
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("settings:update")),
):
    current = get_bim_revit_integration(db)
    return {
        "ok": True,
        "settings": _masked_bim_revit_payload(current),
        "defaults": _masked_bim_revit_payload(DEFAULT_BIM_REVIT_INTEGRATION),
    }


@router.post("/bim-revit")
def save_bim_revit_settings(
    payload: BimRevitSettingsIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    before = get_bim_revit_integration(db)
    incoming = payload.model_dump(exclude_unset=True)
    merged = dict(before)

    endpoint = incoming.get("api_endpoint_url")
    if endpoint is not None:
        endpoint_text = _norm(endpoint)
        if endpoint_text and not _is_valid_http_url(endpoint_text):
            raise HTTPException(status_code=400, detail="api_endpoint_url must be absolute http/https URL or root-relative path.")
        merged["api_endpoint_url"] = endpoint_text

    if "enabled" in incoming:
        merged["enabled"] = bool(incoming.get("enabled"))
    if "require_plugin_signature" in incoming:
        merged["require_plugin_signature"] = bool(incoming.get("require_plugin_signature"))

    key_id = incoming.get("plugin_key_id")
    if key_id is not None:
        merged["plugin_key_id"] = _norm(key_id)

    if "default_category_id" in incoming:
        merged["default_category_id"] = incoming.get("default_category_id")
    if "default_folder_id" in incoming:
        merged["default_folder_id"] = incoming.get("default_folder_id")
    if "allowed_mime" in incoming:
        merged["allowed_mime"] = [str(item or "").strip().lower() for item in (incoming.get("allowed_mime") or [])]
    if "max_batch_size" in incoming and incoming.get("max_batch_size") is not None:
        merged["max_batch_size"] = int(incoming.get("max_batch_size"))

    plugin_secret = incoming.get("plugin_secret")
    if plugin_secret is not None:
        raw_secret = _norm(plugin_secret)
        if raw_secret:
            merged["plugin_secret_encrypted"] = encrypt_plugin_secret(
                raw_secret,
                secret_key=_runtime_secret_key_or_500(),
            )
        else:
            merged["plugin_secret_encrypted"] = str(before.get("plugin_secret_encrypted") or "")

    if bool(merged.get("require_plugin_signature")):
        if not _norm(merged.get("plugin_key_id")):
            raise HTTPException(status_code=400, detail="plugin_key_id is required when signature is enabled.")
        if not _norm(merged.get("plugin_secret_encrypted")):
            raise HTTPException(status_code=400, detail="plugin secret is required when signature is enabled.")

    saved = set_bim_revit_integration(db, merged)
    masked = _masked_bim_revit_payload(saved)
    _audit_log(
        db,
        actor=current_user,
        action="bim_revit_settings.update",
        target_type="bim_revit_settings",
        target_key="bim_revit.v1",
        before=_masked_bim_revit_payload(before),
        after=masked,
    )
    db.commit()
    return {"ok": True, "settings": masked}


@router.post("/bim-revit/rotate-secret")
def rotate_bim_revit_secret(
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    before = get_bim_revit_integration(db)
    merged = dict(before)

    key_id = _norm(merged.get("plugin_key_id"))
    if not key_id:
        key_id = f"BIM-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    plain_secret = generate_plugin_secret(64)
    merged["plugin_key_id"] = key_id
    merged["plugin_secret_encrypted"] = encrypt_plugin_secret(
        plain_secret,
        secret_key=_runtime_secret_key_or_500(),
    )

    saved = set_bim_revit_integration(db, merged)
    masked = _masked_bim_revit_payload(saved)
    _audit_log(
        db,
        actor=current_user,
        action="bim_revit_settings.rotate_secret",
        target_type="bim_revit_settings",
        target_key="bim_revit.v1",
        before=_masked_bim_revit_payload(before),
        after=masked,
    )
    db.commit()
    return {
        "ok": True,
        "plugin_key_id": key_id,
        "plugin_secret": plain_secret,
        "settings": masked,
    }


@router.get("/site-log-catalogs")
def get_site_log_catalogs(
    db: Session = Depends(get_db),
):
    return {
        "ok": True,
        "catalogs": _load_site_log_catalogs_payload(db),
        "catalog_titles": dict(SITE_LOG_CATALOG_TITLES),
    }


@router.post("/site-log-catalogs/upsert")
def upsert_site_log_catalog(
    payload: SiteLogCatalogUpsertIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    catalog_type = _normalize_site_log_catalog_type_or_400(payload.catalog_type)
    model = _site_log_catalog_model(catalog_type)
    code = _upper(payload.code)
    label = _norm(payload.label)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")
    if not label:
        raise HTTPException(status_code=400, detail="label is required")

    row = None
    if payload.id:
        row = db.query(model).filter(model.id == payload.id).first()
        if not row:
            raise HTTPException(status_code=404, detail="Catalog item not found")

    duplicate = (
        db.query(model)
        .filter(func.upper(model.code) == code)
        .first()
    )
    if duplicate and (not row or int(duplicate.id) != int(row.id)):
        raise HTTPException(status_code=409, detail="Catalog code already exists")

    before = _serialize_site_log_catalog_row(row) if row else None
    if not row:
        row = model(code=code)
        db.add(row)

    row.code = code
    row.label = label
    row.sort_order = int(payload.sort_order or 0)
    row.is_active = bool(payload.is_active)

    db.flush()
    after = _serialize_site_log_catalog_row(row)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_catalog.upsert",
        target_type=f"site_log_catalog.{catalog_type}",
        target_key=str(after["id"]),
        before=before,
        after=after,
    )
    db.commit()
    return {
        "ok": True,
        "item": after,
        "catalog_type": catalog_type,
        "catalogs": _load_site_log_catalogs_payload(db),
    }


@router.post("/site-log-catalogs/bulk-upsert")
def bulk_upsert_site_log_catalog(
    payload: SiteLogCatalogBulkUpsertIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    catalog_type = _normalize_site_log_catalog_type_or_400(payload.catalog_type)
    if catalog_type not in SITE_LOG_BULK_CATALOG_PREFIXES:
        raise HTTPException(status_code=400, detail="Bulk add is only supported for material and equipment catalogs")

    model = _site_log_catalog_model(catalog_type)
    overwrite_existing = bool(payload.overwrite_existing)
    existing_by_code = {
        _upper(row.code): row
        for row in db.query(model).all()
        if _upper(row.code)
    }
    used_codes = set(existing_by_code.keys())
    created: List[Dict[str, Any]] = []
    updated: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    touched_rows: List[Any] = []
    seen_payload_codes: set[str] = set()

    for index, item in enumerate(payload.items, start=1):
        label = _norm(item.label)
        if not label:
            skipped.append({"row": index, "reason": "label_required"})
            continue

        code = _upper(item.code)
        if not code:
            code = _next_site_log_bulk_code(db, model, SITE_LOG_BULK_CATALOG_PREFIXES[catalog_type], used_codes)
        elif code in seen_payload_codes:
            skipped.append({"row": index, "code": code, "label": label, "reason": "duplicate_in_payload"})
            continue
        elif len(code) > 32:
            skipped.append({"row": index, "code": code, "label": label, "reason": "code_too_long"})
            continue
        else:
            used_codes.add(code)

        seen_payload_codes.add(code)
        row = existing_by_code.get(code)
        if row and not overwrite_existing:
            skipped.append({"row": index, "code": code, "label": label, "reason": "duplicate_code"})
            continue

        before = _serialize_site_log_catalog_row(row) if row else None
        if row is None:
            row = model(code=code)
            db.add(row)

        row.code = code
        row.label = label
        row.sort_order = int(item.sort_order if item.sort_order is not None else index * 10)
        row.is_active = bool(item.is_active)
        db.flush()

        after = _serialize_site_log_catalog_row(row)
        touched_rows.append(row)
        if before:
            updated.append(after)
        else:
            created.append(after)

    after_items = [_serialize_site_log_catalog_row(row) for row in touched_rows]
    _audit_log(
        db,
        actor=current_user,
        action="site_log_catalog.bulk_upsert",
        target_type=f"site_log_catalog.{catalog_type}",
        target_key=None,
        before=None,
        after={
            "created": created,
            "updated": updated,
            "skipped": skipped,
        },
    )
    db.commit()
    return {
        "ok": True,
        "catalog_type": catalog_type,
        "created": len(created),
        "updated": len(updated),
        "skipped": len(skipped),
        "items": after_items,
        "errors": skipped,
        "catalogs": _load_site_log_catalogs_payload(db),
    }


@router.post("/site-log-catalogs/delete")
def delete_site_log_catalog(
    payload: SiteLogCatalogDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    catalog_type = _normalize_site_log_catalog_type_or_400(payload.catalog_type)
    model = _site_log_catalog_model(catalog_type)
    row = db.query(model).filter(model.id == payload.id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Catalog item not found")

    before = _serialize_site_log_catalog_row(row)
    row.is_active = False
    db.flush()
    after = _serialize_site_log_catalog_row(row)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_catalog.delete",
        target_type=f"site_log_catalog.{catalog_type}",
        target_key=str(payload.id),
        before=before,
        after=after,
    )
    db.commit()
    return {
        "ok": True,
        "catalog_type": catalog_type,
        "item": after,
        "catalogs": _load_site_log_catalogs_payload(db),
    }


@router.get("/site-log-activity-catalog")
def get_site_log_activity_catalog(
    project_code: Optional[str] = Query(default=None),
    organization_id: Optional[int] = Query(default=None, ge=1),
    organization_contract_id: Optional[int] = Query(default=None, ge=1),
    pms_status: Optional[str] = Query(default=None),
    pms_template_id: Optional[int] = Query(default=None, ge=1),
    default_unit: Optional[str] = Query(default=None),
    default_location: Optional[str] = Query(default=None),
    reference_search: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    project_value = _check_optional_project_or_404(db, project_code)
    organization_value = _check_optional_organization_or_404(db, organization_id)
    contract = None
    if organization_contract_id:
        contract = (
            db.query(OrganizationContract)
            .options(joinedload(OrganizationContract.block))
            .filter(OrganizationContract.id == int(organization_contract_id))
            .first()
        )
        if not contract:
            raise HTTPException(status_code=404, detail="Organization contract not found")
        if organization_value and int(contract.organization_id or 0) != int(organization_value):
            raise HTTPException(status_code=400, detail="Selected contract does not belong to the selected organization.")
        if organization_value is None:
            organization_value = int(contract.organization_id or 0) or None
    return {
        "ok": True,
        **_load_site_log_activity_catalog_payload(
            db,
            project_code=project_value,
            organization_id=organization_value,
            organization_contract_id=int(contract.id) if contract else None,
            pms_status=pms_status,
            pms_template_id=pms_template_id,
            default_unit=default_unit,
            default_location=default_location,
            reference_search=reference_search,
        ),
    }


@router.post("/site-log-activity-catalog/upsert")
def upsert_site_log_activity_catalog(
    payload: SiteLogActivityCatalogUpsertIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    project_value = _check_optional_project_or_404(db, payload.project_code)
    if not project_value:
        raise HTTPException(status_code=400, detail="project_code is required")
    organization_value = _check_optional_organization_or_404(db, payload.organization_id)
    contract = None
    if payload.organization_contract_id:
        contract = (
            db.query(OrganizationContract)
            .options(joinedload(OrganizationContract.block))
            .filter(OrganizationContract.id == int(payload.organization_contract_id))
            .first()
        )
        if not contract:
            raise HTTPException(status_code=404, detail="Organization contract not found")
        if organization_value and int(contract.organization_id or 0) != int(organization_value):
            raise HTTPException(status_code=400, detail="Selected contract does not belong to the selected organization.")
        if organization_value is None:
            organization_value = int(contract.organization_id or 0) or None

    activity_code = _upper(payload.activity_code)
    activity_title = _norm(payload.activity_title)
    if not activity_code:
        raise HTTPException(status_code=400, detail="activity_code is required")
    if not activity_title:
        raise HTTPException(status_code=400, detail="activity_title is required")

    row = None
    if payload.id:
        row = db.query(SiteLogActivityCatalog).filter(SiteLogActivityCatalog.id == int(payload.id)).first()
        if not row:
            raise HTTPException(status_code=404, detail="Activity catalog item not found")

    duplicate_query = db.query(SiteLogActivityCatalog).filter(
        SiteLogActivityCatalog.project_code == project_value,
        func.upper(SiteLogActivityCatalog.activity_code) == activity_code,
    )
    if organization_value is None:
        duplicate_query = duplicate_query.filter(SiteLogActivityCatalog.organization_id.is_(None))
    else:
        duplicate_query = duplicate_query.filter(SiteLogActivityCatalog.organization_id == organization_value)
    if contract is None:
        duplicate_query = duplicate_query.filter(SiteLogActivityCatalog.organization_contract_id.is_(None))
    else:
        duplicate_query = duplicate_query.filter(SiteLogActivityCatalog.organization_contract_id == int(contract.id))
    duplicate = duplicate_query.first()
    if duplicate and (not row or int(duplicate.id) != int(row.id)):
        raise HTTPException(status_code=409, detail="Activity code already exists in the selected scope.")

    before = _serialize_site_log_activity_catalog_row(_load_site_log_activity_catalog_row_or_404(db, int(row.id))) if row else None
    if not row:
        row = SiteLogActivityCatalog(project_code=project_value, activity_code=activity_code)
        db.add(row)

    row.project_code = project_value
    row.organization_id = organization_value
    row.organization_contract_id = int(contract.id) if contract else None
    row.activity_code = activity_code
    row.activity_title = activity_title
    row.default_location = _norm(payload.default_location) or None
    row.default_unit = _norm(payload.default_unit) or None
    row.sort_order = int(payload.sort_order or 0)
    row.is_active = bool(payload.is_active)

    db.flush()
    after_row = _load_site_log_activity_catalog_row_or_404(db, int(row.id))
    after = _serialize_site_log_activity_catalog_row(after_row)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_catalog.upsert",
        target_type="site_log_activity_catalog",
        target_key=str(after["id"]),
        before=before,
        after=after,
    )
    db.commit()
    return {
        "ok": True,
        "item": after,
        **_load_site_log_activity_catalog_payload(
            db,
            project_code=project_value,
            organization_id=organization_value,
            organization_contract_id=int(contract.id) if contract else None,
        ),
    }


@router.get("/site-log-activity-catalog/template")
def download_site_log_activity_catalog_template(
    _: DbUser = Depends(require_permission("settings:read")),
):
    return Response(
        content=_site_log_activity_catalog_template_bytes(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="site_log_activity_catalog_template.xlsx"',
        },
    )


@router.post("/site-log-activity-catalog/import")
async def import_site_log_activity_catalog(
    project_code: str = Form(...),
    organization_id: Optional[int] = Form(default=None),
    organization_contract_id: Optional[int] = Form(default=None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    project_value, organization_value, contract = _resolve_activity_catalog_scope(
        db,
        project_code=project_code,
        organization_id=organization_id,
        organization_contract_id=organization_contract_id,
    )
    content = await file.read()
    rows = _parse_site_log_activity_catalog_import(content, str(file.filename or "activities.xlsx"))
    if not rows:
        raise HTTPException(status_code=400, detail="No activity rows found in the Excel file")

    scope_contract_id = int(contract.id) if contract else None
    max_sort_query = db.query(func.max(SiteLogActivityCatalog.sort_order)).filter(
        SiteLogActivityCatalog.project_code == project_value
    )
    if organization_value is None:
        max_sort_query = max_sort_query.filter(SiteLogActivityCatalog.organization_id.is_(None))
    else:
        max_sort_query = max_sort_query.filter(SiteLogActivityCatalog.organization_id == organization_value)
    if scope_contract_id is None:
        max_sort_query = max_sort_query.filter(SiteLogActivityCatalog.organization_contract_id.is_(None))
    else:
        max_sort_query = max_sort_query.filter(SiteLogActivityCatalog.organization_contract_id == scope_contract_id)
    max_sort_order = max_sort_query.scalar() or 0
    created = 0
    updated = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []
    imported_ids: List[int] = []

    for index, raw_row in enumerate(rows, start=1):
        row_no = int(raw_row.get("row_no") or index)
        activity_code = _upper(raw_row.get("activity_code"))
        activity_title = _norm(raw_row.get("activity_title"))
        if not activity_code or not activity_title:
            skipped += 1
            errors.append(
                {
                    "row_no": row_no,
                    "message": "activity_code and activity_title are required",
                }
            )
            continue

        row = _find_activity_catalog_by_scope_code(
            db,
            project_code=project_value,
            organization_id=organization_value,
            organization_contract_id=scope_contract_id,
            activity_code=activity_code,
        )
        is_new = row is None
        if row is None:
            row = SiteLogActivityCatalog(project_code=project_value, activity_code=activity_code)
            db.add(row)
            created += 1
        else:
            updated += 1

        row.project_code = project_value
        row.organization_id = organization_value
        row.organization_contract_id = scope_contract_id
        row.activity_code = activity_code
        row.activity_title = activity_title
        row.default_location = _norm(raw_row.get("default_location")) or None
        row.default_unit = _norm(raw_row.get("default_unit")) or None
        row.sort_order = _activity_import_int(raw_row.get("sort_order"), max_sort_order + (index * 10))
        row.is_active = _activity_import_bool(raw_row.get("is_active"), True)
        db.flush()
        imported_ids.append(int(row.id or 0))
        if is_new:
            max_sort_order = max(max_sort_order, int(row.sort_order or 0))

    if not imported_ids:
        first_error = _norm(errors[0].get("message")) if errors else ""
        detail = "No valid activity rows were imported"
        if first_error:
            detail = f"{detail}: {first_error}"
        raise HTTPException(status_code=400, detail=detail)

    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_catalog.import",
        target_type="site_log_activity_catalog",
        target_key=f"{project_value}:{organization_value or 'project'}:{scope_contract_id or 'all'}",
        before=None,
        after={
            "source_file": file.filename,
            "project_code": project_value,
            "organization_id": organization_value,
            "organization_contract_id": scope_contract_id,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:20],
        },
    )
    db.commit()
    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "imported_ids": imported_ids,
        **_load_site_log_activity_catalog_payload(
            db,
            project_code=project_value,
            organization_id=organization_value,
            organization_contract_id=scope_contract_id,
        ),
    }


@router.post("/site-log-activity-catalog/delete")
def delete_site_log_activity_catalog(
    payload: SiteLogActivityCatalogDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    row = _load_site_log_activity_catalog_row_or_404(db, int(payload.id))
    before = _serialize_site_log_activity_catalog_row(row)
    row.is_active = False
    db.flush()
    after_row = _load_site_log_activity_catalog_row_or_404(db, int(payload.id))
    after = _serialize_site_log_activity_catalog_row(after_row)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_catalog.delete",
        target_type="site_log_activity_catalog",
        target_key=str(payload.id),
        before=before,
        after=after,
    )
    db.commit()
    return {
        "ok": True,
        "item": after,
        **_load_site_log_activity_catalog_payload(
            db,
            project_code=_upper(row.project_code),
            organization_id=int(row.organization_id or 0) or None,
            organization_contract_id=int(row.organization_contract_id or 0) or None,
        ),
    }


@router.get("/site-log-pms/templates")
def list_site_log_pms_templates(
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("settings:read")),
):
    templates = _load_pms_templates(db)
    return {
        "ok": True,
        "items": [_serialize_pms_template(row) for row in templates],
    }


@router.post("/site-log-pms/templates/upsert")
def upsert_site_log_pms_template(
    payload: SiteLogPmsTemplateUpsertIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    title = _norm(payload.title)
    if not code:
        raise HTTPException(status_code=400, detail="PMS Template code is required")
    if not title:
        raise HTTPException(status_code=400, detail="PMS Template title is required")
    step_codes: set[str] = set()
    for step in payload.steps:
        step_code = _upper(step.step_code)
        if step_code in step_codes:
            raise HTTPException(status_code=400, detail=f"Duplicate PMS Step code: {step_code}")
        step_codes.add(step_code)

    row = None
    if payload.id:
        row = (
            db.query(SiteLogPmsTemplate)
            .options(selectinload(SiteLogPmsTemplate.steps))
            .filter(SiteLogPmsTemplate.id == int(payload.id))
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="PMS Template not found")

    duplicate = db.query(SiteLogPmsTemplate).filter(func.upper(SiteLogPmsTemplate.code) == code).first()
    if duplicate and (not row or int(duplicate.id) != int(row.id)):
        raise HTTPException(status_code=409, detail="PMS Template code already exists")

    before = _serialize_pms_template(row) if row else None
    if not row:
        row = SiteLogPmsTemplate(code=code, version=1)
        db.add(row)
    else:
        row.version = int(row.version or 1) + 1
    row.code = code
    row.title = title
    row.description = _norm(payload.description) or None
    row.sort_order = int(payload.sort_order or 0)
    row.is_active = bool(payload.is_active)
    row.updated_at = datetime.utcnow()
    if int(row.id or 0) > 0 and row.steps:
        row.steps[:] = []
        db.flush()
    for index, step in enumerate(payload.steps):
        row.steps.append(
            SiteLogPmsTemplateStep(
                step_code=_upper(step.step_code),
                step_title=_norm(step.step_title),
                weight_pct=float(step.weight_pct or 0),
                sort_order=int(step.sort_order if step.sort_order is not None else index),
                is_active=bool(step.is_active),
            )
        )
    db.flush()
    after = _serialize_pms_template(row)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_pms_template.upsert",
        target_type="site_log_pms_template",
        target_key=str(after["id"]),
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "item": after, "items": [_serialize_pms_template(row) for row in _load_pms_templates(db)]}


@router.post("/site-log-pms/templates/delete")
def delete_site_log_pms_template(
    payload: SiteLogPmsTemplateDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    row = _load_pms_template_or_404(db, template_id=payload.id)
    before = _serialize_pms_template(row)
    row.is_active = False
    row.version = int(row.version or 1) + 1
    row.updated_at = datetime.utcnow()
    db.flush()
    after = _serialize_pms_template(row)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_pms_template.delete",
        target_type="site_log_pms_template",
        target_key=str(payload.id),
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "item": after, "items": [_serialize_pms_template(row) for row in _load_pms_templates(db)]}


@router.post("/site-log-pms/mappings/apply")
def apply_site_log_activity_pms_mapping(
    payload: SiteLogPmsMappingApplyIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    template = _load_pms_template_or_404(db, template_id=payload.template_id, template_code=payload.template_code)
    _ensure_pms_template_ready(template)
    activity_ids = sorted({int(item) for item in payload.activity_ids if int(item or 0) > 0})
    activities = (
        db.query(SiteLogActivityCatalog)
        .options(
            joinedload(SiteLogActivityCatalog.pms_mapping).joinedload(SiteLogActivityPmsMapping.template),
            joinedload(SiteLogActivityCatalog.pms_mapping).selectinload(SiteLogActivityPmsMapping.steps),
        )
        .filter(SiteLogActivityCatalog.id.in_(activity_ids))
        .all()
    )
    found_ids = {int(row.id or 0) for row in activities}
    missing = [item for item in activity_ids if item not in found_ids]
    if missing:
        raise HTTPException(status_code=404, detail=f"Activity catalog items not found: {missing}")
    existing = [row.activity_code for row in activities if row.pms_mapping is not None]
    if existing and not payload.overwrite:
        raise HTTPException(status_code=409, detail=f"Activities already have PMS: {', '.join(existing[:10])}")

    before = [_serialize_site_log_activity_catalog_row(row) for row in activities]
    for activity in activities:
        _copy_template_to_activity_mapping(db, activity=activity, template=template, overwrite=True)
    db.flush()
    refreshed = [
        _load_site_log_activity_catalog_row_or_404(db, int(activity.id or 0))
        for activity in activities
    ]
    after = [_serialize_site_log_activity_catalog_row(row) for row in refreshed]
    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_pms_mapping.apply",
        target_type="site_log_activity_pms_mapping",
        target_key=",".join(str(item) for item in activity_ids),
        before=before,
        after={"template": _serialize_pms_template(template, include_steps=False), "activities": after},
    )
    db.commit()
    return {"ok": True, "applied": len(activities), "items": after}


@router.post("/site-log-pms/mappings/delete")
def delete_site_log_activity_pms_mapping(
    payload: SiteLogPmsMappingDeleteIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    activity = _load_site_log_activity_catalog_row_or_404(db, int(payload.activity_id))
    before = _serialize_site_log_activity_catalog_row(activity)
    if activity.pms_mapping:
        db.delete(activity.pms_mapping)
        activity.pms_mapping = None
    db.flush()
    after = _serialize_site_log_activity_catalog_row(_load_site_log_activity_catalog_row_or_404(db, int(activity.id)))
    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_pms_mapping.delete",
        target_type="site_log_activity_pms_mapping",
        target_key=str(payload.activity_id),
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "item": after}


@router.post("/site-log-pms/mappings/reapply")
def reapply_site_log_activity_pms_mapping(
    payload: SiteLogPmsMappingReapplyIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    activity_ids = sorted({int(item) for item in payload.activity_ids if int(item or 0) > 0})
    activities = (
        db.query(SiteLogActivityCatalog)
        .options(
            joinedload(SiteLogActivityCatalog.pms_mapping).joinedload(SiteLogActivityPmsMapping.template),
            joinedload(SiteLogActivityCatalog.pms_mapping).selectinload(SiteLogActivityPmsMapping.steps),
        )
        .filter(SiteLogActivityCatalog.id.in_(activity_ids))
        .all()
    )
    if len(activities) != len(activity_ids):
        found_ids = {int(row.id or 0) for row in activities}
        missing = [item for item in activity_ids if item not in found_ids]
        raise HTTPException(status_code=404, detail=f"Activity catalog items not found: {missing}")
    missing_pms = [row.activity_code for row in activities if not row.pms_mapping]
    if missing_pms:
        raise HTTPException(status_code=400, detail=f"Activities without PMS cannot be reapplied: {', '.join(missing_pms[:10])}")
    before = [_serialize_site_log_activity_catalog_row(row) for row in activities]
    for activity in activities:
        assert activity.pms_mapping is not None
        _copy_template_to_activity_mapping(db, activity=activity, template=activity.pms_mapping.template, overwrite=True)
    db.flush()
    after = [
        _serialize_site_log_activity_catalog_row(_load_site_log_activity_catalog_row_or_404(db, int(activity.id or 0)))
        for activity in activities
    ]
    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_pms_mapping.reapply",
        target_type="site_log_activity_pms_mapping",
        target_key=",".join(str(item) for item in activity_ids),
        before=before,
        after=after,
    )
    db.commit()
    return {"ok": True, "reapplied": len(activities), "items": after}


@router.get("/site-log-pms/mappings/template")
def download_site_log_pms_mapping_template(
    _: DbUser = Depends(require_permission("settings:read")),
):
    return Response(
        content=_site_log_pms_mapping_workbook_bytes(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="site_log_pms_mapping_template.xlsx"'},
    )


@router.get("/site-log-pms/mappings/export")
def export_site_log_pms_mappings(
    project_code: Optional[str] = Query(default=None),
    organization_id: Optional[int] = Query(default=None, ge=1),
    organization_contract_id: Optional[int] = Query(default=None, ge=1),
    pms_status: Optional[str] = Query(default=None),
    pms_template_id: Optional[int] = Query(default=None, ge=1),
    default_unit: Optional[str] = Query(default=None),
    default_location: Optional[str] = Query(default=None),
    reference_search: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("settings:read")),
):
    payload = _load_site_log_activity_catalog_payload(
        db,
        project_code=_check_optional_project_or_404(db, project_code),
        organization_id=_check_optional_organization_or_404(db, organization_id),
        organization_contract_id=organization_contract_id,
        pms_status=pms_status,
        pms_template_id=pms_template_id,
        default_unit=default_unit,
        default_location=default_location,
        reference_search=reference_search,
    )
    return Response(
        content=_site_log_pms_mapping_workbook_bytes(list(payload.get("items") or [])),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="site_log_pms_mappings.xlsx"'},
    )


@router.post("/site-log-pms/mappings/import")
async def import_site_log_pms_mappings(
    project_code: str = Form(...),
    organization_id: Optional[int] = Form(default=None),
    organization_contract_id: Optional[int] = Form(default=None),
    overwrite: bool = Form(default=False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    project_value, organization_value, contract = _resolve_activity_catalog_scope(
        db,
        project_code=project_code,
        organization_id=organization_id,
        organization_contract_id=organization_contract_id,
    )
    scope_contract_id = int(contract.id) if contract else None
    rows = _parse_site_log_pms_mapping_import(await file.read(), str(file.filename or "pms_mapping.xlsx"))
    if not rows:
        raise HTTPException(status_code=400, detail="No PMS mapping rows found in the Excel file")

    imported = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []
    imported_ids: List[int] = []
    for index, raw_row in enumerate(rows, start=1):
        row_no = int(raw_row.get("row_no") or index)
        activity_code = _upper(raw_row.get("activity_code"))
        template_code = _upper(raw_row.get("pms_template_code"))
        if not activity_code or not template_code:
            skipped += 1
            errors.append({"row_no": row_no, "message": "activity_code and pms_template_code are required"})
            continue
        activity = _find_activity_catalog_by_scope_code(
            db,
            project_code=project_value,
            organization_id=organization_value,
            organization_contract_id=scope_contract_id,
            activity_code=activity_code,
        )
        if not activity:
            skipped += 1
            errors.append({"row_no": row_no, "message": f"Activity not found: {activity_code}"})
            continue
        try:
            template = _load_pms_template_or_404(db, template_code=template_code)
            _copy_template_to_activity_mapping(db, activity=activity, template=template, overwrite=overwrite)
            imported += 1
            imported_ids.append(int(activity.id or 0))
        except HTTPException as exc:
            skipped += 1
            errors.append({"row_no": row_no, "message": str(exc.detail)})

    if not imported:
        detail = "No valid PMS mappings were imported"
        if errors:
            detail = f"{detail}: {errors[0].get('message')}"
        raise HTTPException(status_code=400, detail=detail)
    _audit_log(
        db,
        actor=current_user,
        action="site_log_activity_pms_mapping.import",
        target_type="site_log_activity_pms_mapping",
        target_key=f"{project_value}:{organization_value or 'project'}:{scope_contract_id or 'all'}",
        before=None,
        after={
            "source_file": file.filename,
            "project_code": project_value,
            "organization_id": organization_value,
            "organization_contract_id": scope_contract_id,
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:20],
        },
    )
    db.commit()
    return {
        "ok": True,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "imported_ids": imported_ids,
        **_load_site_log_activity_catalog_payload(
            db,
            project_code=project_value,
            organization_id=organization_value,
            organization_contract_id=scope_contract_id,
        ),
    }


@router.get("/permissions/matrix")
def get_permissions_matrix(
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:read")),
):
    category_key = _normalize_permission_category_or_400(category)
    matrix = _load_permission_matrix(db, category=category_key)
    baseline_matrix = _default_permission_matrix(category_key)
    return {
        "ok": True,
        "category": category_key,
        "category_label": _permission_category_labels().get(category_key, category_key),
        "categories": list(CANONICAL_PERMISSION_CATEGORIES),
        "read_only": False,
        "roles": list(CANONICAL_MATRIX_ROLES),
        "role_labels": _matrix_role_labels(),
        "permissions": _permission_keys(),
        "permissions_meta": _permission_meta(),
        "feature_catalog": _feature_catalog(),
        "baseline_matrix": baseline_matrix,
        "matrix": matrix,
    }


@router.post("/permissions/matrix")
def save_permissions_matrix(
    payload: PermissionMatrixIn,
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("permissions:update")),
):
    category_key = _normalize_permission_category_or_400(category)
    before = _load_permission_matrix(db, category=category_key)
    matrix = _normalize_permission_matrix(payload.matrix, category_key)
    perms = _permission_keys()
    db.query(RoleCategoryPermission).filter(
        RoleCategoryPermission.category == category_key
    ).delete(synchronize_session=False)
    for role in MATRIX_ROLES:
        for perm in perms:
            db.add(
                RoleCategoryPermission(
                    category=category_key,
                    role=role,
                    permission=perm,
                    allowed=bool(matrix[role][perm]),
                )
            )
    _audit_log(
        db,
        actor=current_user,
        action="permissions.matrix.save",
        target_type="permissions_matrix_category",
        target_key=category_key,
        before={"category": category_key, "matrix": before},
        after={"category": category_key, "matrix": matrix},
    )
    db.commit()
    return {
        "ok": True,
        "message": "Permissions matrix saved",
        "category": category_key,
        "matrix": matrix,
    }


@router.get("/permissions/scope")
def get_permissions_scope(
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:read")),
):
    category_key = _normalize_permission_category_or_400(category)
    scope = _load_scope_rules(db, category=category_key)
    projects = [
        {"code": p.code, "name": p.name_e or p.name_p or p.code}
        for p in db.query(Project).filter(Project.is_active == True).order_by(Project.code).all()
    ]
    disciplines = [
        {"code": d.code, "name": d.name_e or d.name_p or d.code}
        for d in db.query(Discipline).order_by(Discipline.code).all()
    ]
    return {
        "ok": True,
        "category": category_key,
        "category_label": _permission_category_labels().get(category_key, category_key),
        "categories": list(CANONICAL_PERMISSION_CATEGORIES),
        "roles": list(CANONICAL_MATRIX_ROLES),
        "role_labels": _matrix_role_labels(),
        "scope_read_only": False,
        "scope": scope,
        "projects": projects,
        "disciplines": disciplines,
    }


@router.post("/permissions/scope")
def save_permissions_scope(
    payload: PermissionScopeIn,
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("permissions:update")),
):
    category_key = _normalize_permission_category_or_400(category)
    before = _load_scope_rules(db, category=category_key)
    scope = _normalize_scope_rules(payload.scope)
    project_codes = {code for (code,) in db.query(Project.code).all()}
    discipline_codes = {code for (code,) in db.query(Discipline.code).all()}

    db.query(RoleCategoryProjectScope).filter(
        RoleCategoryProjectScope.category == category_key
    ).delete(synchronize_session=False)
    db.query(RoleCategoryDisciplineScope).filter(
        RoleCategoryDisciplineScope.category == category_key
    ).delete(synchronize_session=False)

    for role in MATRIX_ROLES:
        for code in scope.get(role, {}).get("projects", []):
            if code in project_codes:
                db.add(
                    RoleCategoryProjectScope(
                        category=category_key,
                        role=role,
                        project_code=code,
                    )
                )
        for code in scope.get(role, {}).get("disciplines", []):
            if code in discipline_codes:
                db.add(
                    RoleCategoryDisciplineScope(
                        category=category_key,
                        role=role,
                        discipline_code=code,
                    )
                )
    _audit_log(
        db,
        actor=current_user,
        action="permissions.scope.save",
        target_type="role_scope_category",
        target_key=category_key,
        before={"category": category_key, "scope": before},
        after={"category": category_key, "scope": scope},
    )
    db.commit()
    return {
        "ok": True,
        "message": "Permission scope saved",
        "category": category_key,
        "scope": scope,
    }


@router.get("/permissions/user-scope")
def get_permissions_user_scope(
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:read")),
):
    scope = _load_user_scope_rules(db)
    users = db.query(DbUser).order_by(DbUser.id).all()
    return {
        "ok": True,
        "users": [
            {
                "id": u.id,
                "email": u.email,
                "full_name": u.full_name,
                "role": u.role,
                "is_active": u.is_active,
            }
            for u in users
        ],
        "scope": scope,
    }


@router.post("/permissions/user-scope/upsert")
def upsert_permissions_user_scope(
    payload: UserPermissionScopeIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("permissions:update")),
):
    user = db.query(DbUser).filter(DbUser.id == payload.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    before_all_scope = _load_user_scope_rules(db)
    before_user_scope = before_all_scope.get(str(payload.user_id), {"projects": [], "disciplines": []})
    normalized_projects = _normalize_scope_values(payload.projects)
    normalized_disciplines = _normalize_scope_values(payload.disciplines)
    project_codes = {code for (code,) in db.query(Project.code).all()}
    discipline_codes = {code for (code,) in db.query(Discipline.code).all()}

    db.query(UserProjectScope).filter(UserProjectScope.user_id == payload.user_id).delete(synchronize_session=False)
    db.query(UserDisciplineScope).filter(UserDisciplineScope.user_id == payload.user_id).delete(synchronize_session=False)

    saved_projects: List[str] = []
    for code in normalized_projects:
        if code in project_codes:
            db.add(UserProjectScope(user_id=payload.user_id, project_code=code))
            saved_projects.append(code)

    saved_disciplines: List[str] = []
    for code in normalized_disciplines:
        if code in discipline_codes:
            db.add(UserDisciplineScope(user_id=payload.user_id, discipline_code=code))
            saved_disciplines.append(code)

    after_user_scope = {
        "projects": sorted(set(saved_projects)),
        "disciplines": sorted(set(saved_disciplines)),
    }
    _audit_log(
        db,
        actor=current_user,
        action="permissions.user_scope.upsert",
        target_type="user_scope",
        target_key=str(payload.user_id),
        before=before_user_scope,
        after=after_user_scope,
    )
    db.commit()
    return {
        "ok": True,
        "message": "User scope saved",
        "user_id": payload.user_id,
        "scope": after_user_scope,
    }


@router.get("/permissions/access-report")
def permissions_access_report(
    project_code: Optional[str] = Query(default=None),
    discipline_code: Optional[str] = Query(default=None),
    include_inactive: bool = Query(default=False),
    include_denied: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:audit_read")),
):
    project = _upper(project_code)
    discipline = _upper(discipline_code)
    if not project and not discipline:
        raise HTTPException(
            status_code=400,
            detail="At least one filter is required: project_code or discipline_code",
        )
    items = _build_access_report_items(
        db,
        project_code=project,
        discipline_code=discipline,
        include_inactive=include_inactive,
        include_denied=include_denied,
    )

    return {
        "ok": True,
        "filters": {"project_code": project or None, "discipline_code": discipline or None},
        "count": len(items),
        "items": items,
    }


@router.get("/permissions/access-report.csv")
def permissions_access_report_csv(
    project_code: Optional[str] = Query(default=None),
    discipline_code: Optional[str] = Query(default=None),
    include_inactive: bool = Query(default=False),
    include_denied: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:audit_read")),
):
    project = _upper(project_code)
    discipline = _upper(discipline_code)
    if not project and not discipline:
        raise HTTPException(
            status_code=400,
            detail="At least one filter is required: project_code or discipline_code",
        )

    items = _build_access_report_items(
        db,
        project_code=project,
        discipline_code=discipline,
        include_inactive=include_inactive,
        include_denied=include_denied,
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "user_id",
            "email",
            "full_name",
            "role",
            "category",
            "is_active",
            "has_access",
            "project_allowed",
            "discipline_allowed",
            "projects_restricted",
            "disciplines_restricted",
            "effective_projects",
            "effective_disciplines",
            "project_filter",
            "discipline_filter",
        ]
    )
    for item in items:
        writer.writerow(
            [
                item.get("user_id"),
                item.get("email"),
                item.get("full_name"),
                item.get("role"),
                item.get("category"),
                item.get("is_active"),
                item.get("has_access"),
                item.get("project_allowed"),
                item.get("discipline_allowed"),
                item.get("projects_restricted"),
                item.get("disciplines_restricted"),
                ",".join(item.get("effective_projects", [])),
                ",".join(item.get("effective_disciplines", [])),
                project or "",
                discipline or "",
            ]
        )

    filename = f"permissions_access_report_{project or 'ALL'}_{discipline or 'ALL'}.csv"
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/permissions/user-access/{user_id}")
def permissions_user_access_report(
    user_id: int,
    include_catalog: bool = Query(default=True),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:audit_read")),
):
    user = (
        db.query(DbUser)
        .options(joinedload(DbUser.organization))
        .filter(DbUser.id == user_id)
        .first()
    )
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    access = resolve_effective_access(user)
    role = access.effective_role
    category = access.permission_category
    permission_source = "full_access"
    if not access.full_access:
        has_category_rows = (
            db.query(RoleCategoryPermission.id)
            .filter(
                RoleCategoryPermission.category == category,
                RoleCategoryPermission.role == role,
            )
            .first()
            is not None
        )
        if has_category_rows:
            permission_source = "category_matrix"
        else:
            has_role_rows = (
                db.query(RolePermission.id)
                .filter(RolePermission.role == role)
                .first()
                is not None
            )
            permission_source = "role_matrix" if has_role_rows else "static_fallback"
    allowed = _load_allowed_permissions(db, role, category=category)
    capabilities = {
        permission: ("*" in allowed or permission in allowed)
        for permission in _permission_keys()
    }
    navigation = build_navigation_state(
        capabilities,
        category=category,
        effective_role=role,
    )
    role_scope = _load_scope_rules(db, category=category).get(role, {"projects": [], "disciplines": []})
    user_scope = _load_user_scope_rules(db).get(str(user_id), {"projects": [], "disciplines": []})
    projects, projects_restricted = _effective_scope_values(
        role_scope.get("projects", []),
        user_scope.get("projects", []),
    )
    disciplines, disciplines_restricted = _effective_scope_values(
        role_scope.get("disciplines", []),
        user_scope.get("disciplines", []),
    )

    payload: Dict[str, Any] = {
        "ok": True,
        "category": category,
        "user": {
            "id": user.id,
            "email": user.email,
            "full_name": user.full_name,
            "role": role,
            "organization_role": user.organization_role,
            "effective_role": role,
            "category": category,
            "organization_type": access.organization_type,
            "is_system_admin": access.is_system_admin,
            "is_active": user.is_active,
            "permission_source": permission_source,
        },
        "capabilities": capabilities,
        "granted_permissions": sorted(permission for permission, is_allowed in capabilities.items() if is_allowed),
        "denied_permissions_sample": sorted(permission for permission, is_allowed in capabilities.items() if not is_allowed)[:25],
        "navigation": navigation,
        "navigation_diagnostics": build_navigation_diagnostics(navigation, category=category),
        "role_scope": role_scope,
        "user_scope": user_scope,
        "effective_scope": {
            "projects": projects,
            "disciplines": disciplines,
            "projects_restricted": projects_restricted,
            "disciplines_restricted": disciplines_restricted,
        },
    }

    if include_catalog:
        all_projects = {
            code: (name_e or name_p or code)
            for code, name_e, name_p in db.query(Project.code, Project.name_e, Project.name_p).all()
        }
        all_disciplines = {
            code: (name_e or name_p or code)
            for code, name_e, name_p in db.query(Discipline.code, Discipline.name_e, Discipline.name_p).all()
        }
        payload["effective_scope_catalog"] = {
            "projects": [{"code": c, "name": all_projects.get(c, c)} for c in projects],
            "disciplines": [{"code": c, "name": all_disciplines.get(c, c)} for c in disciplines],
        }

    return payload


@router.get("/audit-logs")
@router.get("/permissions/audit-logs")
def permissions_audit_logs(
    limit: Optional[int] = Query(default=None, ge=1, le=1000),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=100, ge=1, le=1000),
    offset: Optional[int] = Query(default=None, ge=0),
    action: Optional[str] = Query(default=None),
    target_type: Optional[str] = Query(default=None),
    target_key: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: DbUser = Depends(require_permission("permissions:audit_read")),
):
    q = db.query(SettingsAuditLog)
    if action:
        q = q.filter(SettingsAuditLog.action == _norm(action))
    if target_type:
        q = q.filter(SettingsAuditLog.target_type == _norm(target_type))
    if target_key:
        q = q.filter(SettingsAuditLog.target_key == _norm(target_key))

    effective_page_size = int(limit) if limit is not None else int(page_size)
    effective_offset = int(offset) if offset is not None else (int(page) - 1) * effective_page_size
    total = q.count()
    rows = (
        q.order_by(SettingsAuditLog.created_at.desc())
        .offset(effective_offset)
        .limit(effective_page_size)
        .all()
    )

    total_pages = max(1, (total + effective_page_size - 1) // effective_page_size)
    current_page = (effective_offset // effective_page_size) + 1
    return {
        "ok": True,
        "pagination": {
            "total": total,
            "page": current_page,
            "page_size": effective_page_size,
            "offset": effective_offset,
            "count": len(rows),
            "total_pages": total_pages,
        },
        "items": [
            {
                "id": row.id,
                "action": row.action,
                "target_type": row.target_type,
                "target_key": row.target_key,
                "actor_user_id": row.actor_user_id,
                "actor_email": row.actor_email,
                "actor_name": row.actor_name,
                "before_json": row.before_json,
                "after_json": row.after_json,
                "created_at": row.created_at.isoformat() if row.created_at else None,
            }
            for row in rows
        ],
    }


@router.get("/workflow-statuses")
def list_workflow_statuses(
    db: Session = Depends(get_db),
):
    rows = (
        db.query(WorkflowStatus)
        .order_by(
            WorkflowStatus.item_type.asc(),
            WorkflowStatus.sort_order.asc(),
            WorkflowStatus.code.asc(),
        )
        .all()
    )
    return {
        "ok": True,
        "data": [
            {
                "id": row.id,
                "item_type": row.item_type,
                "code": row.code,
                "label": row.label,
                "is_terminal": bool(row.is_terminal),
                "sort_order": row.sort_order,
                "is_active": bool(row.is_active),
            }
            for row in rows
        ],
    }


@router.post("/workflow-statuses")
def upsert_workflow_status(
    payload: WorkflowStatusIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    item_type = _normalize_workflow_item_type_or_400(payload.item_type)
    code = _upper(payload.code)
    label = _norm(payload.label)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")
    if not label:
        raise HTTPException(status_code=400, detail="label is required")

    row = None
    if payload.id:
        row = db.query(WorkflowStatus).filter(WorkflowStatus.id == payload.id).first()
    if row is None:
        row = (
            db.query(WorkflowStatus)
            .filter(WorkflowStatus.item_type == item_type, WorkflowStatus.code == code)
            .first()
        )

    before = _as_dict(
        row,
        ["id", "item_type", "code", "label", "is_terminal", "sort_order", "is_active"],
    )
    if row is None:
        row = WorkflowStatus(item_type=item_type, code=code)
        db.add(row)
    row.item_type = item_type
    row.code = code
    row.label = label
    row.is_terminal = bool(payload.is_terminal)
    row.sort_order = int(payload.sort_order)
    row.is_active = bool(payload.is_active)
    db.flush()

    _audit_log(
        db,
        actor=current_user,
        action="workflow.status.upsert",
        target_type="workflow_status",
        target_key=f"{row.item_type}:{row.code}",
        before=before,
        after=_as_dict(
            row,
            ["id", "item_type", "code", "label", "is_terminal", "sort_order", "is_active"],
        ),
    )
    db.commit()
    return {
        "ok": True,
        "data": {
            "id": row.id,
            "item_type": row.item_type,
            "code": row.code,
            "label": row.label,
            "is_terminal": bool(row.is_terminal),
            "sort_order": row.sort_order,
            "is_active": bool(row.is_active),
        },
    }


@router.get("/workflow-transitions")
def list_workflow_transitions(
    db: Session = Depends(get_db),
):
    rows = (
        db.query(WorkflowTransition)
        .order_by(
            WorkflowTransition.item_type.asc(),
            WorkflowTransition.from_status_code.asc(),
            WorkflowTransition.to_status_code.asc(),
        )
        .all()
    )
    return {
        "ok": True,
        "data": [
            {
                "id": row.id,
                "item_type": row.item_type,
                "from_status_code": row.from_status_code,
                "to_status_code": row.to_status_code,
                "requires_note": bool(row.requires_note),
                "is_active": bool(row.is_active),
            }
            for row in rows
        ],
    }


@router.post("/workflow-transitions")
def upsert_workflow_transition(
    payload: WorkflowTransitionIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    item_type = _normalize_workflow_item_type_or_400(payload.item_type)
    from_status = _upper(payload.from_status_code)
    to_status = _upper(payload.to_status_code)
    if not from_status or not to_status:
        raise HTTPException(status_code=400, detail="from_status_code and to_status_code are required")

    row = None
    if payload.id:
        row = db.query(WorkflowTransition).filter(WorkflowTransition.id == payload.id).first()
    if row is None:
        row = (
            db.query(WorkflowTransition)
            .filter(
                WorkflowTransition.item_type == item_type,
                WorkflowTransition.from_status_code == from_status,
                WorkflowTransition.to_status_code == to_status,
            )
            .first()
        )

    before = _as_dict(
        row,
        ["id", "item_type", "from_status_code", "to_status_code", "requires_note", "is_active"],
    )
    if row is None:
        row = WorkflowTransition(
            item_type=item_type,
            from_status_code=from_status,
            to_status_code=to_status,
        )
        db.add(row)
    row.item_type = item_type
    row.from_status_code = from_status
    row.to_status_code = to_status
    row.requires_note = bool(payload.requires_note)
    row.is_active = bool(payload.is_active)
    db.flush()

    _audit_log(
        db,
        actor=current_user,
        action="workflow.transition.upsert",
        target_type="workflow_transition",
        target_key=f"{row.item_type}:{row.from_status_code}->{row.to_status_code}",
        before=before,
        after=_as_dict(
            row,
            ["id", "item_type", "from_status_code", "to_status_code", "requires_note", "is_active"],
        ),
    )
    db.commit()
    return {
        "ok": True,
        "data": {
            "id": row.id,
            "item_type": row.item_type,
            "from_status_code": row.from_status_code,
            "to_status_code": row.to_status_code,
            "requires_note": bool(row.requires_note),
            "is_active": bool(row.is_active),
        },
    }


@router.get("/tech-subtypes")
def list_tech_subtypes(
    db: Session = Depends(get_db),
):
    rows = db.query(TechSubtype).order_by(TechSubtype.sort_order.asc(), TechSubtype.code.asc()).all()
    return {
        "ok": True,
        "data": [
            {
                "code": row.code,
                "label": row.label,
                "sort_order": row.sort_order,
                "is_active": bool(row.is_active),
            }
            for row in rows
        ],
    }


@router.post("/tech-subtypes")
def upsert_tech_subtype(
    payload: TechSubtypeIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    label = _norm(payload.label)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")
    if not label:
        raise HTTPException(status_code=400, detail="label is required")

    row = db.query(TechSubtype).filter(TechSubtype.code == code).first()
    before = _as_dict(row, ["code", "label", "sort_order", "is_active"])
    if row is None:
        row = TechSubtype(code=code)
        db.add(row)
    row.code = code
    row.label = label
    row.sort_order = int(payload.sort_order)
    row.is_active = bool(payload.is_active)
    db.flush()

    _audit_log(
        db,
        actor=current_user,
        action="workflow.tech_subtype.upsert",
        target_type="tech_subtype",
        target_key=row.code,
        before=before,
        after=_as_dict(row, ["code", "label", "sort_order", "is_active"]),
    )
    db.commit()
    return {
        "ok": True,
        "data": {
            "code": row.code,
            "label": row.label,
            "sort_order": row.sort_order,
            "is_active": bool(row.is_active),
        },
    }


@router.get("/review-results")
def list_review_results(
    db: Session = Depends(get_db),
):
    rows = db.query(ReviewResult).order_by(ReviewResult.sort_order.asc(), ReviewResult.code.asc()).all()
    return {
        "ok": True,
        "data": [
            {
                "code": row.code,
                "label": row.label,
                "sort_order": row.sort_order,
                "is_active": bool(row.is_active),
            }
            for row in rows
        ],
    }


@router.post("/review-results")
def upsert_review_result(
    payload: ReviewResultIn,
    db: Session = Depends(get_db),
    current_user: DbUser = Depends(require_permission("settings:update")),
):
    code = _upper(payload.code)
    label = _norm(payload.label)
    if not code:
        raise HTTPException(status_code=400, detail="code is required")
    if not label:
        raise HTTPException(status_code=400, detail="label is required")

    row = db.query(ReviewResult).filter(ReviewResult.code == code).first()
    before = _as_dict(row, ["code", "label", "sort_order", "is_active"])
    if row is None:
        row = ReviewResult(code=code)
        db.add(row)
    row.code = code
    row.label = label
    row.sort_order = int(payload.sort_order)
    row.is_active = bool(payload.is_active)
    db.flush()

    _audit_log(
        db,
        actor=current_user,
        action="workflow.review_result.upsert",
        target_type="review_result",
        target_key=row.code,
        before=before,
        after=_as_dict(row, ["code", "label", "sort_order", "is_active"]),
    )
    db.commit()
    return {
        "ok": True,
        "data": {
            "code": row.code,
            "label": row.label,
            "sort_order": row.sort_order,
            "is_active": bool(row.is_active),
        },
    }

