from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


TASK_TABLE_SHEET = "Task_Table1"

_HEADER_ALIASES: dict[str, str] = {
    "name": "subject",
    "subject": "subject",
    "duration": "duration",
    "startdate": "start_date",
    "finishdate": "finish_date",
    "duedate": "finish_date",
    "predecessors": "predecessors",
    "resourcenames": "resource_names",
    "type": "type",
    "priority": "priority",
    "wbs": "wbs",
    "wbscode": "wbs",
    "%complete": "done_ratio",
    "done": "done_ratio",
    "doneratio": "done_ratio",
    "percentcomplete": "done_ratio",
    "complete": "done_ratio",
}

_DATE_FORMATS: tuple[str, ...] = (
    "%d %B %Y %I:%M %p",
    "%d %B %Y %H:%M",
    "%d %B %Y",
    "%d %b %Y",
    "%a %d/%m/%y",
    "%d/%m/%y",
    "%d/%m/%Y",
    "%Y-%m-%d",
)

_PREDECESSOR_TOKEN_RE = re.compile(
    r"^(?P<ref>[A-Za-z0-9][A-Za-z0-9.\-_]*)(?:(?P<rel>[A-Za-z]{2})(?P<lag>[+-]\d+)?)?$",
    flags=re.IGNORECASE,
)
_WBS_RE = re.compile(r"^\d+(?:\.\d+)*$")


def _norm_text(value: Any) -> str:
    return str(value or "").strip()


def _header_key(value: Any) -> str:
    raw = _norm_text(value).lower()
    if not raw:
        return ""
    return re.sub(r"[\s_\-]+", "", raw)


def _normalize_excel_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date().isoformat()
            if isinstance(parsed, date):
                return parsed.isoformat()
        except Exception:
            return None
    raw = _norm_text(value)
    if not raw:
        return None

    normalized = (
        raw.replace("ق.ظ", "AM")
        .replace("ب.ظ", "PM")
        .replace("\u200c", " ")
        .replace("،", " ")
    )
    normalized = re.sub(r"\s+", " ", normalized).strip()

    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(normalized, fmt).date().isoformat()
        except Exception:
            continue

    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", normalized)
    if m:
        value_only = m.group(1)
        for fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value_only, fmt).date().isoformat()
            except Exception:
                continue

    m = re.search(r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})", normalized)
    if m:
        value_only = m.group(1)
        for fmt in ("%d %B %Y", "%d %b %Y"):
            try:
                return datetime.strptime(value_only, fmt).date().isoformat()
            except Exception:
                continue

    return None


def _parse_duration_days(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        days = int(value)
        return days if days >= 0 else None
    raw = _norm_text(value)
    if not raw:
        return None
    match = re.search(r"(-?\d+)", raw)
    if not match:
        return None
    days = int(match.group(1))
    return days if days >= 0 else None


def _parse_done_ratio(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        ratio = float(value)
        if ratio < 0 or ratio > 100:
            return None
        return ratio
    raw = _norm_text(value)
    if not raw:
        return None
    cleaned = raw.replace("%", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    ratio = float(match.group(0))
    if ratio < 0 or ratio > 100:
        return None
    return ratio


def _parse_wbs_segments(value: str) -> list[int] | None:
    raw = _norm_text(value)
    if not raw or not _WBS_RE.fullmatch(raw):
        return None
    try:
        return [int(part) for part in raw.split(".")]
    except Exception:
        return None


def _split_multi_values(raw_value: Any) -> list[str]:
    raw = _norm_text(raw_value)
    if not raw:
        return []
    values: list[str] = []
    for part in re.split(r"[;,]", raw):
        value = _norm_text(part)
        if not value:
            continue
        if value not in values:
            values.append(value)
    return values


def _parse_predecessor_tokens(raw_value: str) -> list[dict[str, Any]]:
    tokens: list[dict[str, Any]] = []
    for token in _split_multi_values(raw_value):
        match = _PREDECESSOR_TOKEN_RE.fullmatch(token)
        if not match:
            tokens.append(
                {
                    "raw": token,
                    "ref": None,
                    "relation_type": None,
                    "lag_days": None,
                    "valid": False,
                    "error": "Token format is invalid.",
                }
            )
            continue

        relation = _norm_text(match.group("rel")).upper() or "FS"
        if relation != "FS":
            tokens.append(
                {
                    "raw": token,
                    "ref": _norm_text(match.group("ref")),
                    "relation_type": relation,
                    "lag_days": None,
                    "valid": False,
                    "error": "Only FS relation is supported in this phase.",
                }
            )
            continue

        lag_text = _norm_text(match.group("lag"))
        lag_days = int(lag_text) if lag_text else 0
        tokens.append(
            {
                "raw": token,
                "ref": _norm_text(match.group("ref")),
                "relation_type": "FS",
                "lag_days": int(lag_days),
                "valid": True,
                "error": None,
            }
        )
    return tokens


def parse_task_table_from_bytes(
    file_bytes: bytes,
    *,
    source_file_name: str,
) -> dict[str, Any]:
    if not file_bytes:
        raise ValueError("Uploaded file is empty.")

    digest = hashlib.sha256(file_bytes).hexdigest()
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid Excel file: {exc}") from exc

    if TASK_TABLE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet `{TASK_TABLE_SHEET}` not found in workbook.")

    sheet = workbook[TASK_TABLE_SHEET]
    rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Task sheet header is empty.") from exc

    header_labels = [_norm_text(cell_value) for cell_value in header_row]
    headers: dict[str, int] = {}
    alias_by_index: dict[int, str] = {}
    for idx, header_label in enumerate(header_labels):
        alias = _HEADER_ALIASES.get(_header_key(header_label))
        if alias and alias not in headers:
            headers[alias] = idx
        if alias:
            alias_by_index[idx] = alias

    if "subject" not in headers:
        raise ValueError("Task sheet must contain `Name` or `Subject` column.")

    has_wbs_column = "wbs" in headers

    parsed_rows: list[dict[str, Any]] = []
    total_rows = 0
    valid_rows = 0
    invalid_rows = 0
    mapping_warnings: list[str] = []
    if not has_wbs_column:
        mapping_warnings.append("WBS column not found; sequential WBS values were auto-generated.")
    seen_wbs: set[str] = set()

    for excel_row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(not _norm_text(v) for v in row):
            continue

        total_rows += 1
        row_index = int(total_rows)

        def value_of(field: str) -> Any:
            pos = headers.get(field)
            if pos is None or pos >= len(row):
                return None
            return row[pos]

        subject = _norm_text(value_of("subject"))
        type_text = _norm_text(value_of("type"))
        priority_text = _norm_text(value_of("priority"))
        duration_raw = _norm_text(value_of("duration"))
        start_raw = _norm_text(value_of("start_date"))
        finish_raw = _norm_text(value_of("finish_date"))
        predecessors_raw = _norm_text(value_of("predecessors"))
        resource_names_raw = _norm_text(value_of("resource_names"))
        done_ratio_raw = _norm_text(value_of("done_ratio"))

        wbs_source = "provided" if has_wbs_column else "auto_generated"
        if has_wbs_column:
            wbs_code = _norm_text(value_of("wbs"))
        else:
            wbs_code = str(row_index)

        normalized_start = _normalize_excel_date(value_of("start_date"))
        normalized_finish = _normalize_excel_date(value_of("finish_date"))
        duration_days = _parse_duration_days(value_of("duration"))
        done_ratio = _parse_done_ratio(value_of("done_ratio"))

        predecessor_tokens = _parse_predecessor_tokens(predecessors_raw)
        resource_items = _split_multi_values(resource_names_raw)

        excel_raw: dict[str, Any] = {}
        custom_fields: dict[str, Any] = {}
        for idx, header_label in enumerate(header_labels):
            label = _norm_text(header_label)
            if not label:
                continue
            value = row[idx] if idx < len(row) else None
            text = _norm_text(value)
            if text:
                excel_raw[label] = text
            if alias_by_index.get(idx):
                continue
            if text:
                custom_fields[label] = text

        row_errors: list[str] = []
        if not subject:
            row_errors.append("Subject is required.")
        if has_wbs_column and not wbs_code:
            row_errors.append("WBS is required.")

        wbs_segments = _parse_wbs_segments(wbs_code)
        if not wbs_segments:
            row_errors.append("WBS format is invalid.")
        elif wbs_code in seen_wbs:
            row_errors.append("WBS must be unique in file.")
        else:
            seen_wbs.add(wbs_code)

        if start_raw and not normalized_start:
            row_errors.append("Start_Date is invalid.")
        if finish_raw and not normalized_finish:
            row_errors.append("Finish_Date is invalid.")
        if normalized_start and normalized_finish and normalized_finish < normalized_start:
            row_errors.append("Finish_Date must be greater than or equal to Start_Date.")

        if done_ratio_raw and done_ratio is None:
            row_errors.append("%complete must be a number between 0 and 100.")

        invalid_pred_tokens = [token for token in predecessor_tokens if not bool(token.get("valid"))]
        if invalid_pred_tokens:
            row_errors.append("Predecessors format is invalid. Only FS and optional lag are supported.")

        parent_wbs = ".".join(str(part) for part in (wbs_segments or [])[:-1]) if wbs_segments and len(wbs_segments) > 1 else None

        mapped_fields = {
            "row_index": int(row_index),
            "subject": subject or None,
            "type_text": type_text or None,
            "priority_text": priority_text or None,
            "duration_days": duration_days,
            "start_date": normalized_start,
            "due_date": normalized_finish,
            "done_ratio": done_ratio,
            "wbs_code": wbs_code or None,
            "predecessor_raw": predecessors_raw or None,
            "resource_names_raw": resource_names_raw or None,
            "resources": resource_items,
        }

        payload = {
            "sheet": TASK_TABLE_SHEET,
            "excel_raw": excel_raw,
            "mapped_fields": mapped_fields,
            "custom_fields": custom_fields,
            "wbs_meta": {
                "source": wbs_source,
                "wbs_code": wbs_code or None,
                "level": int(len(wbs_segments or [])),
                "parent_wbs": parent_wbs,
            },
            "predecessor_tokens": predecessor_tokens,
            "execution_meta": {
                "pass1": {},
                "relations": [],
            },
            "errors": list(row_errors),
        }

        validation_status = "VALID" if not row_errors else "INVALID"
        if validation_status == "VALID":
            valid_rows += 1
        else:
            invalid_rows += 1

        parsed_rows.append(
            {
                "row_no": int(excel_row_no),
                "task_name": subject or None,
                "duration_raw": duration_raw or None,
                "start_raw": start_raw or None,
                "finish_raw": finish_raw or None,
                "predecessors_raw": predecessors_raw or None,
                "resource_names_raw": resource_names_raw or None,
                "normalized_start_date": normalized_start,
                "normalized_finish_date": normalized_finish,
                "validation_status": validation_status,
                "execution_status": "PENDING",
                "error_message": "; ".join(row_errors) if row_errors else None,
                "payload_json": json.dumps(payload, ensure_ascii=False),
            }
        )

    summary = {
        "run_type": "excel_import",
        "sheet": TASK_TABLE_SHEET,
        "source_file_name": _norm_text(source_file_name) or "openproject_import.xlsx",
        "source_sha256": digest,
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "created_rows": 0,
        "failed_rows": 0,
        "pass1_created_rows": 0,
        "pass1_failed_rows": 0,
        "pass2_relation_created": 0,
        "pass2_relation_failed": 0,
        "mapping_warnings": mapping_warnings,
    }
    return {
        "summary": summary,
        "rows": parsed_rows,
        "source_sha256": digest,
    }


def parse_summary_json(value: str | None) -> dict[str, Any]:
    raw = _norm_text(value)
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def build_work_package_create_payload(
    *,
    row_task_name: str,
    row_start_date: str | None,
    row_finish_date: str | None,
    row_done_ratio: float | None,
    row_no: int,
    run_no: str,
    parent_work_package_id: int,
    project_href: str,
    type_href: str,
    priority_href: str | None = None,
    done_ratio_field: str = "doneRatio",
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "subject": row_task_name,
        "description": {
            "format": "markdown",
            "raw": f"Imported from Excel run {run_no} (row {row_no}).",
        },
        "_links": {
            "project": {"href": project_href},
            "type": {"href": type_href},
            "parent": {"href": f"/api/v3/work_packages/{int(parent_work_package_id)}"},
        },
    }
    if priority_href:
        payload["_links"]["priority"] = {"href": str(priority_href)}
    if row_start_date:
        payload["startDate"] = row_start_date
    if row_finish_date:
        payload["dueDate"] = row_finish_date
    if row_done_ratio is not None:
        payload[str(done_ratio_field or "doneRatio")] = float(row_done_ratio)
    return payload
