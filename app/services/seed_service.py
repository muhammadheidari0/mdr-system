# app/services/seed_service.py
from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.config import settings
from app.db.models import (
    Project, Phase, Discipline, Package, Level,
    Block, MdrCategory, DocStatus, SettingsKV
)

# -----------------------------
# Helpers
# -----------------------------
def _norm(s: Any) -> str:
    if s is None: return ""
    return str(s).strip()

def _upper(s: Any) -> str:
    return _norm(s).upper()

def _safe_int(val: Any, default: int = 0) -> int:
    try:
        if val is None: return default
        return int(float(str(val).strip() or 0))
    except (ValueError, TypeError):
        return default

def _read_excel_sheet(file_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    """خواندن یک شیت اکسل و تبدیل به لیست دیکشنری"""
    if not file_path.exists():
        print(f"[WARN] Excel file not found: {file_path}")
        return []
    
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        
        ws = wb[sheet_name]
        rows = list(ws.rows)
        if not rows:
            return []

        # ردیف اول هدر است
        headers = [cell.value for cell in rows[0]]
        data = []
        for row in rows[1:]:
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    key = str(headers[i]).strip().lower()
                    row_data[key] = cell.value
            
            if any(row_data.values()):
                data.append(row_data)
        
        return data
    except Exception as e:
        print(f"[ERR] Error reading excel sheet '{sheet_name}': {e}")
        return []

# -----------------------------
# Main Seeding Logic
# -----------------------------
def seed_from_excel(db: Session) -> Dict[str, Any]:
    """
    خواندن فایل مستر دیتا و آپدیت جداول پایه
    """
    # مسیر فایل اکسل (نسبت به BASE_DIR در config)
    excel_path = settings.BASE_DIR / "data_sources" / "master_data.xlsx"
    
    if not excel_path.exists():
        return {"error": f"File not found: {excel_path}"}

    stats = {}

    # 1. Projects
    rows = _read_excel_sheet(excel_path, "Projects")
    count = 0
    for r in rows:
        code = _upper(r.get("code"))
        if not code: continue
        proj = db.query(Project).filter(Project.code == code).first()
        ne, np, root = _norm(r.get("name_e")), _norm(r.get("name_p")), _norm(r.get("root_path"))
        
        if not proj:
            db.add(Project(code=code, name_e=ne, name_p=np, root_path=root, is_active=True))
            count += 1
        else:
            proj.name_e = ne; proj.name_p = np
            if root: proj.root_path = root
    stats["projects"] = count

    # 2. MDR Categories
    rows = _read_excel_sheet(excel_path, "MdrCategories")
    count = 0
    for r in rows:
        code = _upper(r.get("code"))
        if not code: continue
        mdr = db.query(MdrCategory).filter(MdrCategory.code == code).first()
        ne, np, fld = _norm(r.get("name_e")), _norm(r.get("name_p")), _norm(r.get("folder_name"))
        so = _safe_int(r.get("sort_order"))
        
        if not mdr:
            db.add(MdrCategory(code=code, name_e=ne, name_p=np, folder_name=fld, sort_order=so, is_active=True))
            count += 1
        else:
            mdr.name_e=ne; mdr.name_p=np; mdr.folder_name=fld; mdr.sort_order=so
    stats["mdr_categories"] = count

    # 3. Phases
    rows = _read_excel_sheet(excel_path, "Phases")
    count = 0
    for r in rows:
        code = _upper(r.get("ph_code"))
        if not code: continue
        ph = db.query(Phase).filter(Phase.ph_code == code).first()
        ne, np = _norm(r.get("name_e")), _norm(r.get("name_p"))
        
        if not ph:
            db.add(Phase(ph_code=code, name_e=ne, name_p=np))
            count += 1
        else:
            ph.name_e=ne; ph.name_p=np
    stats["phases"] = count

    # 4. Levels
    rows = _read_excel_sheet(excel_path, "Levels")
    count = 0
    for r in rows:
        code = _norm(r.get("code"))
        if not code: continue
        l = db.query(Level).filter(Level.code == code).first()
        ne, np = _norm(r.get("name_e")), _norm(r.get("name_p"))
        so = _safe_int(r.get("sort_order"))
        
        if not l:
            db.add(Level(code=code, name_e=ne, name_p=np, sort_order=so))
            count += 1
        else:
            l.name_e=ne; l.name_p=np; l.sort_order=so
    stats["levels"] = count

    # 5. Blocks
    rows = _read_excel_sheet(excel_path, "Blocks")
    count = 0
    for r in rows:
        pc, bc = _upper(r.get("project_code")), _upper(r.get("code"))
        if not pc or not bc: continue
        # Project must exist
        if not db.query(Project).filter(Project.code == pc).first(): continue

        blk = db.query(Block).filter(Block.project_code == pc, Block.code == bc).first()
        ne, np = _norm(r.get("name_e")), _norm(r.get("name_p"))
        so = _safe_int(r.get("sort_order"))
        
        if not blk:
            db.add(Block(project_code=pc, code=bc, name_e=ne, name_p=np, sort_order=so, is_active=True))
            count += 1
        else:
            blk.name_e=ne; blk.name_p=np; blk.sort_order=so
    stats["blocks"] = count

    # 6. Disciplines
    rows = _read_excel_sheet(excel_path, "Disciplines")
    count = 0
    for r in rows:
        code = _upper(r.get("code"))
        if not code: continue
        d = db.query(Discipline).filter(Discipline.code == code).first()
        ne, np = _norm(r.get("name_e")), _norm(r.get("name_p"))
        
        if not d:
            db.add(Discipline(code=code, name_e=ne, name_p=np))
            count += 1
        else:
            d.name_e=ne; d.name_p=np
    stats["disciplines"] = count
    
    db.flush()

    # 7. Packages
    rows = _read_excel_sheet(excel_path, "Packages")
    count = 0
    for r in rows:
        dc, pc = _upper(r.get("discipline_code")), _upper(r.get("package_code"))
        if not dc or not pc: continue
        if not db.query(Discipline).filter(Discipline.code == dc).first(): continue
        
        pkg = db.query(Package).filter(Package.discipline_code == dc, Package.package_code == pc).first()
        ne, np = _norm(r.get("name_e")), _norm(r.get("name_p"))
        
        if not pkg:
            db.add(Package(discipline_code=dc, package_code=pc, name_e=ne, name_p=np))
            count += 1
        else:
            pkg.name_e=ne; pkg.name_p=np
    stats["packages"] = count

    # 8. Document Statuses
    rows = _read_excel_sheet(excel_path, "Statuses")
    count = 0
    for r in rows:
        code = _upper(r.get("code"))
        if not code: continue
        
        st = db.query(DocStatus).filter(DocStatus.code == code).first()
        name = _norm(r.get("name"))
        desc = _norm(r.get("description"))
        so = _safe_int(r.get("order")) or _safe_int(r.get("sort_order"))
        
        if not st:
            db.add(DocStatus(code=code, name=name, description=desc, sort_order=so))
            count += 1
        else:
            st.name = name
            st.description = desc
            st.sort_order = so
    stats["statuses"] = count

    return stats